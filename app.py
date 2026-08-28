import streamlit as st

# Force calendar header visible to fix caching issues
st.markdown('''
    <style>
        div[data-baseweb="calendar"] header {
            visibility: visible !important;
            display: flex !important;
        }
    </style>
''', unsafe_allow_html=True)

from datetime import date
import kpi_reports
import pandas as pd
import os
import re
import json
import plotly.express as px
import sqlite3

# Persistent Settings
SETTINGS_FILE = 'local_settings.json'

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_settings(settings):
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=4)

try:
    import google.generativeai as genai
except ImportError:
    st.error("Thư viện google-generativeai chưa được cài đặt. Vui lòng kiểm tra file requirements.txt.")

from datetime import datetime, date, timedelta

def calculate_time_progress(start_date, deadline_date, is_completed=False):
    """Tính % sức khỏe thời gian: Đang tốt (99), Sắp tới hạn (50), Trễ hạn (0)"""
    if is_completed:
        return 100.0
    try:
        today = date.today()
        if isinstance(deadline_date, str):
            deadline_date = datetime.strptime(deadline_date, "%Y-%m-%d").date()
            
        if not deadline_date:
            return 0.0
            
        days_left = (deadline_date - today).days
        
        if days_left < 0:
            return 0.0  # Trễ hạn
        elif 0 <= days_left <= 3:
            return 50.0  # Sắp tới hạn
        else:
            return 99.0  # Còn nhiều hạn
    except Exception:
        return 0.0

from contextlib import contextmanager
import time

@contextmanager
def acquire_db_lock(timeout=15):
    # Lock local (cho cùng 1 container/server)
    lock_dir = "db_write.lock"
    start_time = time.time()
    locked_local = False
    
    while time.time() - start_time < timeout:
        try:
            os.mkdir(lock_dir)
            locked_local = True
            break
        except FileExistsError:
            time.sleep(0.5)
            
    if not locked_local:
        st.error("⚠️ Hệ thống đang bận. Vui lòng đợi vài giây và thử lại.")
        st.stop()
        
    try:
        yield
    finally:
        try:
            os.rmdir(lock_dir)
        except:
            pass

# Page config - Light Theme is handled natively by Streamlit's default settings
st.set_page_config(
    page_title="Hệ thống Quản lý Tiến độ Công việc & KPI - DMT Group",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Chống dịch tự động của Google (gây lỗi chính tả) và chuẩn hóa Font chữ tiếng Việt
st.markdown("""
    <style>
        /* Cố định Font chuẩn hỗ trợ đầy đủ tiếng Việt và tăng kích thước chữ an toàn (không ghi đè icon) */
        html, body, p, label, div.stMarkdown, div.stText {
            font-family: 'Segoe UI', 'Roboto', 'Arial', sans-serif;
            font-size: 1.08rem;
        }
        /* Ngăn Google Translate tự động dịch làm hỏng văn bản tiếng Việt */
        html {
            translate: no;
        }
    </style>
""", unsafe_allow_html=True)

import streamlit.components.v1 as components
components.html(
    """
    <script>
        // Set ngôn ngữ trang thành tiếng Việt và gắn thẻ meta chống dịch
        
        const meta = window.parent.document.createElement('meta');
        meta.name = 'google';
        meta.content = 'notranslate';
        window.parent.document.getElementsByTagName('head')[0].appendChild(meta);
    </script>
    """,
    height=0,
    width=0,
)

# Standardized companies based on CIENCO, DMT Group, and DMT Marina documents
COMPANIES = {
    "CTY CP ĐẦU TƯ ĐÀ NẴNG - MIỀN TRUNG": {},
    "CTY CP XÂY DỰNG CÔNG TRÌNH GIAO THÔNG ĐN-MT": {},
    "CTY CP DMT - MARINA (Du thuyền Happy Yacht)": {}
}

# Configuration JSON logic for dynamic Projects and Departments
CONFIG_FILE = os.path.join("OUTPUT", "CONFIG_PROJECTS.json")

DEFAULT_PERSONNEL = {
    "Ban Lãnh đạo": ["Trần Quốc Thể", "Đoàn Thị Ngọc Nữ", "Đặng Ngọc Hoàng"],
    "Ban Hành chính Nhân sự": ["Nguyễn Thị Hạnh Tiên", "Nguyễn Băng Trinh", "Lê Ngọc Tú Uyên"],
    "Ban Tài chính Kế toán": ["Đồng Thị Nguyệt Nga", "Huỳnh Thị Hoàng Hà", "Nguyễn Thị Nhật Sang"],
    "Ban Kế hoạch Đầu tư": ["Nguyễn Trần Thức", "Phan Thị Mỹ Hạnh", "Nguyễn Đức Lợi", "Trần Tin"],
    "Ban Chuẩn bị Đầu tư": ["Hồ Văn Khoa", "Phan Thị Mỹ Hạnh", "Cao Thuỷ Tiên"],
    "Ban Kỹ thuật": ["Nguyễn Văn Bồn"],
    "Ban Đền bù Giải tỏa": ["Nguyễn Ngọc Tôn", "Đặng Công Nhựt", "Đặng Thị Mỹ Hạnh", "Đặng Thanh Quang"],
    "Ban chỉ huy Công trường": ["Nguyễn Phong Trung", "Phạm Văn Long", "Lê Đông"],
    "Xí nghiệp xe máy thiết bị": ["Đặng Hiền"],
    "Ban Dự án": ["Nguyễn Đình Thắng", "Nguyễn Đình Hiếu"],
    "Xí nghiệp DTBD": ["Mai Văn Châu"],
    "Sàn GDBĐS": ["Ngô Thị Tâm"],
    "Tổ KPI": []
}

def is_gsheets_configured():
    try:
        # Check if connections.gsheets exists in secrets
        if "connections" in st.secrets and "gsheets" in st.secrets["connections"]:
            return True
    except Exception:
        pass
    return False

def get_gsheets_conn():
    from supabase import create_client
    SUPABASE_URL = 'https://xlfnxyerpcebqxgmfngd.supabase.co'
    SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InhsZm54eWVycGNlYnF4Z21mbmdkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NjYwNTAzNSwiZXhwIjoyMTAyMTgxMDM1fQ.qZsoZu8HaFpbvsG6siw76M5QXmX5bwipLV1qWeGG89s'
    try:
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception as e:
        import streamlit as st


        st.warning(f"Chưa cấu hình Supabase Connection: {e}")
        return None

def _get_table_name(worksheet):
    mapping = {
        "Sheet1": "tasks",
        "GANTT_KHDT": "gantt_tasks",
        "VAN_BAN_DEN": "documents",
        "CONFIG": "kpi_config",
        "KPI_ADJUSTMENTS": "kpi_adjustments"
    }
    return mapping.get(worksheet, worksheet)

@st.cache_resource
def get_global_state():
    return {}

def safe_gsheets_read(conn, worksheet, ttl=15, fallback_df=None):
    if fallback_df is None:
        import pandas as pd
        fallback_df = pd.DataFrame()
    
    import streamlit as st


    import time
    
    cache_key = f"cached_df_{worksheet}"
    

            
    try:
        table_name = _get_table_name(worksheet)
        res = conn.table(table_name).select('*').execute()
        data = res.data
        if not data:
            return fallback_df
            
        import pandas as pd
        df = pd.DataFrame(data)
        
        import numpy as np
        df = df.replace("", np.nan).dropna(how='all')
        
        if worksheet == "Sheet1":
            col_mapping = {
                'ID': 'Mã CV',
                'TenCongViec': 'Tên công việc',
                'PhanTramHoanThanh': 'Tiến độ %',
                'TrangThai': 'Trạng thái',
                'Deadline': 'Hạn chót'
            }
            df.rename(columns=col_mapping, inplace=True)
            if "NgayBatDau" in df.columns:
                df["NgayBatDau"] = pd.to_datetime(df["NgayBatDau"], errors="coerce").dt.strftime('%d/%m/%Y')
            if "Hạn chót" in df.columns:
                df["Hạn chót"] = pd.to_datetime(df["Hạn chót"], errors="coerce").dt.strftime('%d/%m/%Y')
                
        elif worksheet == "KPI_ADJUSTMENTS":
            if "SoDiem" in df.columns:
                df.rename(columns={"SoDiem": "DiemDieuChinh"}, inplace=True)
            if "NhanSu" in df.columns:
                df.rename(columns={"NhanSu": "TenNhanVien"}, inplace=True)
            if "LoaiDieuChinh" in df.columns:
                df.rename(columns={"LoaiDieuChinh": "LoaiHanhVi"}, inplace=True)
                

        
        return df
    except Exception as e:
        import streamlit as st


        st.warning(f"Lỗi đọc Supabase ({worksheet}): {e}")
        return fallback_df

def safe_gsheets_update(conn, worksheet, data):
    import streamlit as st


    import time
    import pandas as pd
    import numpy as np
    import math
    
    cache_key = f"cached_df_{worksheet}"
    table_name = _get_table_name(worksheet)
    
    try:
        df = data.copy()
        if worksheet == "Sheet1":
            col_mapping = {
                'Mã CV': 'ID', 'MaCV': 'ID',
                'Tên công việc': 'TenCongViec', 'Nội dung': 'TenCongViec', 'Công việc': 'TenCongViec',
                'Tiến độ %': 'PhanTramHoanThanh', 'Progress': 'PhanTramHoanThanh', 'Tiến độ': 'PhanTramHoanThanh',
                'Trạng thái': 'TrangThai', 'Status': 'TrangThai',
                'Hạn chót': 'Deadline', 'Ngày hoàn thành': 'Deadline'
            }
            rename_dict = {k: v for k, v in col_mapping.items() if k in df.columns}
            df.rename(columns=rename_dict, inplace=True)
            
            
            allowed_cols = ['ID', 'DonVi', 'PhongBan', 'NguoiChuTri', 'TenDuAn', 'MocTienDo', 'SanPhamBanGiao', 'TenCongViec', 'PhanLoaiChiSo', 'NgayBatDau', 'Deadline', 'DoUuTien', 'PhanTramHoanThanh', 'TrangThai', 'LinkKetQua', 'GiaiTrinhDeXuat', 'NgayCapNhat', 'ChuKyTheoDoi', 'PhanLoaiTreHan', 'TyTrongKPI', 'NguonGiaoViec', 'MucDoGhiNhan']
            df = df[[c for c in df.columns if c in allowed_cols]]
            
        elif worksheet == "KPI_ADJUSTMENTS":
            if "DiemDieuChinh" in df.columns:
                df.rename(columns={"DiemDieuChinh": "SoDiem"}, inplace=True)
            if "TenNhanVien" in df.columns:
                df.rename(columns={"TenNhanVien": "NhanSu"}, inplace=True)
            if "LoaiHanhVi" in df.columns:
                df.rename(columns={"LoaiHanhVi": "LoaiDieuChinh"}, inplace=True)
            allowed_cols = ['ID', 'NhanSu', 'Thang', 'Nam', 'LoaiDieuChinh', 'SoDiem', 'LyDo', 'NguoiCapNhat', 'ThoiGianCapNhat']
            df = df[[c for c in df.columns if c in allowed_cols]]
            
        elif worksheet == "CONFIG":
            allowed_cols = ['PhongBan', 'NhanSu', 'Role', 'config_json']
            df = df[[c for c in df.columns if c in allowed_cols]]
            
        elif worksheet == "GANTT_KHDT":
            allowed_cols = ['ID', 'TenDuAn', 'TenCongViec', 'GiaiDoan', 'NgayBatDau', 'Deadline', 'PhanTramHoanThanh', 'Milestone', 'NgayCapNhat']
            df = df[[c for c in df.columns if c in allowed_cols]]
            
        elif worksheet == "VAN_BAN_DEN":
            allowed_cols = ['ID', 'SoKyHieu', 'NgayBanHanh', 'CoQuanBanHanh', 'TrichYeu', 'NguoiChuTri', 'ThoiHanGiaiQuyet', 'TrangThai', 'GhiChu']
            df = df[[c for c in df.columns if c in allowed_cols]]

        pk = 'NhanSu' if worksheet == 'CONFIG' else 'ID'
        if pk in df.columns:
            df = df[df[pk].notna()]
            df = df[df[pk] != '']

        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            
        records = df.to_dict(orient='records')
        for r in records:
            for k, v in list(r.items()):
                if isinstance(v, float):
                    if math.isnan(v):
                        r[k] = None
                    elif v.is_integer():
                        r[k] = int(v)
                elif pd.isna(v):
                    r[k] = None
                    
        if records:
            conn.table(table_name).upsert(records).execute()
            
            if pk in df.columns:
                current_ids = [str(x) for x in df[pk].tolist()]
                if len(current_ids) > 0:
                    res = conn.table(table_name).select(pk).execute()
                    db_ids = [(row[pk], str(row[pk])) for row in res.data]
                    ids_to_delete = [orig for orig, string_val in db_ids if string_val not in current_ids]
                    if ids_to_delete:
                        for i in range(0, len(ids_to_delete), 100):
                            batch_del = ids_to_delete[i:i+100]
                            conn.table(table_name).delete().in_(pk, batch_del).execute()

        st.session_state[cache_key] = data.copy()
        st.session_state[cache_key + "_time"] = time.time()
        
        global_state = get_global_state()
        global_state[cache_key] = data.copy()
        global_state[cache_key + "_time"] = time.time()
        
        if worksheet == "Sheet1":
            if hasattr(read_db, "clear"): read_db.clear()
        elif worksheet == "GANTT_KHDT":
            if hasattr(read_gantt_db, "clear"): read_gantt_db.clear()
        elif worksheet == "KPI_ADJUSTMENTS":
            if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
        elif worksheet == "VAN_BAN_DEN":
            if hasattr(read_incoming_docs_db, "clear"): read_incoming_docs_db.clear()
            
        return True
    except Exception as e:
        err_msg = str(e)
        import streamlit as st


        st.error(f"Lỗi khi lưu vào Supabase ({worksheet}): {err_msg}")
        return False


def save_config(config_data):
    conn = get_gsheets_conn()
    if conn is None:
        st.error("Chưa cấu hình Google Sheets (secrets.toml).")
        return False
    try:
        import json
        import pandas as pd
        
        config_data_copy = config_data.copy()
        job_descriptions = config_data_copy.pop("job_descriptions", {})
        
        rows = []
        rows.append({
            "NhanSu": "APP_GLOBAL_CONFIG",
            "PhongBan": "SYSTEM",
            "Role": "SYSTEM",
            "config_json": json.dumps(config_data_copy, ensure_ascii=False)
        })
        
        for company, personnel_dict in job_descriptions.items():
            for person, jd_data in personnel_dict.items():
                if jd_data:
                    rows.append({
                        "NhanSu": person,
                        "PhongBan": company,
                        "Role": "JD",
                        "config_json": json.dumps(jd_data, ensure_ascii=False) if isinstance(jd_data, (dict, list)) else json.dumps({"jd_text": jd_data}, ensure_ascii=False)
                    })
                    
        df_save = pd.DataFrame(rows)
        success = safe_gsheets_update(conn, worksheet="CONFIG", data=df_save)
        if not success:
            st.error("⚠️ Lỗi: Không tìm thấy trang tính 'CONFIG' trên Google Sheets! Vui lòng mở Google Sheets, tạo một Sheet mới đặt tên là 'CONFIG', sau đó lưu lại.")
            return False
            
        st.cache_data.clear()
        config_data["job_descriptions"] = job_descriptions
        return True
    except Exception as e:
        st.error(f'Lỗi lưu Google Sheets: {e}')
        return False

def load_config():
    default_config = {
        "companies": {
            "CTY CP ĐẦU TƯ ĐÀ NẴNG - MIỀN TRUNG": {
                "projects_by_category": {
                    "BĐS & KDC": ["KDC Bàu Mạc", "KDC Nam Bàu Mạc", "KĐT Phước Lý & Phước Lý MR", "TĐC Phước Lý 2 & Hoà Liên 5", "Dự án Phong Nam", "Khu BT ST Hoà Ninh"],
                    "HẠ TẦNG & GIAO THÔNG": ["Tuyến đường Lê Trọng Tấn", "Tuyến đường Lê Trọng Tấn - Hoà Nhơn", "Tuyến đường Trần Hưng Đạo (BT)", "Trục I Tây Bắc", "Khu TĐC Hoà Vang"],
                    "THƯƠNG MẠI & KHÁCH SẠN": ["Khách sạn DMT-Group"]
                },
                "departments": ["Ban Lãnh đạo", "Ban Hành chính Nhân sự", "Ban Tài chính Kế toán", "Ban Kế hoạch Đầu tư", "Ban Chuẩn bị Đầu tư", "Ban Kỹ thuật", "Ban Đền bù Giải tỏa", "Tổ KPI", "Ban chỉ huy Công trường", "Xí nghiệp xe máy thiết bị", "Ban Dự án", "Xí nghiệp DTBD", "Sàn GDBĐS"],
                "personnel_by_department": DEFAULT_PERSONNEL.copy()
            },
            "CTY CP DMT - MARINA (Du thuyền Happy Yacht)": {
                "projects_by_category": {
                    "THƯƠNG MẠI & KHÁCH SẠN": ["Du thuyền Happy Yacht (DMT Marina)", "Du thuyền Happy Yacht", "HCNS", "TCKT"]
                },
                "departments": ["Ban Lãnh đạo", "Ban Hành chính Nhân sự", "Ban Tài chính Kế toán"],
                "personnel_by_department": {
                    "Ban Hành chính Nhân sự": ["Nguyễn Thị Hạnh Tiên"],
                    "Ban Tài chính Kế toán": ["Lê Thị Hải"],
                    "Ban Lãnh đạo": ["Trần Cường", "Đặng Ngọc Hoàng"],
                    "Tổ KPI": []
                }
            },
            "CTY CP XÂY DỰNG CÔNG TRÌNH GIAO THÔNG ĐN-MT": {
                "projects_by_category": {},
                "departments": ["Ban Lãnh đạo", "Ban Kỹ thuật", "Ban chỉ huy Công trường", "Xí nghiệp xe máy thiết bị", "Ban Hành chính Nhân sự", "Ban Tài chính Kế toán"],
                "personnel_by_department": {
                    "Ban Lãnh đạo": ["Thái Văn Thành", "Trần Văn Trọng", "Đặng Thị Lan Ngọc"],
                    "Ban Kỹ thuật": ["Trần Văn Trọng", "Phạm Quang Nghĩa"],
                    "Ban chỉ huy Công trường": ["Nguyễn Phong Trung", "Phạm Văn Long", "Lê Đông"],
                    "Xí nghiệp xe máy thiết bị": ["Đặng Hiền"],
                    "Ban Tài chính Kế toán": ["Nguyễn Thị Ngọc Hà", "Nguyễn Thị Như Can"],
                    "Ban Hành chính Nhân sự": ["Nguyễn Thị Mỹ Phương"]
                }
            }
        },
        "cv_gsheet_url": ""
    }
    
    conn = get_gsheets_conn()
    if conn is None:
        return default_config
        
    try:
        import json
        df = safe_gsheets_read(conn, worksheet="CONFIG", ttl=600)
        if df is None or df.empty:
            return default_config
            
        if "config_json" in df.columns:
            config_rows = df[df["config_json"].notna()]
            if not config_rows.empty:
                if "NhanSu" in df.columns:
                    app_row = config_rows[config_rows["NhanSu"] == "APP_GLOBAL_CONFIG"]
                    if not app_row.empty:
                        json_str = app_row.iloc[0]["config_json"]
                    else:
                        json_str = config_rows.iloc[0]["config_json"]
                else:
                    json_str = config_rows.iloc[0]["config_json"]
            else:
                json_str = df.iloc[0]["config_json"]
        else:
            json_str = df.iloc[0]["config_json"]
            
        data = json.loads(json_str)
        
        if "NhanSu" in df.columns and "Role" in df.columns:
            jd_rows = df[df["Role"] == "JD"]
            if not jd_rows.empty:
                if "job_descriptions" not in data:
                    data["job_descriptions"] = {}
                for _, row in jd_rows.iterrows():
                    company = row.get("PhongBan", "")
                    person = row.get("NhanSu", "")
                    jd_json = row.get("config_json", "{}")
                    try:
                        jd_data = json.loads(jd_json)
                        if company not in data["job_descriptions"]:
                            data["job_descriptions"][company] = {}
                        data["job_descriptions"][company][person] = jd_data
                    except:
                        pass
        
        needs_save = False
        if "personnel_by_department" not in data:
            data["personnel_by_department"] = DEFAULT_PERSONNEL.copy()
            needs_save = True
            
        if "cv_gsheet_url" not in data:
            data["cv_gsheet_url"] = ""
            needs_save = True
            
        if "companies" not in data:
            data["companies"] = default_config["companies"]
            needs_save = True
            
        return data
    except Exception as e:
        print("Error parsing DB config:", e)
        return default_config


# Load current config dynamically
config = load_config()

DEPT_ABBR = {
    "Ban Lãnh đạo": "BLĐ",
    "Lãnh đạo": "BLĐ",
    "Ban Hành chính Nhân sự": "HCNS",
    "Ban Tài chính Kế toán": "TCKT",
    "Ban Kế hoạch Đầu tư": "KHĐT",
    "Ban Chuẩn bị Đầu tư": "CBĐT",
    "Ban Kỹ thuật": "KT",
    "Ban Đền bù Giải tỏa": "ĐBGT",
    "Ban Dự án": "DA",
    "Xí nghiệp DTBD": "XN DTBD",
    "Sàn GDBĐS": "Sàn GDBĐS",
    "Tổ KPI": "Tổ KPI",
    "Ban chỉ huy Công trường": "BCH CT",
    "Xí nghiệp xe máy thiết bị": "XN XMTB",
    "Xí nghiệp xe thiết bị": "XN XMTB"
}

for comp_name, comp_data in config.get("companies", {}).items():
    if "departments" in comp_data:
        comp_data["departments"] = [DEPT_ABBR.get(d, d) for d in comp_data["departments"]]
    if "personnel_by_department" in comp_data:
        new_personnel = {}
        for d, p in comp_data["personnel_by_department"].items():
            new_personnel[DEPT_ABBR.get(d, d)] = p
        comp_data["personnel_by_department"] = new_personnel


# Default owners by department and company for autofill
DEPT_ABBR = {
    "Ban Lãnh đạo": "BLĐ",
    "Ban Hành chính Nhân sự": "HCNS",
    "Ban Tài chính Kế toán": "TCKT",
    "Ban Kế hoạch Đầu tư": "KHĐT",
    "Ban Chuẩn bị Đầu tư": "CBĐT",
    "Ban Kỹ thuật": "KT",
    "Ban Đền bù Giải tỏa": "ĐBGT",
    "Ban Dự án": "DA",
    "Xí nghiệp DTBD": "XN DTBD",
    "Sàn GDBĐS": "Sàn GDBĐS",
    "Tổ KPI": "Tổ KPI"
}

DEPT_LEADS = {
    "CTY CP ĐẦU TƯ ĐÀ NẴNG - MIỀN TRUNG": {
        "BLĐ": "Trần Quốc Thể",
        "HCNS": "Nguyễn Thị Hạnh Tiên",
        "TCKT": "Đồng Thị Nguyệt Nga",
        "KHĐT": "Nguyễn Trần Thức",
        "CBĐT": "Hồ Văn Khoa",
        "KT": "Nguyễn Văn Bồn",
        "ĐBGT": "Nguyễn Ngọc Tôn",
        "DA": "Nguyễn Đình Thắng",
        "XN DTBD": "Mai Văn Châu",
        "Sàn GDBĐS": "Ngô Thị Tâm",
        "Tổ KPI": ""
    }
}

def get_personnel_for_company_dept(company, dept, config):
    companies = config.get("companies", {})
    if company in companies:
        return companies[company].get("personnel_by_department", {}).get(dept, [])
    # Fallback to global config if any, or empty list
    return config.get("personnel_by_department", {}).get(dept, [])

def get_departments_for_company(company, config):
    companies = config.get("companies", {})
    if company in companies:
        return companies[company].get("departments", [])
    # Fallback to global departments
    return config.get("departments", [])


def get_filtered_projects(company_name, config, db_projs):
    companies = config.get("companies", {})
    projs_by_cat = {}
    if company_name in companies:
        projs_by_cat = companies[company_name].get("projects_by_category", {})
    else:
        projs_by_cat = config.get("projects_by_category", {})
        
    all_projs = []
    for cat, projs in projs_by_cat.items():
        all_projs.extend(projs)
        
    merged = list(set(all_projs + db_projs))
    return sorted(merged)

DB_FILE = os.path.join("OUTPUT", "DATA_TIEN_DO_KPI.xlsx")

# Gantt DB Configuration
GANTT_DB_FILE = os.path.join("OUTPUT", "DATA_TIEN_DO_KPI.xlsx")

def read_gantt_db():
    required_cols = ["ID", "TenDuAn", "TenCongViec", "GiaiDoan", "NgayBatDau", "Deadline", "PhanTramHoanThanh", "Milestone", "QuanTrong", "KhanCap", "NgayCapNhat"]
    conn = get_gsheets_conn()
    if conn is None:
        return pd.DataFrame(columns=required_cols)
        
    try:
        df = safe_gsheets_read(conn, worksheet="GANTT_KHDT", ttl=15)
        if df is None or df.empty or len(df.columns) < 2:
            df = pd.DataFrame(columns=required_cols)
        else:
            
            df.columns = [str(c).strip() for c in df.columns]
            
            # 1. Tự động chuẩn hóa và ánh xạ tên cột
            col_mapping = {
                'Mã CV': 'ID', 'MaCV': 'ID',
                'Tên công việc': 'TenCongViec', 'TenCongViec': 'TenCongViec', 'Nội dung': 'TenCongViec', 'Công việc': 'TenCongViec',
                'Tiến độ %': 'PhanTramHoanThanh', 'Progress': 'PhanTramHoanThanh', 'Tiến độ': 'PhanTramHoanThanh',
                'Trạng thái': 'TrangThai', 'Status': 'TrangThai',
                'Hạn chót': 'Deadline', 'Ngày hoàn thành': 'Deadline'
            }
            new_cols = []
            for c in df.columns:
                matched = c
                for k, v in col_mapping.items():
                    if c.lower() == k.lower():
                        matched = v
                        break
                new_cols.append(matched)
            df.columns = new_cols
            
            # 3. Xử lý dữ liệu rỗng / NaN an toàn
            if 'PhanTramHoanThanh' in df.columns:
                df['PhanTramHoanThanh'] = pd.to_numeric(df['PhanTramHoanThanh'], errors='coerce').fillna(0)
            if 'TrangThai' in df.columns:
                df['TrangThai'] = df['TrangThai'].fillna('Đang thực hiện')
                df['TrangThai'] = df['TrangThai'].replace('', 'Đang thực hiện')

    except Exception as e:
        import streamlit as st


        st.error(f"Lỗi khi đọc dữ liệu GANTT_KHDT: {e}")
        raise e
        
    # Khởi tạo các cột thiếu
    for col in ["ID", "TenDuAn", "TenCongViec", "GiaiDoan", "NgayBatDau", "Deadline", "PhanTramHoanThanh", "Milestone", "NgayCapNhat"]:
        if col not in df.columns:
            df[col] = ""

            
    df['NgayBatDau'] = pd.to_datetime(df['NgayBatDau'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))).dt.date
    df['Deadline'] = pd.to_datetime(df['Deadline'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))).dt.date
    df['NgayCapNhat'] = df['NgayCapNhat'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))
    df['ID'] = df['ID'].astype(str)
    df['TenDuAn'] = df['TenDuAn'].fillna('Dự án mặc định')
    df['TenCongViec'] = df['TenCongViec'].fillna('')
    df['GiaiDoan'] = df['GiaiDoan'].fillna('Khác')
    df['Milestone'] = df['Milestone'].fillna('')
    df['PhanTramHoanThanh'] = pd.to_numeric(df['PhanTramHoanThanh'], errors='coerce').fillna(0).astype(int)
    
    phase_mapping = {
        "Concept Dev": "1. Chuẩn bị Đầu tư & Nghiên cứu Tiền khả thi",
        "1. Phát triển Ý tưởng & Khảo sát": "1. Chuẩn bị Đầu tư & Nghiên cứu Tiền khả thi",
        "System Design": "2. Pháp lý Dự án & Quy hoạch 1/500",
        "2. Thiết kế Cơ sở & Quy hoạch": "2. Pháp lý Dự án & Quy hoạch 1/500",
        "Detail Design": "3. Thiết kế Cơ sở & Báo cáo Tự đánh giá / ĐTM",
        "3. Thiết kế Chi tiết & Lập Báo cáo": "3. Thiết kế Cơ sở & Báo cáo Tự đánh giá / ĐTM",
        "Legal / Regulatory": "4. Thiết kế Bản vẽ Thi công & Thẩm định",
        "4. Phê duyệt Pháp lý & Thẩm định": "4. Thiết kế Bản vẽ Thi công & Thẩm định",
        "Test & Refine": "5. Cấp phép Xây dựng & Lựa chọn Nhà thầu",
        "5. Thử nghiệm & Chỉnh sửa": "5. Cấp phép Xây dựng & Lựa chọn Nhà thầu",
        "Produce": "6. Thi công Xây lắp & Lắp đặt Thiết bị",
        "Produce / Execute": "6. Thi công Xây lắp & Lắp đặt Thiết bị",
        "6. Triển khai & Thực thi": "6. Thi công Xây lắp & Lắp đặt Thiết bị"
    }
    df['GiaiDoan'] = df['GiaiDoan'].replace(phase_mapping)
    
    return df



def read_kpi_adjustments():
    import pandas as pd
    conn = get_gsheets_conn()
    empty_df = pd.DataFrame(columns=["ID", "TenNhanVien", "Thang", "Nam", "LoaiHanhVi", "DiemDieuChinh", "LyDo"])
    if conn is None:
        return empty_df
    try:
        df = safe_gsheets_read(conn, worksheet="KPI_ADJUSTMENTS", ttl=600)
        if df is None or df.empty:
            return empty_df
        for col in ["Thang", "Nam", "DiemDieuChinh"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        return df
    except Exception as e:
        import streamlit as st


        st.error(f"Lỗi khi đọc trang tính KPI_ADJUSTMENTS: {e}")
        raise e

def add_kpi_adjustment(ten, thang, nam, loai, diem, lydo):
    if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
    import pandas as pd
    df = read_kpi_adjustments()
    if not df.empty:
        dup = df[(df['TenNhanVien'] == ten) & (df['Thang'] == thang) & (df['Nam'] == nam) & (df['LoaiHanhVi'] == loai) & (df['DiemDieuChinh'] == diem) & (df['LyDo'] == lydo)]
        if not dup.empty:
            return True, ""
    if df.empty:
        new_id = 1
    else:
        max_id = pd.to_numeric(df['ID'], errors='coerce').max(skipna=True)
        new_id = 1 if pd.isna(max_id) else int(max_id) + 1
    new_row = {
        "ID": new_id,
        "TenNhanVien": ten,
        "Thang": thang,
        "Nam": nam,
        "LoaiHanhVi": loai,
        "DiemDieuChinh": diem,
        "LyDo": lydo
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    conn = get_gsheets_conn()
    if conn is not None:
        try:
            success = safe_gsheets_update(conn, worksheet="KPI_ADJUSTMENTS", data=df)
            if success:
                import streamlit as st


                if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
                return True, ""
            else:
                return False, "Không thể cập nhật lên Google Sheets (lỗi đã ghi log)"
        except Exception as e:
            return False, str(e)
    return False, "Không kết nối được Google Sheets"

def edit_kpi_adjustment(adj_id, ten, thang, nam, loai, diem, lydo):
    if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
    import pandas as pd
    df = read_kpi_adjustments()
    if df.empty: return False, "Dữ liệu trống"
    idx = df[df['ID'] == adj_id].index
    if len(idx) == 0: return False, "Không tìm thấy ID"
    df.loc[idx[0], 'TenNhanVien'] = ten
    df.loc[idx[0], 'Thang'] = thang
    df.loc[idx[0], 'Nam'] = nam
    df.loc[idx[0], 'LoaiHanhVi'] = loai
    df.loc[idx[0], 'DiemDieuChinh'] = diem
    df.loc[idx[0], 'LyDo'] = lydo
    conn = get_gsheets_conn()
    if conn is not None:
        try:
            success = safe_gsheets_update(conn, worksheet="KPI_ADJUSTMENTS", data=df)
            if success:
                import streamlit as st


                if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
                return True, ""
        except Exception as e:
            return False, str(e)
    return False, "Lỗi kết nối"

def delete_kpi_adjustment(adj_id):
    if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
    import pandas as pd
    df = read_kpi_adjustments()
    if df.empty: return False, "Dữ liệu trống"
    df = df[df['ID'].astype(str).str.strip() != str(adj_id).strip()]
    conn = get_gsheets_conn()
    if conn is not None:
        try:
            success = safe_gsheets_update(conn, worksheet="KPI_ADJUSTMENTS", data=df)
            if success:
                import streamlit as st


                if hasattr(read_kpi_adjustments, "clear"): read_kpi_adjustments.clear()
                return True, ""
        except Exception as e:
            return False, str(e)
    return False, "Lỗi kết nối"


def save_gantt_db(df):
    conn = get_gsheets_conn()
    if conn is None:
        st.error("Chưa kết nối Google Sheets.")
        return False
    try:
        df_save = df.copy()
        df_save['NgayBatDau'] = df_save['NgayBatDau'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['Deadline'] = df_save['Deadline'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['NgayCapNhat'] = df_save['NgayCapNhat'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        
        safe_gsheets_update(conn, worksheet="GANTT_KHDT", data=df_save)
        return True
    except Exception as e:
        st.error(f'Lỗi lưu Google Sheets: {e}')
        return False
def read_sqlite_table(table_name):
    try:
        conn = sqlite3.connect("database.db")
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
        conn.close()
        return df
    except Exception:
        return None

def save_sqlite_table(df, table_name):
    try:
        conn = sqlite3.connect("database.db")
        df.to_sql(table_name, conn, if_exists="replace", index=False)
        conn.close()
        
        if hasattr(read_sqlite_table, "clear"): 
            read_sqlite_table.clear()
            
        return True
    except Exception:
        return False

def convert_gsheet_to_csv_url(url):
    url = url.strip()
    if not url:
        return ""
    # Look for /spreadsheets/d/{spreadsheetId}
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        return url
    key = match.group(1)
    
    # Check if there is a gid parameter (specifying sheet ID)
    gid_match = re.search(r"gid=(\d+)", url)
    if gid_match:
        gid = gid_match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{key}/export?format=csv&gid={gid}"
    else:
        return f"https://docs.google.com/spreadsheets/d/{key}/export?format=csv"

def sync_incoming_docs_from_df(import_df, selected_company, today):
    # Normalize column names
    import_df.columns = [str(c).strip() for c in import_df.columns]
    
    # Map columns using a case-insensitive check
    mapping = {}
    fields = {
        "NGÀY": ["NGÀY", "Ngày", "Ngay"],
        "ĐƠN VỊ": ["ĐƠN VỊ", "Đơn vị", "Don vi", "Co quan gui", "Cơ quan gửi"],
        "NỘI DUNG": ["NỘI DUNG", "Nội dung", "Noi dung", "Trich yeu", "Trích yếu"],
        "Số ký hiệu": ["Số ký hiệu", "Số / Ký hiệu", "So ky hieu", "SỐ KÝ HIỆU", "SoKyHieu"],
        "Thời hạn hoàn thành": ["Thời hạn hoàn thành", "THỜI HẠN HOÀN THÀNH", "Ngày hoàn thành", "NGÀY HOÀN THÀNH", "Deadline"],
        "Trạng thái": ["Trạng thái", "Trang thai", "TRẠNG THÁI", "TrangThai"],
        "Người/ Ban thực hiện": ["Người/ Ban thực hiện", "Nguoi/ Ban thuc hien", "NGƯỜI/ BAN THỰC HIỆN", "Bộ phận chủ trì", "Ban chủ trì", "BanChuTri"],
        "Ghi chú": ["Ghi chú", "Ghi chu", "GhiChu", "Note", "Ghi chú khác"]
    }
    
    for key, possibilities in fields.items():
        found_col = None
        for col in import_df.columns:
            if col.lower() in [p.lower() for p in possibilities]:
                found_col = col
                break
        mapping[key] = found_col
        
    # Check if critical columns exist
    critical_fields = ["NỘI DUNG", "Số ký hiệu", "Thời hạn hoàn thành"]
    missing_critical = [f for f in critical_fields if mapping[f] is None]
    if missing_critical:
        return False, f"Thiếu các cột bắt buộc trong bảng dữ liệu: {', '.join(missing_critical)}"
        
    # Keep rows where "Thời hạn hoàn thành" and "NỘI DUNG" are not null / empty
    deadline_col = mapping["Thời hạn hoàn thành"]
    content_col = mapping["NỘI DUNG"]
    so_ky_hieu_col = mapping["Số ký hiệu"]
    ghi_chu_col = mapping["Ghi chú"]
    
    # Drop rows that are completely empty or have null deadline/content
    valid_df = import_df.dropna(subset=[deadline_col])
    valid_df = valid_df[valid_df[deadline_col].astype(str).str.strip() != ""]
    valid_df = valid_df[valid_df[content_col].notna() & (valid_df[content_col].astype(str).str.strip() != "")]
    
    if valid_df.empty:
        return False, "Không tìm thấy dòng hợp lệ nào chứa đầy đủ thông tin 'Thời hạn hoàn thành' và 'Nội dung'."
        
    with acquire_db_lock():
        
        docs_df = read_incoming_docs_db()
        tasks_df = read_db()
        
        success_count = 0
        update_count = 0
    
        # Convert today to date if it is datetime
        if isinstance(today, datetime):
            today = today.date()
            
        for _, row in valid_df.iterrows():
            # Parse fields
            date_col = mapping["NGÀY"]
            if date_col and not pd.isna(row[date_col]):
                try:
                    ngay_ban_hanh = pd.to_datetime(row[date_col]).date()
                except Exception:
                    ngay_ban_hanh = today
            else:
                ngay_ban_hanh = today
                
            try:
                deadline_val = pd.to_datetime(row[deadline_col]).date()
            except Exception:
                continue
                
            so_ky_hieu = str(row[so_ky_hieu_col]).strip() if not pd.isna(row[so_ky_hieu_col]) else f"VB-{(datetime.utcnow() + timedelta(hours=7)).strftime('%M%S')}"
            co_quan_gui = str(row[mapping["ĐƠN VỊ"]]).strip() if mapping["ĐƠN VỊ"] and not pd.isna(row[mapping["ĐƠN VỊ"]]) else ""
            trich_yeu = str(row[content_col]).strip()
            
            ban_chu_tri_raw = str(row[mapping["Người/ Ban thực hiện"]]).strip() if mapping["Người/ Ban thực hiện"] and not pd.isna(row[mapping["Người/ Ban thực hiện"]]) else ""
            config = load_settings()
            all_depts = set(config.get("departments", []))
            for comp_data in config.get("companies", {}).values():
                all_depts.update(comp_data.get("departments", []))
                
            if ban_chu_tri_raw in all_depts:
                ban_chu_tri = ban_chu_tri_raw
            else:
                ban_chu_tri = "Ban Lãnh đạo"
                
            trang_thai_raw = str(row[mapping["Trạng thái"]]).strip() if mapping["Trạng thái"] and not pd.isna(row[mapping["Trạng thái"]]) else "⏳ Đang xử lý"
            ghi_chu = str(row[ghi_chu_col]).strip() if ghi_chu_col and not pd.isna(row[ghi_chu_col]) else ""
            
            is_completed = trang_thai_raw in ["Đã xong", "Hoàn thành", "Đã hoàn thành", "✅ Đã xong"]
            
            trang_thai = "⏳ Đang xử lý"
            if is_completed:
                trang_thai = "✅ Đã xong"
            else:
                if pd.notna(deadline_val) and deadline_val < today:
                    days_late = (today - deadline_val).days
                    trang_thai = f"⚠️ Trễ hạn xử lý CV (Trễ {days_late} ngày)"
                else:
                    trang_thai = "⏳ Đang xử lý"
                    
            # Check duplicate in docs_df
            duplicate_doc = docs_df[docs_df['SoKyHieu'] == so_ky_hieu]
            
            if duplicate_doc.empty:
                # Generate next DOC ID
                next_doc_id = 1
                if not docs_df.empty:
                    ids = docs_df['ID'].tolist()
                    nums = [int(m[0]) for idx in ids for m in [re.findall(r'\d+', str(idx))] if m]
                    if nums:
                        next_doc_id = max(nums) + 1
                doc_id = f"DOC-{next_doc_id:03d}"
                
                new_doc_row = {
                    "ID": doc_id,
                    "DonVi": selected_company if selected_company != "Tất cả đơn vị" else "CTY CP ĐẦU TƯ ĐÀ NẴNG - MIỀN TRUNG",
                    "SoKyHieu": so_ky_hieu,
                    "NgayBanHanh": ngay_ban_hanh,
                    "CoQuanGui": co_quan_gui,
                    "TrichYeu": trich_yeu,
                    "TenDuAn": "Quản lý Công văn đến",
                    "GanttTaskId": "",
                    "BanChuTri": ban_chu_tri,
                    "Deadline": deadline_val,
                    "LinkFile": "",
                    "TrangThai": trang_thai,
                    "NgayCapNhat": (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S'),
                    "GhiChu": ghi_chu
                }
                docs_df = pd.concat([docs_df, pd.DataFrame([new_doc_row])], ignore_index=True)
                success_count += 1
            else:
                # Update existing document
                doc_id = duplicate_doc.iloc[0]['ID']
                idx = docs_df[docs_df['ID'] == doc_id].index[0]
                docs_df.at[idx, "DonVi"] = selected_company if selected_company != "Tất cả đơn vị" else docs_df.at[idx, "DonVi"]
                docs_df.at[idx, "NgayBanHanh"] = ngay_ban_hanh
                docs_df.at[idx, "CoQuanGui"] = co_quan_gui
                docs_df.at[idx, "TrichYeu"] = trich_yeu
                docs_df.at[idx, "BanChuTri"] = ban_chu_tri
                docs_df.at[idx, "Deadline"] = deadline_val
                docs_df.at[idx, "TrangThai"] = trang_thai
                docs_df.at[idx, "NgayCapNhat"] = (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
                docs_df.at[idx, "GhiChu"] = ghi_chu
                update_count += 1
                
            # Update or create the associated Task in tasks_df
            task_name = f"📩 [Công văn đến] {trich_yeu} (Số: {so_ky_hieu})"
            duplicate_task = tasks_df[tasks_df['TenCongViec'].str.contains(so_ky_hieu, na=False)]
            
            task_status = "Đang thực hiện"
            if trang_thai == "✅ Đã xong":
                task_status = "Hoàn thành"
            elif pd.notna(deadline_val) and deadline_val < today:
                task_status = "Quá hạn"
                
            if duplicate_task.empty:
                next_tsk_id = 1
                if not tasks_df.empty:
                    t_ids = tasks_df['ID'].tolist()
                    t_nums = [int(m[0]) for idx in t_ids for m in [re.findall(r'\d+', str(idx))] if m]
                    if t_nums:
                        next_tsk_id = max(t_nums) + 1
                task_id = f"TSK-{next_tsk_id:03d}"
                
                new_task_row = {
                    "ID": task_id,
                    "DonVi": selected_company if selected_company != "Tất cả đơn vị" else "CTY CP ĐẦU TƯ ĐÀ NẴNG - MIỀN TRUNG",
                    "PhongBan": ban_chu_tri,
                    "NguoiChuTri": "Ban Lãnh đạo",
                    "TenDuAn": "Quản lý Công văn đến",
                    "MocTienDo": "Tự do",
                    "SanPhamBanGiao": "Xem chi tiết văn bản",
                    "TenCongViec": task_name,
                    "PhanLoaiChiSo": "Chỉ số kết quả (Outcome Metric)",
                    "NgayBatDau": ngay_ban_hanh,
                    "Deadline": deadline_val,
                    "DoUuTien": "Trung bình",
                    "PhanTramHoanThanh": 100 if task_status == "Hoàn thành" else 99,
                    "TrangThai": task_status,
                    "LinkKetQua": "",
                    "GiaiTrinhDeXuat": "",
                    "NgayCapNhat": (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S'),
                    "ChuKyTheoDoi": "Theo dự án / Tự do",
                    "PhanLoaiTreHan": "🟢 Không trễ hạn / Đúng tiến độ" if task_status != "Quá hạn" else "👤 Do chủ quan"
                }
                tasks_df = pd.concat([tasks_df, pd.DataFrame([new_task_row])], ignore_index=True)
            else:
                task_id = duplicate_task.iloc[0]['ID']
                t_idx = tasks_df[tasks_df['ID'] == task_id].index[0]
                tasks_df.at[t_idx, "PhongBan"] = ban_chu_tri
                tasks_df.at[t_idx, "TenCongViec"] = task_name
                tasks_df.at[t_idx, "NgayBatDau"] = ngay_ban_hanh
                tasks_df.at[t_idx, "Deadline"] = deadline_val
                tasks_df.at[t_idx, "TrangThai"] = task_status
                tasks_df.at[t_idx, "PhanTramHoanThanh"] = 100 if task_status == "Hoàn thành" else 99
            tasks_df.at[t_idx, "NgayCapNhat"] = (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
            
        if save_incoming_docs_db(docs_df) and save_db(tasks_df):
            return True, f"Đồng bộ thành công! Đã thêm mới {success_count} văn bản và cập nhật {update_count} văn bản."
        else:
            return False, "Không thể lưu dữ liệu vào cơ sở dữ liệu."

def read_incoming_docs_db():
    required_cols = [
        "ID", "DonVi", "SoKyHieu", "NgayBanHanh", "CoQuanGui", "TrichYeu", 
        "TenDuAn", "GanttTaskId", "BanChuTri", "Deadline", "LinkFile", 
        "TrangThai", "NgayCapNhat", "GhiChu"
    ]
    conn = get_gsheets_conn()
    if conn is None:
        return pd.DataFrame(columns=required_cols)
        
    try:
        df = safe_gsheets_read(conn, worksheet="VAN_BAN_DEN", ttl=600)
        if df is None or df.empty or len(df.columns) < 2:
            df = pd.DataFrame(columns=required_cols)
        else:
            
            df.columns = [str(c).strip() for c in df.columns]
            
            # 1. Tự động chuẩn hóa và ánh xạ tên cột
            col_mapping = {
                'Mã CV': 'ID', 'MaCV': 'ID',
                'Tên công việc': 'TenCongViec', 'TenCongViec': 'TenCongViec', 'Nội dung': 'TenCongViec', 'Công việc': 'TenCongViec',
                'Tiến độ %': 'PhanTramHoanThanh', 'Progress': 'PhanTramHoanThanh', 'Tiến độ': 'PhanTramHoanThanh',
                'Trạng thái': 'TrangThai', 'Status': 'TrangThai',
                'Hạn chót': 'Deadline', 'Ngày hoàn thành': 'Deadline'
            }
            new_cols = []
            for c in df.columns:
                matched = c
                for k, v in col_mapping.items():
                    if c.lower() == k.lower():
                        matched = v
                        break
                new_cols.append(matched)
            df.columns = new_cols
            
            # 3. Xử lý dữ liệu rỗng / NaN an toàn
            if 'PhanTramHoanThanh' in df.columns:
                df['PhanTramHoanThanh'] = pd.to_numeric(df['PhanTramHoanThanh'], errors='coerce').fillna(0)
            if 'TrangThai' in df.columns:
                df['TrangThai'] = df['TrangThai'].fillna('Đang thực hiện')
                df['TrangThai'] = df['TrangThai'].replace('', 'Đang thực hiện')

    except Exception as e:
        pass
        df = pd.DataFrame(columns=required_cols)

    # Khởi tạo các cột thiếu để tránh KeyError
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""

        
    df['NgayBanHanh'] = pd.to_datetime(df['NgayBanHanh'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, errors='coerce'))).dt.date
    df['Deadline'] = pd.to_datetime(df['Deadline'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))).dt.date
    df['NgayCapNhat'] = df['NgayCapNhat'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))
    df['ID'] = df['ID'].astype(str)
    
    today_dt = date.today()
    for idx, row in df.iterrows():
        deadline_val = row['Deadline']
        status_val = str(row['TrangThai']).strip()
        
        if isinstance(deadline_val, str):
            try:
                deadline_val = datetime.strptime(deadline_val, '%Y-%m-%d').date()
            except Exception:
                pass
                
        if isinstance(deadline_val, datetime):
            deadline_val = deadline_val.date()
            
        if isinstance(deadline_val, date):
            if deadline_val < today_dt and "✅ Đã xong" not in status_val and "Đã xong" not in status_val:
                days_late = (today_dt - deadline_val).days
                df.at[idx, 'TrangThai'] = f"⚠️ Trễ hạn xử lý CV (Trễ {days_late} ngày)"
                
    return df

def save_incoming_docs_db(df):
    conn = get_gsheets_conn()
    if conn is None:
        st.error("Chưa kết nối Google Sheets.")
        return False
    try:
        df_save = df.copy()
        df_save['NgayBanHanh'] = df_save['NgayBanHanh'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['Deadline'] = df_save['Deadline'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['NgayCapNhat'] = df_save['NgayCapNhat'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        
        safe_gsheets_update(conn, worksheet="VAN_BAN_DEN", data=df_save)
        return True
    except Exception as e:
        st.error(f'Lỗi lưu Google Sheets: {e}')
        return False

def read_db():
    # Force cache clear for new progress calculation rules
    required_cols = [
        "ID", "DonVi", "PhongBan", "NguoiChuTri", "TenDuAn", "MocTienDo", "SanPhamBanGiao",
        "TenCongViec", "PhanLoaiChiSo", "NgayBatDau", "Deadline", "DoUuTien", 
        "PhanTramHoanThanh", "TrangThai", "LinkKetQua", "GiaiTrinhDeXuat", "NgayCapNhat", "ChuKyTheoDoi", "PhanLoaiTreHan", "TyTrongKPI", "NguonGiaoViec", "MucDoGhiNhan"
    ]
    conn = get_gsheets_conn()
    if conn is None:
        return pd.DataFrame(columns=required_cols)
        
    try:
        df = safe_gsheets_read(conn, worksheet="Sheet1", ttl=15)
        if df is None or df.empty or len(df.columns) < 2:
            df = pd.DataFrame(columns=required_cols)
        else:
            
            df.columns = [str(c).strip() for c in df.columns]
            
            # 1. Tự động chuẩn hóa và ánh xạ tên cột
            col_mapping = {
                'Mã CV': 'ID', 'MaCV': 'ID',
                'Tên công việc': 'TenCongViec', 'TenCongViec': 'TenCongViec', 'Nội dung': 'TenCongViec', 'Công việc': 'TenCongViec',
                'Tiến độ %': 'PhanTramHoanThanh', 'Progress': 'PhanTramHoanThanh', 'Tiến độ': 'PhanTramHoanThanh',
                'Trạng thái': 'TrangThai', 'Status': 'TrangThai',
                'Hạn chót': 'Deadline', 'Ngày hoàn thành': 'Deadline'
            }
            new_cols = []
            for c in df.columns:
                matched = c
                for k, v in col_mapping.items():
                    if c.lower() == k.lower():
                        matched = v
                        break
                new_cols.append(matched)
            df.columns = new_cols
            
            # 3. Xử lý dữ liệu rỗng / NaN an toàn
            if 'PhanTramHoanThanh' in df.columns:
                df['PhanTramHoanThanh'] = pd.to_numeric(df['PhanTramHoanThanh'], errors='coerce').fillna(0)
            if 'TrangThai' in df.columns:
                df['TrangThai'] = df['TrangThai'].fillna('Đang thực hiện')
                df['TrangThai'] = df['TrangThai'].replace('', 'Đang thực hiện')

    except Exception as e:
        import streamlit as st


        st.error(f"Lỗi khi đọc dữ liệu Sheet1: {e}")
        raise e

    # Khởi tạo các cột thiếu để tránh KeyError
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""


    # Check and initialize missing columns dynamically
    if "ChuKyTheoDoi" not in df.columns:
        df["ChuKyTheoDoi"] = "Theo dự án / Tự do"
    if "PhanLoaiTreHan" not in df.columns:
        df["PhanLoaiTreHan"] = "🟢 Không trễ hạn / Đúng tiến độ"
    if "NguonGiaoViec" not in df.columns:
        df["NguonGiaoViec"] = "Công việc được giao / định kì"
    if "MucDoGhiNhan" not in df.columns:
        df["MucDoGhiNhan"] = "0% (Không ghi nhận)"
    else:
        def clean_mucdo(val):
            val_str = str(val).strip()
            if val_str == "0.5" or "50" in val_str: return "50%"
            if val_str == "0.8" or "80" in val_str: return "80%"
            if val_str == "0.9" or "90" in val_str: return "90%"
            if "miễn" in val_str.lower() or "loại bỏ" in val_str.lower(): return "Miễn trừ (Loại bỏ KPI)"
            return "0% (Không ghi nhận)"
        df["MucDoGhiNhan"] = df["MucDoGhiNhan"].apply(clean_mucdo)

    for col in required_cols:
        if col not in df.columns:
            df[col] = ""

    # Clean data formats
    df['NgayBatDau'] = pd.to_datetime(df['NgayBatDau'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))).dt.date
    df['Deadline'] = pd.to_datetime(df['Deadline'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))).dt.date
    df['NgayCapNhat'] = df['NgayCapNhat'].astype(str).str.replace('T', ' ', regex=False).str.slice(0, 19).apply(lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce'))
    df['DonVi'] = df['DonVi'].fillna('CTY CP DMT - MARINA (Du thuyền Happy Yacht)')
    df['TenDuAn'] = df['TenDuAn'].fillna('')
    df['MocTienDo'] = df['MocTienDo'].fillna('Tự do')
    df['SanPhamBanGiao'] = df['SanPhamBanGiao'].fillna('Xem chi tiết')
    df['LinkKetQua'] = df['LinkKetQua'].fillna('')
    df['GiaiTrinhDeXuat'] = df['GiaiTrinhDeXuat'].fillna('')
    df['ChuKyTheoDoi'] = df['ChuKyTheoDoi'].fillna('Theo dự án / Tự do')
    df['PhanLoaiTreHan'] = df['PhanLoaiTreHan'].fillna('🟢 Không trễ hạn / Đúng tiến độ')
    df['ID'] = df['ID'].astype(str)
    
    # 🧹 Auto-healing: Dọn dẹp hoàn toàn các công việc trùng lặp do lỗi mạng / click đúp (nếu có)
    if 'ID' in df.columns:
        df['ID'] = df['ID'].astype(str).str.strip()
        df = df.drop_duplicates(subset=['ID'], keep='last').reset_index(drop=True)
    
    for idx, row in df.iterrows():
        is_comp = str(row['TrangThai']).strip() == "Hoàn thành"
        start_d = row['NgayBatDau']
        end_d = row['Deadline']
        df.at[idx, 'PhanTramHoanThanh'] = calculate_time_progress(start_d, end_d, is_comp)
        
    return df

def save_db(df):
    conn = get_gsheets_conn()
    if conn is None:
        st.error("Chưa kết nối Google Sheets.")
        return False
    try:
        df_save = df.copy()
        df_save['NgayBatDau'] = df_save['NgayBatDau'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['Deadline'] = df_save['Deadline'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['NgayCapNhat'] = df_save['NgayCapNhat'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_save['ChuKyTheoDoi'] = df_save['ChuKyTheoDoi'].fillna('Theo dự án / Tự do')
        df_save['PhanLoaiTreHan'] = df_save['PhanLoaiTreHan'].fillna('🟢 Không trễ hạn / Đúng tiến độ')
        if 'NguonGiaoViec' not in df_save.columns: df_save['NguonGiaoViec'] = 'Công việc được giao / định kì'
        df_save['NguonGiaoViec'] = df_save['NguonGiaoViec'].fillna('Công việc được giao / định kì')
        if 'MucDoGhiNhan' not in df_save.columns: df_save['MucDoGhiNhan'] = '0% (Không ghi nhận)'
        df_save['MucDoGhiNhan'] = df_save['MucDoGhiNhan'].fillna('0% (Không ghi nhận)')
        
        df_save = df_save.where(pd.notnull(df_save), None)
        return safe_gsheets_update(conn, worksheet="Sheet1", data=df_save)
    except Exception as e:
        st.error(f'Lỗi lưu Google Sheets: {e}')
        return False

# CSS DMT GROUP Branding Theme (Navy Blue & Orange Gold Accent)
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
    html, body, [class*="css"], .stApp {
        font-family: 'Be Vietnam Pro', sans-serif !important;
    }
    
    /* Make Tab headers bolder and clearer */
    button[data-baseweb="tab"] {
        font-weight: 700 !important;
        font-size: 15px !important;
    }
    button[data-baseweb="tab"] p {
        font-weight: 700 !important;
        font-size: 15px !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] p {
        color: #e65100 !important;
    }
    
    /* Hide Streamlit top-right icons */
    .stDeployButton {display:none !important;}
    .viewerBadge_container__1QSob {display: none !important;}
    .viewerBadge_link__1S137 {display: none !important;}
    
    .block-container {
        padding-top: 4rem !important;
        padding-bottom: 3rem !important;
    }
    .main-title {
        font-size: 28px;
        font-weight: 800;
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 50%, #f97316 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 25px;
        letter-spacing: -0.5px;
    }
    /* Style Streamlit primary button to have brand orange-to-gold gradient */
    div.stButton > button:first-child {
        background: linear-gradient(135deg, #f97316 0%, #f59e0b 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 700 !important;
        font-size: 16px !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button:first-child:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(249, 115, 22, 0.3) !important;
    }
    
    /* Style Streamlit download button to have brand Navy-Blue gradient and White text */
    div.stDownloadButton > button:first-child {
        background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 700 !important;
        font-size: 16px !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
    }
    div.stDownloadButton > button:first-child:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.3) !important;
    }
    
    /* Style metrics cards to feel premium with Navy Blue border */
    div[data-testid="stMetric"] {
        background-color: #f8fafc;
        border: 1.8px solid #1e3a8a;
        border-radius: 10px;
        padding: 14px 18px;
        box-shadow: 0 2px 5px rgba(30, 58, 138, 0.05);
    }
    div[data-testid="stMetric"] div[data-testid="stMetricLabel"] p {
        font-size: 16px !important;
        font-weight: 700 !important;
        color: #1e3a8a !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        font-size: 26px !important;
        font-weight: 800 !important;
        color: #f97316 !important;
    }
    
    /* Input field labels - bold and larger font */
    div[data-testid="stWidgetLabel"] p {
        font-size: 16px !important;
        font-weight: 700 !important;
        color: #0f172a !important;
    }
    
    /* Headers styling */
    h1, h2, h3, h4 {
        font-weight: 700 !important;
        color: #1e3a8a !important;
    }
    
    /* Style sidebar with Navy theme styling & high contrast text */
    section[data-testid="stSidebar"] {
        background-color: #0f172a !important;
    }
    section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p, 
    section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
    section[data-testid="stSidebar"] span[data-baseweb="select"] div,
    section[data-testid="stSidebar"] div[role="radiogroup"] label p {
        color: #ffffff !important;
        font-size: 16px !important;
        font-weight: 600 !important;
    }
    section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 {
        color: #ffd700 !important;
        font-size: 18px !important;
        font-weight: 800 !important;
        border-bottom: 2px solid #f97316;
        padding-bottom: 5px;
    }
    
    /* Pull logo up to the very top of Sidebar */
    [data-testid="stSidebarContent"] {
        padding-top: 10px !important;
    }
    [data-testid="stSidebarContent"] img {
        margin-top: -30px !important;
    }
</style>
""", unsafe_allow_html=True)

# Main Header Title with DMT branding
st.markdown('<div class="main-title">DMT GROUP — QUẢN LÝ TIẾN ĐỘ</div>', unsafe_allow_html=True)

# Sidebar layout with logo image and fallback
logo_path = "logo.png" if os.path.exists("logo.png") else ("INPUT/logo.png" if os.path.exists("INPUT/logo.png") else None)
if logo_path:
    st.sidebar.image(logo_path, use_container_width=True)
else:
    st.sidebar.warning("💡 Vui lòng đặt file logo.png vào thư mục gốc của dự án để hiển thị logo.")
    st.sidebar.markdown("### DMT GROUP")
st.sidebar.markdown("---")

# Link Google Sheets Config (Silently initialize for all users)
if "gsheet_url" not in st.session_state:
    st.session_state["gsheet_url"] = load_settings().get("gsheet_url", "")

company_options = ["Tất cả đơn vị"] + list(COMPANIES.keys())
selected_company = st.sidebar.selectbox(
    "CHỌN CÔNG TY / THÀNH VIÊN", 
    company_options, 
    index=1,
    format_func=lambda x: str(x).replace("CTY CP", "CÔNG TY CP")
)

role_mode = st.sidebar.selectbox("QUYỀN TRUY CẬP", ["Nhân viên", "Quản lý", "HR", "Cá nhân (Thử nghiệm)"], index=0)


if "is_admin_authenticated" not in st.session_state:
    st.session_state.is_admin_authenticated = False
if "is_manager_authenticated" not in st.session_state:
    st.session_state.is_manager_authenticated = False
if "is_personal_authenticated" not in st.session_state:
    st.session_state.is_personal_authenticated = False
if "personal_user" not in st.session_state:
    st.session_state.personal_user = None

if role_mode == "Quản lý":
    st.session_state.is_personal_authenticated = False
    st.session_state.personal_user = None
    st.session_state.is_admin_authenticated = False
    if not st.session_state.is_manager_authenticated:
        mgr_pwd = st.sidebar.text_input("Nhập Mật khẩu Quản lý", type="password")
        if mgr_pwd:
            if mgr_pwd == "quanly123":
                st.session_state.is_manager_authenticated = True
                st.rerun()
            else:
                st.sidebar.error("Mật khẩu không đúng!")
    
    if st.session_state.is_manager_authenticated:
        st.sidebar.success("Đã xác thực quyền Quản lý!")
        
        if st.sidebar.button("Đăng xuất"):
            st.session_state.is_manager_authenticated = False
            st.rerun()

elif role_mode == "HR":
    st.session_state.is_personal_authenticated = False
    st.session_state.personal_user = None
    st.session_state.is_manager_authenticated = False
    if not st.session_state.is_admin_authenticated:
        admin_pwd = st.sidebar.text_input("Nhập Mật khẩu HR", type="password")
        if admin_pwd:
            if admin_pwd == "admindmt123":
                st.session_state.is_admin_authenticated = True
                st.rerun()
            else:
                st.sidebar.error("Mật khẩu không đúng!")
    
    if st.session_state.is_admin_authenticated:
        st.sidebar.success("Đã xác thực toàn quyền (HR)!")
        
        if st.sidebar.button("Đăng xuất", key="logout_hr"):
            st.session_state.is_admin_authenticated = False
            st.rerun()

elif role_mode == "Cá nhân (Thử nghiệm)":
    st.session_state.is_admin_authenticated = False
    st.session_state.is_manager_authenticated = False
    if not st.session_state.is_personal_authenticated:
        pers_pwd = st.sidebar.text_input("Nhập MÃ PIN cá nhân (Mặc định: 1234)", type="password")
        if pers_pwd:
            if pers_pwd == "1234":
                st.session_state.is_personal_authenticated = True
                st.rerun()
            else:
                st.sidebar.error("MÃ PIN không đúng!")
    
    if st.session_state.is_personal_authenticated:
        st.sidebar.success("Đã xác thực quyền Cá nhân!")
        
        if st.sidebar.button("Đăng xuất"):
            st.session_state.is_personal_authenticated = False
            st.session_state.personal_user = None
            st.rerun()
else:
    st.session_state.is_admin_authenticated = False
    st.session_state.is_manager_authenticated = False
    st.session_state.is_personal_authenticated = False
    st.session_state.personal_user = None
st.sidebar.markdown("---")

is_mobile = False
try:
    if hasattr(st, "context") and hasattr(st.context, "headers"):
        ua = st.context.headers.get("User-Agent", "").lower()
        if "mobi" in ua or "android" in ua or "iphone" in ua:
            is_mobile = True
except Exception:
    pass

menu_options = [
    "🚀 Bảng theo dõi tiến độ công việc",
    "➕ Thêm / Cập Nhật Công Việc",
    "📖 Sổ tay Hướng dẫn"
]
if is_mobile:
    menu_options.insert(0, "👀 BẢNG TỔNG QUAN (View)")

if st.session_state.get('is_manager_authenticated', False):
    menu_options = [
        "📋 Bảng theo dõi tiến độ công việc",
        "➕ Thêm / Cập Nhật Công Việc",
        "⚖️ Duyệt việc Khách quan",
        "🏆 Đánh giá KPI & Xếp loại",
        "📖 Sổ tay Hướng dẫn"
    ]
    if is_mobile:
        menu_options.insert(0, "📊 BẢNG TỔNG QUAN (View)")

if st.session_state.is_admin_authenticated:
    menu_options = [
        "👀 BẢNG TỔNG QUAN (View)",
        "🚀 Bảng theo dõi tiến độ công việc",
        "➕ Thêm / Cập Nhật Công Việc",
        "✅ Duyệt việc Khách quan",
        "🏆 Đánh giá KPI & Xếp loại",
        "🔍 Quản lý & Đối chiếu JD",
        "⚙️ Quản Lý Cấu Hình",
        "📖 Sổ tay Hướng dẫn"
    ]

menu = st.sidebar.radio(
    "PHÂN HỆ CHỨC NĂNG",
    menu_options,
    index=0
)

st.sidebar.markdown("---")

# Current date
df = read_db()
if not df.empty and "PhongBan" in df.columns:
    df["PhongBan"] = df["PhongBan"].map(lambda x: DEPT_ABBR.get(x, x))

gantt_df = read_gantt_db()
today = date.today()

# Filter display dataframe based on sidebar selected company
if selected_company != "Tất cả đơn vị":
    display_df = df[df['DonVi'] == selected_company].copy()
else:
    display_df = df.copy()
    
if 'NguoiChuTri' not in display_df.columns:
    display_df['NguoiChuTri'] = ''

# -------- LỌC CÁ NHÂN ---------
if role_mode == "Cá nhân (Thử nghiệm)" and st.session_state.is_personal_authenticated:
    all_owners = sorted(list(display_df['NguoiChuTri'].dropna().astype(str).unique()))
    
    # Render selectbox in sidebar
    st.sidebar.markdown("---")
    st.sidebar.markdown("**Cấu hình Cá nhân**")
    
    current_idx = 0
    if st.session_state.personal_user in all_owners:
        current_idx = all_owners.index(st.session_state.personal_user)
    elif "Nguyễn Thị Hạnh Tiên" in all_owners:
        current_idx = all_owners.index("Nguyễn Thị Hạnh Tiên")
        
    selected_user = st.sidebar.selectbox("Bạn là ai?", all_owners, index=current_idx)
    st.session_state.personal_user = selected_user
    
    # Filter display_df
    if st.session_state.personal_user:
        display_df = display_df[display_df['NguoiChuTri'] == st.session_state.personal_user].copy()
        
# ------------------------------

# Statistics helpers
total_v = len(display_df)
done_v = len(display_df[display_df['TrangThai'] == 'Hoàn thành'])
issue_v = len(display_df[display_df['TrangThai'] == 'Có vướng mắc'])
overdue_v = len(display_df[(pd.to_datetime(display_df['Deadline'], errors='coerce') < pd.Timestamp(today)) & (display_df['TrangThai'] != 'Hoàn thành')])
doing_v = total_v - done_v - issue_v - overdue_v
if doing_v < 0:
    doing_v = 0

# Sidebar Excel Download
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_excel(tasks_df, df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write Sheet1 (Drop ID column if exists)
        tasks_df_copy = tasks_df.copy()
        
        # Format the PhanLoaiTreHan column for Excel
        formatted_causes = []
        for _, row in tasks_df_copy.iterrows():
            is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
            deadline = row.get('Deadline')
            if isinstance(deadline, str):
                try:
                    deadline = datetime.strptime(deadline, '%Y-%m-%d').date()
                except Exception:
                    pass
            ref_today = today
            if isinstance(ref_today, datetime):
                ref_today = ref_today.date()
            if isinstance(deadline, datetime):
                deadline = deadline.date()
                
            is_late = False
            if isinstance(deadline, date):
                is_late = (deadline < ref_today) and not is_comp
                
            if not is_late:
                formatted_causes.append("")
            else:
                val = row.get('PhanLoaiTreHan', '')
                if "chủ quan" in str(val).lower():
                    formatted_causes.append("[Do chủ quan]")
                elif "khách quan" in str(val).lower():
                    explain = row.get('GiaiTrinhDeXuat', '')
                    if explain and explain != "--" and str(explain).strip():
                        formatted_causes.append(f"[Do khách quan] - {str(explain).strip()}")
                    else:
                        formatted_causes.append("[Do khách quan]")
                else:
                    formatted_causes.append("")
                    
        tasks_df_copy['PhanLoaiTreHan'] = formatted_causes
        tasks_df_copy = tasks_df_copy.rename(columns={'PhanLoaiTreHan': 'Nguyên nhân trễ hạn'})
        
        # Column filtering and renaming
        export_cols = {
            'NgayBatDau': 'Ngày bắt đầu',
            'Deadline': 'Hạn chót',
            'TrangThai': 'Trạng thái thực hiện',
            'NguoiChuTri': 'Người thực hiện',
            'PhongBan': 'Phòng ban',
            'TenDuAn': 'Dự án / Hạng mục',
            'TenCongViec': 'Tên công việc',
            'GiaiTrinhDeXuat': 'Ghi chú / Giải trình vướng mắc'
        }
        
        cols_to_keep = [c for c in export_cols.keys() if c in tasks_df_copy.columns]
        tasks_df_copy = tasks_df_copy[cols_to_keep]
        tasks_df_copy = tasks_df_copy.rename(columns=export_cols)
        tasks_df_copy.insert(0, 'STT', range(1, len(tasks_df_copy) + 1))
        
        desired_order = ['STT', 'Ngày bắt đầu', 'Hạn chót', 'Trạng thái thực hiện', 'Người thực hiện', 'Phòng ban', 'Dự án / Hạng mục', 'Tên công việc', 'Ghi chú / Giải trình vướng mắc']
        final_cols = [c for c in desired_order if c in tasks_df_copy.columns]
        tasks_df_copy = tasks_df_copy[final_cols]
        
        for col in ['Ngày bắt đầu', 'Hạn chót']:
            if col in tasks_df_copy.columns:
                tasks_df_copy[col] = pd.to_datetime(tasks_df_copy[col], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')
                
        tasks_df_copy.to_excel(writer, sheet_name="Sheet1", index=False)
        
        # Write GANTT_KHDT (Drop ID column if exists)
        df_copy = df.copy()
        if 'ID' in df_copy.columns:
            df_copy = df_copy.drop(columns=['ID'])
        if 'NgayCapNhat' in df_copy.columns:
            df_copy['NgayCapNhat'] = df_copy['NgayCapNhat'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_copy.to_excel(writer, sheet_name="GANTT_KHDT", index=False)
        
        # Write VAN_BAN_DEN (Drop ID column if exists)
        try:
            docs_df = read_incoming_docs_db()
            docs_df_copy = docs_df.copy()
            if 'ID' in docs_df_copy.columns:
                docs_df_copy = docs_df_copy.drop(columns=['ID'])
            if 'NgayCapNhat' in docs_df_copy.columns:
                docs_df_copy['NgayCapNhat'] = docs_df_copy['NgayCapNhat'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
            for col_name in ['NgayBanHanh', 'Deadline']:
                if col_name in docs_df_copy.columns:
                    docs_df_copy[col_name] = docs_df_copy[col_name].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
            docs_df_copy = docs_df_copy.rename(columns={
                'SoKyHieu': 'Số / Ký hiệu văn bản',
                'NgayBanHanh': 'Ngày ban hành',
                'CoQuanGui': 'Cơ quan / Đơn vị gửi',
                'TrichYeu': 'Trích yếu nội dung',
                'TenDuAn': 'Dự án liên quan',
                'GanttTaskId': 'Mã CV Gantt liên kết',
                'BanChuTri': 'Bộ phận chủ trì xử lý',
                'Deadline': 'Hạn xử lý / Phản hồi',
                'LinkFile': 'Đính kèm Link / File',
                'TrangThai': 'Trạng thái xử lý',
                'NgayCapNhat': 'Thời gian cập nhật'
            })
            docs_df_copy.to_excel(writer, sheet_name="VAN_BAN_DEN", index=False)
        except Exception:
            pass
            
        workbook = writer.book
        
        # Helper to style each worksheet
        def style_worksheet(ws):
            navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            normal_font = Font(name="Calibri", size=11)
            
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            thin_border_side = Side(border_style="thin", color="CCCCCC")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            headers = [str(cell.value or '') for cell in ws[1]]
            
            ws.row_dimensions[1].height = 28
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = navy_fill
                cell.font = white_bold_font
                cell.alignment = center_align
                cell.border = thin_border
                
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 22
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    col_name = headers[col - 1]
                    if col_name in ["ID", "STT", "NgayBatDau", "Deadline", "Deadline", "PhanTramHoanThanh", "TrangThai", "NgayCapNhat"]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
                    if col_name == "PhanTramHoanThanh":
                        cell.number_format = '0"%"'
                        
            # Autofit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '').replace('\n', ' ')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)
                
        style_worksheet(workbook["Sheet1"])
        style_worksheet(workbook["GANTT_KHDT"])
        if "VAN_BAN_DEN" in workbook.sheetnames:
            style_worksheet(workbook["VAN_BAN_DEN"])
        
    return output.getvalue()

st.sidebar.markdown("---")
st.sidebar.markdown("### 📥 XUẤT DỮ LIỆU EXCEL")
try:
    df_for_excel = df
    styled_excel_data = generate_styled_excel(df, df_for_excel)
    st.sidebar.download_button(
        label="Tải xuống tệp Excel",
        data=styled_excel_data,
        file_name="DATA_TIEN_DO_KPI.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_sidebar_excel_dl"
    )
except Exception as e:
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "rb") as f:
            st.sidebar.download_button(
                label="Tải xuống tệp Excel (Dự phòng)",
                data=f.read(),
                file_name="DATA_TIEN_DO_KPI.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_sidebar_excel_dl_fallback"
            )



# Helper function to clean project name whitespace
def clean_proj_name(name):
    return name.strip()

# ----------------- 1. DASHBOARD TỔNG QUAN -----------------
if menu in ["🚀 Bảng theo dõi tiến độ công việc", "📋 Bảng theo dõi tiến độ công việc"]:
    st.info("💡 **Gợi ý:** Để xem chi tiết hướng dẫn sử dụng phần mềm, bạn hãy nhấp vào mục **📖 Sổ tay Hướng dẫn** ở thanh Menu bên trái nhé!")
    
    st.markdown(f"### 🚀 Bảng theo dõi tiến độ công việc — {selected_company}")

    
    # Calculate stats based on filtered dash_df
    dash_df = display_df.copy()
    total_dash = len(dash_df)
    done_dash = len(dash_df[dash_df['TrangThai'] == 'Hoàn thành'])
    issue_dash = len(dash_df[dash_df['TrangThai'] == 'Có vướng mắc'])
    overdue_dash = len(dash_df[(pd.to_datetime(dash_df['Deadline'], errors='coerce') < pd.Timestamp(today)) & (dash_df['TrangThai'] != 'Hoàn thành')])
    doing_dash = total_dash - done_dash - issue_dash - overdue_dash
    if doing_dash < 0:
        doing_dash = 0
        
    # 4 metrics cards
    m_col1, m_col2, m_col3, m_col4 = st.columns(4)
    with m_col1:
        st.metric("Tổng số việc", total_dash)
    with m_col2:
        st.metric("Đã xong", done_dash)
    with m_col3:
        st.metric("Đang làm", doing_dash)
    with m_col4:
        st.metric("🔴 Trễ hạn / Vướng mắc", issue_dash + overdue_dash)

    st.markdown("---")
    
    st.markdown("""
        <style>
        button[data-baseweb="tab"] {
            font-size: 18px !important;
            font-weight: bold !important;
            padding: 1rem !important;
        }
        button[data-baseweb="tab"] span {
            font-size: 18px !important;
            font-weight: 800 !important;
            color: #555555 !important;
        }
        button[data-baseweb="tab"][aria-selected="true"] span {
            color: #1976d2 !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    tab_report, tab_giaoban, tab_data = st.tabs(["📊 CÔNG VIỆC TỚI HẠN", "📢 BÁO CÁO GIAO BAN", "📋 BẢNG THEO DÕI TIẾN ĐỘ CÔNG VIỆC"])
    
    with tab_report:
        st.markdown(f"### 📊 Dashboard Tổng Quan — {selected_company}")
    
        pass
        
        # Overdue and due today/tomorrow alerts scanning (Group 1 & 2)
        def get_badge_and_urgency(deadline_val, today_dt):
            if pd.isna(deadline_val):
                return None, None
            if not isinstance(deadline_val, date):
                if isinstance(deadline_val, datetime):
                    deadline_val = deadline_val.date()
                else:
                    return None, None
            if deadline_val < today_dt:
                days_late = (today_dt - deadline_val).days
                return f"🔴 [⚠️ Trễ {days_late} ngày]", 1
            elif deadline_val == today_dt:
                return "⏳ [Hạn hôm nay]", 2
            elif deadline_val == today_dt + timedelta(days=1):
                return "⚠️ [Hạn ngày mai]", 3
            return None, None

        alert_list = []
        for _, row in dash_df[dash_df['TrangThai'] != 'Hoàn thành'].iterrows():
            badge, urgency = get_badge_and_urgency(row['Deadline'], today)
            if badge:
                row_copy = row.copy()
                row_copy['Badge'] = badge
                row_copy['Urgency'] = urgency
                alert_list.append(row_copy)

        if alert_list:
            alert_df_show = pd.DataFrame(alert_list)
            if 'NgayCapNhat' in alert_df_show.columns:
                alert_df_show = alert_df_show.sort_values(by=["Urgency", "NgayCapNhat", "ID"], ascending=[True, False, False])
            else:
                alert_df_show = alert_df_show.sort_values(by=["Urgency", "Deadline"])
            st.error(f"🚨 **CẢNH BÁO: DỰ ÁN CÓ {len(alert_df_show)} HẠNG MỤC CẦN LƯU Ý (TRỄ HẠN / SẮP ĐẾN HẠN)**")        
        st.markdown("---")
    
        # Critical alert panel
        st.markdown("### ⚠️ Hạng mục cần lưu ý (Trễ hạn hoặc Sắp đến hạn)")
    
        if alert_list:
            alert_df_show = pd.DataFrame(alert_list)
            if 'NgayCapNhat' in alert_df_show.columns:
                alert_df_show = alert_df_show.sort_values(by=["Urgency", "NgayCapNhat", "ID"], ascending=[True, False, False])
            else:
                alert_df_show = alert_df_show.sort_values(by=["Urgency", "Deadline"])
            crit_display = pd.DataFrame()
            crit_display['Ngày bắt đầu'] = alert_df_show['NgayBatDau'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
            crit_display['Hạn chót'] = alert_df_show['Deadline'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
            crit_display['Tiến độ'] = alert_df_show['PhanTramHoanThanh'].apply(lambda x: f"{int(x)}%" if pd.notna(x) else "0%")
            crit_display['Trạng thái thực tế'] = alert_df_show['Badge']
            crit_display['Người thực hiện'] = alert_df_show['NguoiChuTri']
            crit_display['Phòng ban'] = alert_df_show['PhongBan']
            crit_display['Dự án / Hạng mục'] = alert_df_show['TenDuAn']
            crit_display['Tên công việc'] = alert_df_show['TenCongViec']
            crit_display['Ghi chú / Giải trình vướng mắc'] = alert_df_show['GiaiTrinhDeXuat']
        
            st.dataframe(
                crit_display,
                column_config={
                    "Ngày bắt đầu": st.column_config.TextColumn("Ngày bắt đầu", width=90),
                    "Hạn chót": st.column_config.TextColumn("Hạn chót", width=150),
                    "Tiến độ": st.column_config.TextColumn("Tiến độ", width=80),
                    "Trạng thái thực tế": st.column_config.TextColumn("Trạng thái thực tế", width=120),
                    "Người thực hiện": st.column_config.TextColumn("Người thực hiện", width=150),
                    "Phòng ban": st.column_config.TextColumn("Phòng ban", width=80),
                    "Dự án / Hạng mục": st.column_config.TextColumn("Dự án / Hạng mục", width=200),
                    "Tên công việc": st.column_config.TextColumn("Tên công việc", width="large"),
                    "Ghi chú / Giải trình vướng mắc": st.column_config.TextColumn("Ghi chú / Giải trình vướng mắc", width="large")
                },
                use_container_width=True,
                hide_index=True
            )
        else:
            st.success("🎉 Đảm bảo tiến độ: Không có công việc nào bị trễ hạn hoặc sắp đến hạn cần lưu ý!")
        
        st.markdown("---")
    


    # ----------------- 1.5. BÁO CÁO GIAO BAN -----------------
    with tab_giaoban:
        st.markdown(f"### 📢 Báo cáo Giao ban — {selected_company}")
        
        # Lọc các công việc có nguồn giao việc là "Giao ban"
        gb_df = display_df[display_df['NguonGiaoViec'].astype(str).str.contains("Giao ban", na=False, case=False)].copy()
        
        if not gb_df.empty:
            def _get_priority_gb(row):
                st_val = str(row.get('TrangThai', ''))
                if 'Trễ hạn' in st_val or 'Vướng mắc' in st_val or '🔴' in st_val or '⚠️' in st_val:
                    return 0
                return 1
            gb_df['SortPriority'] = gb_df.apply(_get_priority_gb, axis=1)
            if 'NgayCapNhat' in gb_df.columns:
                gb_df = gb_df.sort_values(by=['SortPriority', 'NgayCapNhat', 'ID'], ascending=[True, False, False]).reset_index(drop=True)
            else:
                gb_df = gb_df.sort_values(by=['SortPriority', 'ID'], ascending=[True, False]).reset_index(drop=True)
        
        if gb_df.empty:
            st.info('Chưa có công việc nào có "Nguồn giao việc" là "Giao ban".')
            st.write('💡 Để thêm công việc vào Giao ban, hãy chọn Nguồn giao việc là **Công việc trong "Giao ban"** khi tạo hoặc cập nhật công việc.')
        else:
            total_gb = len(gb_df)
            done_gb = len(gb_df[gb_df['TrangThai'] == 'Hoàn thành'])
            issue_gb = len(gb_df[gb_df['TrangThai'] == 'Có vướng mắc'])
            
            gb_col1, gb_col2, gb_col3, gb_col4 = st.columns(4)
            gb_col1.metric("📌 Tổng Số Việc Giao Ban", total_gb)
            gb_col2.metric("✅ Đã Hoàn Thành", done_gb)
            gb_col3.metric("🔥 Đang Vướng Mắc", issue_gb)
            gb_col4.metric("⏳ Đang Thực Hiện", total_gb - done_gb - issue_gb)
            
            st.markdown("#### 📋 Danh sách chi tiết:")
            
            # Formatting the table for Giao ban
            gb_display = gb_df[['Deadline', 'NguoiChuTri', 'TenCongViec', 'PhanTramHoanThanh', 'TrangThai', 'GiaiTrinhDeXuat']]
            st.dataframe(
                gb_display,
                column_config={
                    "Deadline": st.column_config.DateColumn("Hạn chót", format="DD/MM/YYYY"),
                    "NguoiChuTri": "Người phụ trách",
                    "TenCongViec": st.column_config.TextColumn("Tên công việc", width="large"),
                    "PhanTramHoanThanh": st.column_config.ProgressColumn("Tiến độ", format="%d%%", min_value=0, max_value=100),
                    "Trạng thái": "Trạng thái",
                    "GiaiTrinhDeXuat": st.column_config.TextColumn("Vướng mắc / Giải trình", width="medium")
                },
                use_container_width=True,
                hide_index=True
            )

    # ----------------- 2. BẢNG TIẾN ĐỘ CHI TIẾT -----------------

    with tab_data:
        st.markdown(f"### 📋 Bảng Tiến Độ Công Việc Chi Tiết — {selected_company}")
    
        # Filter tools for Boss
        col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
    
        with col_filter1:
            db_projs = list(display_df["TenDuAn"].dropna().unique()) if not display_df.empty else []
            merged_projs = get_filtered_projects(selected_company, config, db_projs)
            proj_options = ["Tất cả dự án"] + merged_projs
            sel_proj_filter = st.selectbox("Lọc nhanh theo Dự án / Hạng mục", proj_options)
        
        with col_filter2:
            allowed_depts = get_departments_for_company(selected_company, config)
            dept_options = ["Tất cả phòng ban"] + allowed_depts
            sel_dept_filter = st.selectbox("Lọc nhanh theo Phòng ban", dept_options)
        
        with col_filter3:
            owners = ["Tất cả"] + sorted(list(display_df['NguoiChuTri'].dropna().astype(str).unique())) if not display_df.empty else ["Tất cả"]
            sel_owner_filter = st.selectbox("Lọc theo Người phụ trách", owners)
            
        with col_filter4:
            months = set()
            if not display_df.empty:
                for _, row in display_df.iterrows():
                    if pd.notna(row.get('NgayBatDau')) and hasattr(row['NgayBatDau'], 'strftime'):
                        months.add(row['NgayBatDau'].strftime('%m/%Y'))
                    if pd.notna(row.get('Deadline')) and hasattr(row['Deadline'], 'strftime'):
                        months.add(row['Deadline'].strftime('%m/%Y'))
            month_options = ["Tất cả các tháng"] + sorted(list(months), key=lambda x: datetime.strptime(x, '%m/%Y'), reverse=True)
            sel_month_filter = st.selectbox("Lọc nhanh theo Tháng", month_options)
        
        # Apply filters
        table_df = display_df.copy()
        if sel_proj_filter != "Tất cả dự án":
            clean_proj = clean_proj_name(sel_proj_filter)
            table_df = table_df[table_df['TenDuAn'].str.contains(clean_proj, case=False, na=False)]
        
        if sel_dept_filter != "Tất cả phòng ban":
            table_df = table_df[table_df['PhongBan'] == sel_dept_filter]
            
        if sel_owner_filter != "Tất cả":
            table_df = table_df[table_df['NguoiChuTri'] == sel_owner_filter]
        
        if sel_month_filter != "Tất cả các tháng":
            target_month = sel_month_filter
            mask = (
                table_df['NgayBatDau'].apply(lambda x: x.strftime('%m/%Y') if pd.notna(x) and hasattr(x, 'strftime') else '') == target_month
            ) | (
                table_df['Deadline'].apply(lambda x: x.strftime('%m/%Y') if pd.notna(x) and hasattr(x, 'strftime') else '') == target_month
            )
            table_df = table_df[mask]
            
        # Sắp xếp: Ghim Trễ hạn/Vướng mắc lên đầu, sau đó mới đến công việc mới cập nhật
        def _get_priority(row):
            st_val = str(row.get('TrangThai', ''))
            if 'Trễ hạn' in st_val or 'Vướng mắc' in st_val or '🔴' in st_val or '⚠️' in st_val:
                return 0
            return 1
            
        table_df['SortPriority'] = table_df.apply(_get_priority, axis=1)
        if 'NgayCapNhat' in table_df.columns:
            table_df = table_df.sort_values(by=['SortPriority', 'NgayCapNhat', 'ID'], ascending=[True, False, False]).reset_index(drop=True)
        else:
            table_df = table_df.sort_values(by=['SortPriority', 'ID'], ascending=[True, False]).reset_index(drop=True)
        
        if table_df.empty:
            st.info("Không có công việc nào phù hợp với bộ lọc.")
        else:
            df_display = pd.DataFrame()
            df_display['Phòng ban'] = table_df['PhongBan']
            df_display['Người thực hiện'] = table_df['NguoiChuTri']
            df_display['Dự án / Hạng mục'] = table_df['TenDuAn']
            df_display['Tên công việc'] = table_df['TenCongViec']
            df_display['Ngày bắt đầu'] = table_df['NgayBatDau'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
            import pandas as pd
            df_display['Tỷ trọng KPI'] = table_df.apply(lambda row: f"{int(float(str(row.get('TyTrongKPI', 0)).strip() or 0))}%" if pd.to_numeric(row.get('TyTrongKPI', 0), errors='coerce') > 0 else "Tự chia", axis=1)
        
            # Format Hạn chót
            def format_dl(row):
                if pd.isna(row['Deadline']) or not isinstance(row['Deadline'], (date, datetime)): return ""
                prog = int(row['PhanTramHoanThanh'])
                date_str = row['Deadline'].strftime('%d/%m/%Y')
            
                if prog >= 100:
                    return date_str
            
                # prog < 100
                days_left = (row['Deadline'] - today).days
                if days_left < 0:
                    days_late = abs(days_left)
                    return f"🔴 {date_str} (Trễ hạn {days_late} ngày)"
                elif days_left == 0:
                    return f"⏳ {date_str} (Hạn hôm nay)"
                elif 1 <= days_left <= 3:
                    return f"⚠️ {date_str} (Sắp hạn - Còn {days_left} ngày)"
                else:
                    return date_str
            df_display['Hạn chót'] = table_df.apply(format_dl, axis=1)
        
            df_display['Tiến độ'] = table_df['PhanTramHoanThanh']
        
            # Format Trạng thái
            def format_status(row):
                prog = int(row['PhanTramHoanThanh'])
                is_issue = row['TrangThai'] == 'Có vướng mắc'
            
                if prog >= 100:
                    return "✅ Đã xong"
                
                # prog < 100
                if pd.notna(row['Deadline']) and row['Deadline'] < today:
                    return "⚠️ Trễ hạn"
                
                if is_issue:
                    return "🔴 Vướng mắc"
                
                from datetime import date, datetime
                if prog == 0 and pd.notna(row['NgayBatDau']) and isinstance(row['NgayBatDau'], (date, datetime)) and row['NgayBatDau'] > today:
                    return "❌ Chưa bắt đầu"
                
                # Default state based on start date
                if pd.notna(row['NgayBatDau']) and isinstance(row['NgayBatDau'], (date, datetime)):
                    if today >= row['NgayBatDau']:
                        return "⏳ Đang thực hiện"
                    else:
                        return "❌ Chưa bắt đầu"
                else:
                    return "⏳ Đang thực hiện"
            df_display['Trạng thái'] = table_df.apply(format_status, axis=1)
        
            # Format Nguyên nhân trễ hạn
            def format_late_cause(row):
                is_comp = (row['TrangThai'] == 'Hoàn thành')
                is_late = (pd.notna(row['Deadline']) and row['Deadline'] < today) and not is_comp
                if not is_late:
                    return "--"
            
                val = row.get('PhanLoaiTreHan', '')
                if "chủ quan" in str(val).lower():
                    return "🔴 [Do chủ quan]"
                elif "khách quan" in str(val).lower():
                    explain = row.get('GiaiTrinhDeXuat', '')
                    if pd.notna(explain) and str(explain).strip():
                        return f"⚠️ [Do khách quan] - {str(explain).strip()}"
                    return "⚠️ [Do khách quan]"
                else:
                    return "--"
            df_display['Nguyên nhân trễ hạn'] = table_df.apply(format_late_cause, axis=1)
        
            # Format Kết quả / File đính kèm
            def format_notes(row):
                is_comp = (row['TrangThai'] == 'Hoàn thành')
                if is_comp:
                    val = row['LinkKetQua']
                    if not val or pd.isna(val):
                        return "Chưa đính kèm kết quả"
                    if isinstance(val, str) and val.startswith("OUTPUT"):
                        display_name = os.path.basename(val)
                        if "_" in display_name:
                            display_name = display_name.split("_", 1)[1]
                        return f"📁 {display_name}"
                    return str(val)
                else:
                    return row['GiaiTrinhDeXuat'] if (isinstance(row['GiaiTrinhDeXuat'], str) and row['GiaiTrinhDeXuat']) else "--"
            df_display['Kết quả / File đính kèm'] = table_df.apply(format_notes, axis=1)
        
            # Reorder columns
            ordered_cols = [
                'Ngày bắt đầu',
                'Hạn chót',
                'Tiến độ',
                'Trạng thái',
                'Người thực hiện',
                'Phòng ban',
                'Dự án / Hạng mục',
                'Tên công việc',
                'Tỷ trọng KPI',
                'Nguyên nhân trễ hạn',
                'Kết quả / File đính kèm'
            ]
            df_display = df_display[ordered_cols]
        
            # Render clean st.dataframe
            st.dataframe(
                df_display,
                column_config={
                    "Ngày bắt đầu": st.column_config.TextColumn("Ngày bắt đầu", width=90),
                    "Hạn chót": st.column_config.TextColumn("Hạn chót", width=150),
                    "Tiến độ": st.column_config.ProgressColumn(
                        "Tiến độ",
                        format="%d%%",
                        min_value=0,
                        max_value=100,
                        width=100
                    ),
                    "Trạng thái": st.column_config.TextColumn("Trạng thái", width=120),
                    "Người thực hiện": st.column_config.TextColumn("Người thực hiện", width=150),
                    "Phòng ban": st.column_config.TextColumn("Phòng ban", width=80),
                    "Dự án / Hạng mục": st.column_config.TextColumn("Dự án / Hạng mục", width=200),
                    "Tên công việc": st.column_config.TextColumn("Tên công việc", width="large"),
                    "Nguyên nhân trễ hạn": st.column_config.TextColumn("Nguyên nhân trễ hạn", width=150),
                    "Kết quả / File đính kèm": st.column_config.LinkColumn(
                        "Kết quả / File đính kèm",
                        max_chars=300,
                        width="medium"
                    )
                },
                use_container_width=True,
                hide_index=True
            )

elif menu in ["👀 BẢNG TỔNG QUAN (View)", "📊 BẢNG TỔNG QUAN (View)"]:
    # 1. Hide Streamlit UI elements for a clean dashboard view
    st.markdown("""
        <style>
            
            
            
            .block-container {padding-top: 1rem; padding-bottom: 0rem;}
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown(f"### 📊 Bảng Tổng Quan (View) — {selected_company}")
    
    col_f1, col_f2, col_f3, col_auto = st.columns([2, 2, 2, 1])
    with col_f1:
        db_projs = list(display_df["TenDuAn"].dropna().unique()) if not display_df.empty else []
        merged_projs = get_filtered_projects(selected_company, config, db_projs)
        proj_options = ["Tất cả dự án"] + merged_projs
        sel_proj = st.selectbox("Lọc Dự án", proj_options, key="tv_proj")
    with col_f2:
        dept_options = ["Tất cả phòng ban"] + get_departments_for_company(selected_company, config)
        sel_dept = st.selectbox("Lọc Phòng ban", dept_options, key="tv_dept")
    with col_f3:
        status_options = ["Đang thực hiện", "Sắp tới hạn / Trễ hạn", "Hoàn thành", "Vướng mắc", "Tất cả trạng thái"]
        sel_status = st.selectbox("Lọc Trạng thái", status_options, key="tv_status", index=1)
    with col_auto:
        auto_refresh = st.checkbox("🔄 Auto-refresh (5p)", value=True, help="Tự động tải lại trang sau mỗi 5 phút")
        mobile_mode = st.checkbox("📱 Chế độ Điện thoại", value=False, help="Hiển thị dạng thẻ dọc để xem trên mobile")
        if auto_refresh:
            import streamlit.components.v1 as components
            components.html("""
                <script>
                    setTimeout(function(){
                        window.parent.location.reload();
                    }, 300000);
                </script>
            """, height=0, width=0)
            
    # Apply filters
    table_df = display_df.copy()
    if sel_proj != "Tất cả dự án":
        clean_proj = clean_proj_name(sel_proj)
        table_df = table_df[table_df['TenDuAn'].str.contains(clean_proj, case=False, na=False)]
    if sel_dept != "Tất cả phòng ban":
        table_df = table_df[table_df['PhongBan'] == sel_dept]
        
    def get_days_left(d):
        if pd.notna(d) and hasattr(d, 'strftime'):
            if isinstance(d, datetime):
                d = d.date()
            return (d - today).days
        return 999

    if sel_status == "Đang thực hiện":
        table_df = table_df[
            (table_df['TrangThai'] != 'Hoàn thành') & 
            (table_df['TrangThai'] != 'Có vướng mắc') & 
            (table_df['Deadline'].apply(get_days_left) > 3)
        ]
    elif sel_status == "Sắp tới hạn / Trễ hạn":
        table_df = table_df[
            (table_df['TrangThai'] != 'Hoàn thành') & 
            (table_df['Deadline'].apply(get_days_left) <= 3)
        ]
    elif sel_status == "Hoàn thành":
        table_df = table_df[table_df['TrangThai'] == 'Hoàn thành']
    elif sel_status == "Vướng mắc":
        table_df = table_df[table_df['TrangThai'] == 'Có vướng mắc']
        
    def _get_priority(row):
        st_val = str(row.get('TrangThai', ''))
        if 'Trễ hạn' in st_val or 'Vướng mắc' in st_val or '🔴' in st_val or '⚠️' in st_val:
            return 0
        return 1
        
    table_df['SortPriority'] = table_df.apply(_get_priority, axis=1)
    if 'NgayCapNhat' in table_df.columns:
        table_df = table_df.sort_values(by=['SortPriority', 'NgayCapNhat', 'ID'], ascending=[True, False, False]).reset_index(drop=True)
    else:
        table_df = table_df.sort_values(by=['SortPriority', 'ID'], ascending=[True, False]).reset_index(drop=True)
        
    if table_df.empty:
        st.info("Không có công việc nào phù hợp với bộ lọc.")
    else:
        df_display = pd.DataFrame()
        df_display['Phòng ban'] = table_df['PhongBan']
        df_display['Người thực hiện'] = table_df['NguoiChuTri']
        df_display['Dự án / Hạng mục'] = table_df['TenDuAn']
        df_display['Tên công việc'] = table_df['TenCongViec']
        df_display['Ngày bắt đầu'] = table_df['NgayBatDau'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) and isinstance(x, (date, datetime)) else (str(x) if pd.notna(x) else None))
        df_display['Tỷ trọng KPI'] = table_df.apply(lambda row: f"{int(float(str(row.get('TyTrongKPI', 0)).strip() or 0))}%" if pd.to_numeric(row.get('TyTrongKPI', 0), errors='coerce') > 0 else "Tự chia", axis=1)
        
        def format_dl(row):
            if pd.isna(row['Deadline']) or not isinstance(row['Deadline'], (date, datetime)): return ""
            prog = int(row['PhanTramHoanThanh'])
            date_str = row['Deadline'].strftime('%d/%m/%Y')
            if prog >= 100: return date_str
            days_left = (row['Deadline'] - today).days
            if days_left < 0: return f"🔴 {date_str} (Trễ {abs(days_left)} ngày)"
            elif days_left == 0: return f"⏳ {date_str} (Hạn hôm nay)"
            elif 1 <= days_left <= 3: return f"⚠️ {date_str} (Còn {days_left} ngày)"
            else: return date_str
        df_display['Hạn chót'] = table_df.apply(format_dl, axis=1)
        
        df_display['Tiến độ'] = table_df['PhanTramHoanThanh']
        
        def format_status(row):
            prog = int(row['PhanTramHoanThanh'])
            if prog >= 100: return "✅ Đã xong"
            if pd.notna(row['Deadline']) and row['Deadline'] < today: return "⚠️ Trễ hạn"
            if row['TrangThai'] == 'Có vướng mắc': return "🔴 Vướng mắc"
            from datetime import date, datetime
            if prog == 0 and pd.notna(row['NgayBatDau']) and isinstance(row['NgayBatDau'], (date, datetime)) and row['NgayBatDau'] > today: return "❌ Chưa bắt đầu"
            if pd.notna(row['NgayBatDau']) and isinstance(row['NgayBatDau'], (date, datetime)):
                if today >= row['NgayBatDau']: return "⏳ Đang thực hiện"
            return "❌ Chưa bắt đầu"
        df_display['Trạng thái'] = table_df.apply(format_status, axis=1)
        
        ordered_cols = ['Ngày bắt đầu', 'Hạn chót', 'Tiến độ', 'Trạng thái', 'Người thực hiện', 'Phòng ban', 'Dự án / Hạng mục', 'Tên công việc']
        df_display = df_display[ordered_cols]
        
        st.markdown("---")
        if mobile_mode:
            for idx, row in df_display.iterrows():
                prog = int(row['Tiến độ'])
                
                with st.container():
                    st.markdown(f"**📌 {row['Tên công việc']}**")
                    st.markdown(f"📁 *{row['Dự án / Hạng mục']}* | 👤 *{row['Người thực hiện']}*")
                    st.markdown(f"⏳ **Hạn chót:** {row['Hạn chót']} | Trạng thái: **{row['Trạng thái']}**")
                    st.caption(f"Tiến độ: {prog}%")
                    st.progress(prog)
                    st.markdown("---")
        else:
            st.dataframe(
                df_display,
                column_config={
                    "Ngày bắt đầu": st.column_config.TextColumn("Ngày bắt đầu", width=90),
                    "Hạn chót": st.column_config.TextColumn("Hạn chót", width=150),
                    "Tiến độ": st.column_config.ProgressColumn("Tiến độ", format="%d%%", min_value=0, max_value=100, width=100),
                    "Trạng thái": st.column_config.TextColumn("Trạng thái", width=120),
                    "Người thực hiện": st.column_config.TextColumn("Người thực hiện", width=150),
                    "Phòng ban": st.column_config.TextColumn("Phòng ban", width=80),
                    "Dự án / Hạng mục": st.column_config.TextColumn("Dự án / Hạng mục", width=200),
                    "Tên công việc": st.column_config.TextColumn("Tên công việc", width="large")
                },
                use_container_width=True,
                hide_index=True,
                height=700
            )

# ----------------- 3. THÊM / CẬP NHẬT CÔNG VIỆC -----------------


elif menu == "➕ Thêm / Cập Nhật Công Việc":
    st.markdown("### ✏️ Phân hệ Thêm / Cập Nhật Công Việc")
    
    tab_new, tab_update = st.tabs(["➕ Khởi tạo công việc mới", "✏️ Cập nhật tiến độ công việc"])
    
    # Form: Add New
    with tab_new:
        st.markdown("#### Thêm mới công việc tự do")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # 1. Company selection
            company_list = list(COMPANIES.keys())
            default_company_idx = 0
            if selected_company in company_list:
                default_company_idx = company_list.index(selected_company)
            entry_company = st.selectbox(
                "Đơn vị / Công ty thành viên", 
                company_list, 
                index=default_company_idx,
                format_func=lambda x: str(x).replace("CTY CP", "CÔNG TY CP")
            )
            
            # 4. Department
            allowed_depts = get_departments_for_company(entry_company, config)
            is_personal = (role_mode == "Cá nhân (Thử nghiệm)" and st.session_state.is_personal_authenticated and st.session_state.personal_user)
            if is_personal:
                # Deduce their department from their existing tasks or default to first
                user_dept_mode = display_df['PhongBan'].mode()
                user_dept = user_dept_mode[0] if not user_dept_mode.empty else allowed_depts[0]
                task_dept = st.selectbox("Phòng ban chịu trách nhiệm", [user_dept], index=0, disabled=True)
            else:
                task_dept = st.selectbox("Phòng ban chịu trách nhiệm", allowed_depts)
            
            # 5. Owner (based on configuration with custom type option)
            dept_personnel = get_personnel_for_company_dept(entry_company, task_dept, config)
            owner_options = list(dept_personnel) + ["✍️ Nhập tên người khác..."]
            
            # Find default lead index if present in department personnel
            dept_lead = DEPT_LEADS.get(entry_company, {}).get(task_dept, "")
            default_lead_idx = 0
            if dept_lead in dept_personnel:
                default_lead_idx = dept_personnel.index(dept_lead)
            
            is_personal = (role_mode == "Cá nhân (Thử nghiệm)" and st.session_state.is_personal_authenticated and st.session_state.personal_user)
            if is_personal:
                sel_owner_opt = st.selectbox("Người thực hiện / Phụ trách", [st.session_state.personal_user], index=0, disabled=True)
                task_owner = st.session_state.personal_user
            else:
                sel_owner_opt = st.selectbox("Người thực hiện / Phụ trách", owner_options, index=default_lead_idx)
                if sel_owner_opt == "✍️ Nhập tên người khác...":
                    task_owner = st.text_input("✍️ Nhập tên người thực hiện khác...", value="")
                else:
                    task_owner = sel_owner_opt
            
            # 2. Project selection (Categorized dropdown or custom)
            is_marina_co = "CTY CP DMT - MARINA" in entry_company or "Du thuyền Happy Yacht" in entry_company
            if False:
                proj_options_with_custom = ["➕ Tạo / Nhập Dự án mới..."]
            else:
                db_projs = list(display_df["TenDuAn"].dropna().unique()) if not display_df.empty else []
                merged_projs = get_filtered_projects(entry_company, config, db_projs)
                proj_options_with_custom = merged_projs + ["✍️ Tự nhập Dự án / Hạng mục khác..."]
                
            default_proj_opt = st.selectbox("Dự án / Hạng mục", proj_options_with_custom)
            
            if is_marina_co or default_proj_opt in ["✍️ Tự nhập Dự án / Hạng mục khác...", "➕ Tạo / Nhập Dự án mới..."]:
                project_name = st.text_input("Nhập tên Dự án / Hạng mục mới", value="")
            else:
                project_name = clean_proj_name(default_proj_opt)
            
            # 3. Task details
            task_name = st.text_input("Tên công việc (tự nhập tự do)", value="")
            task_nguon = st.selectbox("Nguồn giao việc", ["Công việc được giao / định kì", 'CV giao ban / VB đến'])
            st.caption("💡 **Định kỳ:** Đăng ký đầu tháng / quản lý giao. **Giao ban:** Phát sinh sau khi họp giao ban.")
            
        with col2:
            # 6. Dates
            task_start = st.date_input("Ngày bắt đầu thực hiện", today, format="DD/MM/YYYY")
            task_deadline = st.date_input("Hạn hoàn thành (Deadline)", today, format="DD/MM/YYYY")
            
            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander("🔽 Tùy chọn nâng cao: Ghi nhận trạng thái / Nộp kết quả ngay", expanded=False):
                with st.container():
                    st.markdown("<div style='padding: 15px; border-radius: 8px; border: 1px dashed #ccc; background-color: #f9f9f9; margin-bottom: 20px;'>", unsafe_allow_html=True)
                    # 7 & 8. Status radio
                    st.markdown("<p style='font-size: 1.1rem; font-weight: 600; color: #1e3a8a; margin-top: 0;'>📌 Trạng thái công việc</p>", unsafe_allow_html=True)
                    is_late_for_status = (task_deadline < today)
                    if is_late_for_status:
                        status_opts = ["✅ Xác nhận ĐÃ HOÀN THÀNH công việc", "⚠️ Công việc CHƯA HOÀN THÀNH, đang VƯỚNG MẮC"]
                    else:
                        status_opts = ["✅ Xác nhận ĐÃ HOÀN THÀNH công việc"]
                    task_status_choice = st.radio("Trạng thái công việc", status_opts, index=None, label_visibility="collapsed", key="new_status_choice")
                    task_is_completed = (task_status_choice == "✅ Xác nhận ĐÃ HOÀN THÀNH công việc")
                    task_has_issue = (task_status_choice == "⚠️ Công việc CHƯA HOÀN THÀNH, đang VƯỚNG MẮC")
                    
                    if task_status_choice is not None:
                        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
                        # 10. Ghi chú vướng mắc
                        is_late = (task_deadline < today) and not task_is_completed
                        task_late_cause = "🟢 Không trễ hạn / Đúng tiến độ"
                        if is_late:
                            st.markdown("**⚠️ Phân loại nguyên nhân trễ hạn**")
                            task_late_cause = st.radio(
                                "Phân loại nguyên nhân trễ hạn",
                                ["🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)", "👤 Do chủ quan"],
                                index=0,
                                label_visibility="collapsed",
                                key="new_task_late_cause"
                            )
                        if is_late or task_has_issue:
                            task_explain = st.text_area("📝 Chi tiết vướng mắc / Giải trình nguyên nhân (Bắt buộc)", placeholder="Mô tả chi tiết nguyên nhân trễ hạn hoặc vướng mắc gặp phải...", height=120, key="new_task_explain")
                            if is_late and task_late_cause == "🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)":
                                st.caption("💡 **Lưu ý:** Giải trình này sẽ được hệ thống gửi đến Quản lý để xem xét mức độ ghi nhận KPI.")
                        else:
                            task_explain = ""
                        
                        # 9. Kết quả / File đính kèm
                        if task_has_issue:
                            task_file = None
                            task_link_text = ""
                            result_mode = "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)"
                        else:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if task_is_completed:
                                st.markdown("🚨 **<span style='color:red; font-size: 17px;'>ĐỂ XÁC NHẬN HOÀN THÀNH, BẮT BUỘC NHẬP BÁO CÁO HOẶC TẢI FILE DƯỚI ĐÂY:</span>**", unsafe_allow_html=True)
                            else:
                                st.markdown("**Kết quả / File đính kèm**")
                            result_mode = st.radio("Hình thức nộp kết quả", ["✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)", "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)"], horizontal=True, key="new_result_mode")
                            if result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)":
                                if task_is_completed:
                                    st.warning("⚠️ **VUI LÒNG NHẬP NỘI DUNG KẾT QUẢ / BÁO CÁO VÀO Ô BÊN DƯỚI:**")
                                else:
                                    st.info("💡 **Ghi chú nội dung/tiến độ công việc vào ô bên dưới:**")
                                task_link_text = st.text_area("Nhập tên Báo cáo / Số hiệu Văn bản / Link", height=100, label_visibility="collapsed", placeholder="Ví dụ: Báo cáo số 01/BC-DMT, đã trình sếp, hoặc dán link Google Drive...", key="new_result_text")
                                task_file = None
                            else:
                                task_file = st.file_uploader("Tải file đính kèm (PDF, Word, Excel, Ảnh...)", key="new_result_file")
                                task_link_text = ""
                    else:
                        task_late_cause = "🟢 Không trễ hạn / Đúng tiến độ"
                        task_explain = ""
                        result_mode = "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)"
                        task_link_text = ""
                        task_file = None
                        
                    st.markdown("</div>", unsafe_allow_html=True)
                

            # 11. Chu kỳ theo dõi
            task_cycle = "Theo dự án / Tự do"
            
            # 12. Tỷ trọng KPI
            task_weight = 0
            
            
        submit_new = st.button("💾 Lưu", type="primary", key="btn_save_new_task")
        
        if submit_new:
            if not task_name.strip():
                st.error("⚠️ Vui lòng nhập Tên công việc!")
            elif not task_owner.strip():
                st.error("⚠️ Vui lòng nhập Người thực hiện!")
            else:
                # Calculate status and progress automatically
                if task_is_completed:
                    calc_status = "Hoàn thành"
                elif task_has_issue:
                    calc_status = "Có vướng mắc"
                elif task_deadline < today:
                    calc_status = "Quá hạn"
                elif today >= task_start:
                    calc_status = "Đang thực hiện"
                else:
                    calc_status = "Chưa bắt đầu"
                    
                task_progress = calculate_time_progress(task_start, task_deadline, task_is_completed)
                
                is_late = (task_deadline < today and not task_is_completed)
                    
                # Constraints validation
                has_error = False
                if calc_status == "Hoàn thành":
                    if result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)" and not task_link_text.strip():
                        st.error("⚠️ Bắt buộc điền 'Kết quả / File đính kèm'!")
                        has_error = True
                    elif result_mode == "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)" and task_file is None:
                        st.error("⚠️ Bắt buộc tải file đính kèm!")
                        has_error = True
                        
                if is_late:
                    if task_late_cause == "🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)":
                        if not task_explain.strip() or len(task_explain.strip()) < 5:
                            st.error("⚠️ Bắt buộc nhập 'Chi tiết nguyên nhân khách quan & Đề xuất phương án xử lý' (tối thiểu 5 ký tự)!")
                            has_error = True
                elif calc_status == "Có vướng mắc":
                    if not task_explain.strip() or len(task_explain.strip()) < 5:
                        st.error("⚠️ Bắt buộc nhập 'Chi tiết vướng mắc & Đề xuất hỗ trợ'!")
                        has_error = True
                        
                if not has_error:
                    with acquire_db_lock():
                        
                        fresh_df = read_db()
                        
                        # Kiểm tra trùng lặp để cảnh báo người dùng (tránh click đúp)
                        is_duplicate = False
                        if not fresh_df.empty:
                            dup_df = fresh_df[
                                (fresh_df['TenCongViec'].astype(str).str.strip() == task_name.strip()) & 
                                (fresh_df['NguoiChuTri'].astype(str).str.strip() == task_owner.strip()) & 
                                (fresh_df['TenDuAn'].astype(str).str.strip() == project_name) &
                                (fresh_df['Deadline'].astype(str) == str(task_deadline)) &
                                (fresh_df['NgayBatDau'].astype(str) == str(task_start))
                            ]
                            if not dup_df.empty:
                                is_duplicate = True
                                
                        if is_duplicate:
                            st.error("⚠️ Công việc này đã tồn tại (trùng Tên công việc, Người thực hiện, Dự án và Thời gian)! Để tránh trùng lặp do bấm nhầm, hệ thống đã chặn lại. Nếu bạn thực sự muốn tạo 1 công việc giống hệt, vui lòng sửa lại Tên công việc (ví dụ: thêm số 2 vào cuối).")
                        else:
                            # Auto ID generator
                            next_id = 1
                            if not fresh_df.empty:
                                ids = fresh_df['ID'].tolist()
                                nums = [int(m[0]) for idx in ids for m in [re.findall(r'\d+', str(idx))] if m]
                                if nums:
                                    next_id = max(nums) + 1
                            task_id = f"TSK-{next_id:03d}"
                            
                            saved_result = ""
                            
                            if result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)":
                                saved_result = task_link_text.strip()
                            elif result_mode == "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)" and task_file is not None:
                                    upload_dir = os.path.join("OUTPUT", "UPLOADED_FILES")
                                    if not os.path.exists(upload_dir):
                                        os.makedirs(upload_dir, exist_ok=True)
                                    safe_name = re.sub(r'[^\w\-_.]', '_', task_file.name)
                                    file_name = f"{task_id}_{safe_name}"
                                    file_path = os.path.join(upload_dir, file_name)
                                    with open(file_path, "wb") as f:
                                        f.write(task_file.getbuffer())
                                    saved_result = file_path
                                    
                            new_row = {
                                "ID": task_id,
                                "DonVi": entry_company,
                                "PhongBan": task_dept,
                                "NguoiChuTri": task_owner.strip(),
                                "TenDuAn": project_name,
                                "MocTienDo": "Tự do",
                                "SanPhamBanGiao": "Xem chi tiết",
                                "TenCongViec": task_name.strip(),
                                "PhanLoaiChiSo": "Chỉ số kết quả (Outcome Metric)",
                                "NgayBatDau": task_start,
                                "Deadline": task_deadline,
                                "DoUuTien": "Trung bình",
                                "PhanTramHoanThanh": task_progress,
                                "TrangThai": calc_status,
                                "LinkKetQua": saved_result,
                                "GiaiTrinhDeXuat": task_explain.strip(),
                                "NgayCapNhat": (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S'),
                                "ChuKyTheoDoi": task_cycle,
                                "PhanLoaiTreHan": task_late_cause if is_late else "🟢 Không trễ hạn / Đúng tiến độ",
                                "TyTrongKPI": task_weight,
                                "NguonGiaoViec": task_nguon,
                                "MucDoGhiNhan": "0% (Không ghi nhận)"
                            }
                            
                            df_updated = pd.concat([fresh_df, pd.DataFrame([new_row])], ignore_index=True)
                            if save_db(df_updated):
                                st.session_state["success_msg"] = f"🎉 Đã khởi tạo thành công công việc mã: {task_id}!"
                                st.rerun()


    # Form: Update Progress
    with tab_update:
        st.markdown("#### Cập nhật tiến độ công việc đang chạy")
        
        # Display only items matching selected company
        avail_update_df = display_df.copy()
        if 'NgayCapNhat' in avail_update_df.columns:
            avail_update_df = avail_update_df.sort_values(by=['NgayCapNhat', 'ID'], ascending=[False, False]).reset_index(drop=True)
        
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            departments = ["Tất cả"] + sorted(list(avail_update_df['PhongBan'].dropna().astype(str).unique()))
            filter_dept = st.selectbox("Lọc theo Phòng ban", departments, key="filter_dept_update")
        with col_f2:
            owners = ["Tất cả"] + sorted(list(avail_update_df['NguoiChuTri'].dropna().astype(str).unique()))
            filter_owner = st.selectbox("Lọc theo Người phụ trách", owners, key="filter_owner_update")
        with col_f3:
            projects = ["Tất cả"] + sorted(list(avail_update_df['TenDuAn'].dropna().astype(str).unique()))
            filter_proj = st.selectbox("Lọc theo Dự án", projects, key="filter_proj_update")
            
        if filter_dept != "Tất cả":
            avail_update_df = avail_update_df[avail_update_df['PhongBan'] == filter_dept]
        if filter_owner != "Tất cả":
            avail_update_df = avail_update_df[avail_update_df['NguoiChuTri'] == filter_owner]
        if filter_proj != "Tất cả":
            avail_update_df = avail_update_df[avail_update_df['TenDuAn'] == filter_proj]

        if avail_update_df.empty:
            st.info("Chưa có công việc nào khả dụng.")
        else:
            def format_task_option(task_id):
                row = df[df['ID'] == task_id].iloc[0]
                pic = row.get('NguoiChuTri', 'Chưa rõ')
                return f"{row['TenCongViec']} - Phụ trách: {pic}"
            
            selected_id = st.selectbox("Chọn công việc cần cập nhật", avail_update_df['ID'].tolist(), format_func=format_task_option)
            task_data = df[df['ID'] == selected_id].iloc[0]
            
            with st.container():
                col_u1, col_u2 = st.columns(2)
                
                with col_u1:
                    st.markdown(f"**Mã Hạng mục:** `{task_data['ID']}`")
                    st.markdown(f"**Đơn vị:** {task_data['DonVi']}")
                    st.markdown(f"**Phòng ban phụ trách:** {task_data['PhongBan']}")
                    
                    u_proj = st.text_input("Dự án / Hạng mục", value=task_data['TenDuAn'], key=f"u_proj_{task_data['ID']}")
                    u_name = st.text_input("Tên công việc", value=task_data['TenCongViec'], key=f"u_name_{task_data['ID']}")
                    u_nguon_opts = ["Công việc được giao / định kì", 'CV giao ban / VB đến']
                    current_nguon = task_data.get('NguonGiaoViec', 'Công việc được giao / định kì')
                    u_nguon_idx = u_nguon_opts.index(current_nguon) if current_nguon in u_nguon_opts else 0
                    u_nguon = st.selectbox("Nguồn giao việc", u_nguon_opts, index=u_nguon_idx, key=f"u_nguon_{task_data['ID']}")
                    st.caption("💡 **Định kỳ:** Đăng ký đầu tháng. **Giao ban:** Phát sinh sau khi họp giao ban.")
                    
                    # Owner selection based on configuration
                    u_dept = task_data['PhongBan']
                    u_dept_personnel = get_personnel_for_company_dept(task_data['DonVi'], u_dept, config)
                    u_owner_options = list(u_dept_personnel) + ["✍️ Nhập tên người khác..."]
                    
                    current_owner = task_data['NguoiChuTri']
                    is_personal = (role_mode == "Cá nhân (Thử nghiệm)" and st.session_state.is_personal_authenticated and st.session_state.personal_user)
                    if is_personal:
                        st.selectbox("Người thực hiện / Phụ trách", [st.session_state.personal_user], index=0, disabled=True, key=f"u_owner_sel_{task_data['ID']}")
                        u_owner = st.session_state.personal_user
                    else:
                        if current_owner in u_dept_personnel:
                            u_default_index = u_dept_personnel.index(current_owner)
                            u_sel_owner_opt = st.selectbox("Người thực hiện / Phụ trách", u_owner_options, index=u_default_index, key=f"u_owner_sel_{task_data['ID']}")
                            if u_sel_owner_opt == "✍️ Nhập tên người khác...":
                                u_owner = st.text_input("✍️ Nhập tên người thực hiện khác...", value="", key=f"u_owner_custom_{task_data['ID']}")
                            else:
                                u_owner = u_sel_owner_opt
                        else:
                            u_default_index = len(u_owner_options) - 1
                            u_sel_owner_opt = st.selectbox("Người thực hiện / Phụ trách", u_owner_options, index=u_default_index, key=f"u_owner_sel_{task_data['ID']}")
                            u_owner = st.text_input("✍️ Nhập tên người thực hiện khác...", value=current_owner, key=f"u_owner_custom_{task_data['ID']}")
                    
                with col_u2:
                    u_start = st.date_input("Ngày bắt đầu thực hiện", value=pd.to_datetime(task_data['NgayBatDau']).date() if pd.notna(task_data['NgayBatDau']) and str(task_data['NgayBatDau']).strip() else today, format="DD/MM/YYYY", key=f"u_start_{task_data['ID']}")
                    u_deadline = st.date_input("Hạn hoàn thành (Deadline)", value=pd.to_datetime(task_data['Deadline']).date() if pd.notna(task_data['Deadline']) and str(task_data['Deadline']).strip() else today, format="DD/MM/YYYY", key=f"u_deadline_{task_data['ID']}")
                    
                    st.markdown("<p style='font-size: 1.1rem; font-weight: 600; color: #1e3a8a;'>📌 Trạng thái công việc</p>", unsafe_allow_html=True)
                    u_current_status = task_data.get('TrangThai', 'Đang thực hiện')
                    u_is_late_for_status = (u_deadline is not None and u_deadline < today)
                    if u_is_late_for_status or u_current_status == 'Có vướng mắc':
                        u_status_opts = ["✅ Xác nhận ĐÃ HOÀN THÀNH công việc", "⚠️ Công việc CHƯA HOÀN THÀNH, đang VƯỚNG MẮC"]
                    else:
                        u_status_opts = ["✅ Xác nhận ĐÃ HOÀN THÀNH công việc"]
                    
                    if u_current_status == 'Hoàn thành':
                        u_status_idx = 0
                    elif u_current_status == 'Có vướng mắc':
                        u_status_idx = 1
                    else:
                        u_status_idx = None
                        
                    u_status_choice = st.radio("Trạng thái công việc", u_status_opts, index=u_status_idx, label_visibility="collapsed", key=f"u_status_choice_{task_data['ID']}")
                    u_is_completed = (u_status_choice == "✅ Xác nhận ĐÃ HOÀN THÀNH công việc")
                    u_has_issue = (u_status_choice == "⚠️ Công việc CHƯA HOÀN THÀNH, đang VƯỚNG MẮC")
                    
                    # 11. Chu kỳ theo dõi
                    current_cycle = task_data.get('ChuKyTheoDoi', 'Theo dự án / Tự do')
                    cycle_list = ["Hàng tuần", "Hàng tháng", "Hàng quý", "Theo dự án / Tự do"]
                    default_cycle_idx = cycle_list.index(current_cycle) if current_cycle in cycle_list else 3
                    u_cycle = current_cycle
                    
                    u_weight = task_data.get('TyTrongKPI', '')
                    
                
                u_is_late = (u_deadline is not None and u_deadline < today) and not u_is_completed
                if u_status_choice is not None:
                    st.markdown("<div style='padding: 15px; border-radius: 8px; border: 1px dashed #ccc; background-color: #f9f9f9; margin-top: 15px;'>", unsafe_allow_html=True)
                    u_late_cause = "🟢 Không trễ hạn / Đúng tiến độ"
                    if u_is_late:
                        st.markdown("**⚠️ Phân loại nguyên nhân trễ hạn**")
                        u_options = ["🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)", "👤 Do chủ quan"]
                        u_current_val = task_data.get('PhanLoaiTreHan', "🟢 Không trễ hạn / Đúng tiến độ")
                        u_default_idx = u_options.index(u_current_val) if u_current_val in u_options else 0
                        u_late_cause = st.radio(
                            "Phân loại nguyên nhân trễ hạn",
                            u_options,
                            index=u_default_idx,
                            label_visibility="collapsed",
                            key=f"u_late_cause_sel_{task_data['ID']}"
                        )
                    if u_is_late or u_has_issue:
                        u_explain = st.text_area("📝 Chi tiết vướng mắc / Giải trình nguyên nhân (Bắt buộc)", value=task_data.get('GiaiTrinhDeXuat', ''), placeholder="Mô tả chi tiết nguyên nhân trễ hạn hoặc vướng mắc gặp phải...", height=120, key=f"u_explain_txt_{task_data['ID']}")
                        if u_is_late and u_late_cause == "🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)":
                            if st.session_state.is_admin_authenticated or st.session_state.get('is_manager_authenticated', False):
                                current_chamchuoc = task_data.get('MucDoGhiNhan', '0% (Không ghi nhận)')
                                chamchuoc_opts = ["0% (Không ghi nhận)", "Miễn trừ (Loại bỏ KPI)", "50%", "80%", "90%"]
                                idx_cc = chamchuoc_opts.index(current_chamchuoc) if current_chamchuoc in chamchuoc_opts else 0
                                u_chamchuoc = st.selectbox("Mức độ ghi nhận (Dành cho Quản lý)", chamchuoc_opts, index=idx_cc, key=f"u_cc_{task_data['ID']}")
                            else:
                                current_chamchuoc = task_data.get('MucDoGhiNhan', '0% (Không ghi nhận)')
                                u_chamchuoc = current_chamchuoc
                                if current_chamchuoc != '0% (Không ghi nhận)':
                                    st.info(f"Đã được Quản lý ghi nhận mức độ KPI: **{current_chamchuoc}**")
                                else:
                                    st.caption("💡 **Lưu ý:** Giải trình này sẽ được hệ thống gửi đến Quản lý để xem xét mức độ ghi nhận KPI.")
                        else:
                            u_chamchuoc = '0% (Không ghi nhận)'
                    else:
                        u_explain = ""
                        u_chamchuoc = '0% (Không ghi nhận)'
                    

                    current_link = task_data['LinkKetQua']
                    u_link_text = ""
                    u_file = None
                    u_result_mode = None
                    if not u_has_issue:
                        if u_is_completed:
                            st.markdown("🚨 **<span style='color:red; font-size: 17px;'>ĐỂ XÁC NHẬN HOÀN THÀNH, BẮT BUỘC NHẬP BÁO CÁO HOẶC TẢI FILE DƯỚI ĐÂY:</span>**", unsafe_allow_html=True)
                        else:
                            st.markdown("**Cập nhật Kết quả / File đính kèm**")
                            
                        u_result_mode = st.radio("Hình thức nộp kết quả", ["✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)", "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)"], horizontal=True, key=f"u_result_mode_{task_data['ID']}")
                        
                        if u_result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)":
                            if u_is_completed:
                                st.warning("⚠️ **VUI LÒNG NHẬP NỘI DUNG KẾT QUẢ / BÁO CÁO VÀO Ô BÊN DƯỚI:**")
                            else:
                                st.info("💡 **Ghi chú nội dung/tiến độ công việc vào ô bên dưới:**")
                            u_link_text = st.text_area("Nhập tên Báo cáo / Số hiệu Văn bản / Link mới", height=100, label_visibility="collapsed", placeholder="Ví dụ: Đã hoàn thành 50%, trình ký sếp...", key=f"u_result_text_{task_data['ID']}")
                        elif u_result_mode == "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)":
                            u_file = st.file_uploader("Tải file đính kèm mới", key=f"u_result_file_{task_data['ID']}")
                    st.markdown("</div>", unsafe_allow_html=True)
                else:
                    u_late_cause = "🟢 Không trễ hạn / Đúng tiến độ"
                    u_explain = ""
                    u_chamchuoc = '0% (Không ghi nhận)'
                    u_link_text = ""
                    u_file = None
                    u_result_mode = "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)"

                    

                btn_save, btn_del = st.columns([3, 2])
                
                with btn_save:
                    save_click = st.button("💾 LƯU CẬP NHẬT TIẾN ĐỘ", type="primary", key=f"btn_save_update_{task_data['ID']}")
                with btn_del:
                    del_click = False
                    if st.session_state.is_admin_authenticated:
                        confirm_del = st.checkbox("Xác nhận xóa dữ liệu này", key=f"confirm_del_{task_data['ID']}")
                        if confirm_del:
                            del_click = st.button("🗑️ XÓA CÔNG VIỆC CHỌN", type="secondary", key=f"btn_del_update_{task_data['ID']}")


                if save_click:
                    # Calculate status and progress automatically
                    if u_is_completed:
                        u_status = "Hoàn thành"
                    elif u_has_issue:
                        u_status = "Có vướng mắc"
                    elif u_deadline < today:
                        u_status = "Quá hạn"
                    elif today >= u_start:
                        u_status = "Đang thực hiện"
                    else:
                        u_status = "Chưa bắt đầu"
                        
                    u_progress = calculate_time_progress(u_start, u_deadline, u_is_completed)
                        
                    # Constraints validation
                    has_error = False
                    if u_status == "Hoàn thành":
                        if u_result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)" and not u_link_text.strip() and not current_link:
                            st.error("⚠️ Bắt buộc điền 'Kết quả / File đính kèm' để hoàn thành công việc!")
                            has_error = True
                        elif u_result_mode == "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)" and u_file is None and not current_link:
                            st.error("⚠️ Bắt buộc tải file đính kèm để hoàn thành công việc!")
                            has_error = True
                            
                    if u_is_late:
                        if u_late_cause == "🌧️ Do khách quan (Pháp lý, Đối tác, Thời tiết, Cơ quan nhà nước...)":
                            if not u_explain.strip() or len(u_explain.strip()) < 5:
                                st.error("⚠️ Bắt buộc nhập 'Chi tiết nguyên nhân khách quan & Đề xuất phương án xử lý' (tối thiểu 5 ký tự)!")
                                has_error = True
                    elif u_status == "Có vướng mắc":
                        if not u_explain.strip() or len(u_explain.strip()) < 5:
                            st.error("⚠️ Bắt buộc nhập 'Chi tiết vướng mắc & Đề xuất hỗ trợ'!")
                            has_error = True
                            
                    if not has_error:
                        # Determine final link value
                        final_link = current_link
                        
                        if u_result_mode == "✍️ Nhập tên Báo cáo / Số hiệu Văn bản / Link (Dạng text tự do)":
                            if u_link_text.strip():
                                final_link = u_link_text.strip()
                        elif u_result_mode == "📁 Tải file đính kèm (PDF, Word, Excel, Ảnh...)" and u_file is not None:
                                upload_dir = os.path.join("OUTPUT", "UPLOADED_FILES")
                                if not os.path.exists(upload_dir):
                                    os.makedirs(upload_dir, exist_ok=True)
                                safe_name = re.sub(r'[^\w\-_.]', '_', u_file.name)
                                file_name = f"{selected_id}_{safe_name}"
                                file_path = os.path.join(upload_dir, file_name)
                                with open(file_path, "wb") as f:
                                    f.write(u_file.getbuffer())
                                final_link = file_path
                                
                        with acquire_db_lock():
                            
                            fresh_df = read_db()
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'TenDuAn'] = u_proj.strip()
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'TenCongViec'] = u_name.strip()
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'NguoiChuTri'] = u_owner.strip()
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'NgayBatDau'] = u_start
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'Deadline'] = u_deadline
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'PhanTramHoanThanh'] = u_progress
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'TrangThai'] = u_status
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'LinkKetQua'] = final_link
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'GiaiTrinhDeXuat'] = u_explain.strip()
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'NgayCapNhat'] = (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'ChuKyTheoDoi'] = u_cycle
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'PhanLoaiTreHan'] = u_late_cause if u_is_late else "🟢 Không trễ hạn / Đúng tiến độ"
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'TyTrongKPI'] = str(u_weight)
                            fresh_df.loc[fresh_df['ID'] == selected_id, 'NguonGiaoViec'] = u_nguon
                            if u_is_late:
                                fresh_df.loc[fresh_df['ID'] == selected_id, 'MucDoGhiNhan'] = u_chamchuoc
                            else:
                                fresh_df.loc[fresh_df['ID'] == selected_id, 'MucDoGhiNhan'] = '0% (Không ghi nhận)'

                            if save_db(fresh_df):
                                st.session_state["success_msg"] = f"🎉 Đã lưu cập nhật công việc mã: {selected_id}!"
                                st.rerun()
                            
                    if has_error:
                        st.session_state.is_updating_task = False
                            
                if del_click:
                    with acquire_db_lock():
                        
                        fresh_df = read_db()
                        df_after_del = fresh_df[fresh_df['ID'] != selected_id]
                        if save_db(df_after_del):
                            st.session_state["success_msg"] = f"🗑️ Đã xóa thành công công việc mã: {selected_id}!"
                            st.rerun()

                st.markdown("---")
                with st.expander("🔄 Tái tạo công việc định kỳ (Nhân bản cho kỳ sau)"):
                    st.info("Tính năng này giúp nhân bản công việc hiện tại thành một công việc mới cho kỳ tiếp theo (dành cho các báo cáo tuần, giao ban tháng...).")
                    
                    rep_name = st.text_input("Tên công việc mới", value=f"{task_data['TenCongViec']} (Kỳ tiếp theo)", key=f"rep_name_{task_data['ID']}")
                    
                    try:
                        from dateutil.relativedelta import relativedelta
                        import pandas as pd
                        
                        # Fix for cases where NgayBatDau or Deadline might be NaT or None
                        if pd.notna(task_data.get('NgayBatDau')):
                            default_start = task_data['NgayBatDau'] + relativedelta(months=1)
                        else:
                            default_start = today
                            
                        if pd.notna(task_data.get('Deadline')):
                            default_deadline = task_data['Deadline'] + relativedelta(months=1)
                        else:
                            default_deadline = default_start + timedelta(days=6)
                    except Exception as e:
                        default_start = task_data['Deadline'] + timedelta(days=1) if pd.notna(task_data.get('Deadline')) else today
                        default_deadline = default_start + timedelta(days=6)
                    
                    col_rep1, col_rep2 = st.columns(2)
                    with col_rep1:
                        rep_start = st.date_input("Ngày bắt đầu mới", value=default_start, format="DD/MM/YYYY", key=f"rep_start_{task_data['ID']}")
                    with col_rep2:
                        rep_deadline = st.date_input("Hạn chót mới", value=default_deadline, format="DD/MM/YYYY", key=f"rep_deadline_{task_data['ID']}")
                        
                    if st.button("🔄 TẠO CÔNG VIỆC CHO KỲ SAU", type="primary", key=f"btn_rep_{task_data['ID']}"):
                        with acquire_db_lock():
                            
                            fresh_df = read_db()
                            
                            next_id = 1
                            if not fresh_df.empty:
                                ids = fresh_df['ID'].tolist()
                                import re
                                nums = [int(m[0]) for idx in ids for m in [re.findall(r'\d+', str(idx))] if m]
                                if nums:
                                    next_id = max(nums) + 1
                            new_id = f"TSK-{next_id:03d}"
                            
                            new_row = {
                                "ID": new_id,
                                "DonVi": task_data['DonVi'],
                                "PhongBan": task_data['PhongBan'],
                                "NguoiChuTri": task_data['NguoiChuTri'],
                                "TenDuAn": task_data['TenDuAn'],
                                "MocTienDo": "Tự do",
                                "SanPhamBanGiao": "Xem chi tiết",
                                "TenCongViec": rep_name.strip(),
                                "PhanLoaiChiSo": "Chỉ số kết quả (Outcome Metric)",
                                "NgayBatDau": rep_start,
                                "Deadline": rep_deadline,
                                "DoUuTien": "Trung bình",
                                "PhanTramHoanThanh": 0,
                                "TrangThai": "Chưa bắt đầu" if today < rep_start else "Đang thực hiện",
                                "LinkKetQua": "",
                                "GiaiTrinhDeXuat": "",
                                "NgayCapNhat": (datetime.utcnow() + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S'),
                                "ChuKyTheoDoi": task_data['ChuKyTheoDoi'],
                                "PhanLoaiTreHan": "🟢 Không trễ hạn / Đúng tiến độ"
                            }
                            
                            df_rep = pd.concat([fresh_df, pd.DataFrame([new_row])], ignore_index=True)
                            if save_db(df_rep):
                                st.session_state["success_msg"] = f"🎉 Đã nhân bản thành công công việc mới mã: {new_id}!"
                                st.rerun()

# ----------------- 5. ĐÁNH GIÁ KPI & XẾP LOẠI -----------------
elif menu == "🏆 Đánh giá KPI & Xếp loại":
    st.markdown(f"### 🏆 Đánh giá KPI & Xếp loại Cá nhân — {selected_company}")
    
    if "success_msg" in st.session_state:
        st.success(st.session_state["success_msg"])
        del st.session_state["success_msg"]
    
    if role_mode in ["Quản lý", "HR"] and st.session_state.is_admin_authenticated:
        kpi_tab1, kpi_tab2, kpi_tab3, kpi_tab4 = st.tabs(["📅 Đánh giá theo Tháng", "🏅 Tổng kết KPI Cả Năm (Tháng 13)", "⚖️ Thưởng / Phạt Điểm", "📈 Phân tích & Xuất Báo cáo"])
    elif role_mode == "Quản lý" and st.session_state.get('is_manager_authenticated', False):
        kpi_tab1, = st.tabs(["📅 Đánh giá theo Tháng"])
    else:
        kpi_tab1, kpi_tab2 = st.tabs(["📅 Đánh giá theo Tháng", "🏅 Tổng kết KPI Cả Năm (Tháng 13)"])
    
    with kpi_tab1:
        st.markdown("#### Đánh giá và Xếp loại KPI Tháng")
        
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            selected_month = st.selectbox("Chọn Tháng", list(range(1, 13)), index=today.month - 1)
        with col_m2:
            selected_year = st.selectbox("Chọn Năm", [today.year - 1, today.year, today.year + 1], index=1)
        with col_m3:
            allowed_depts_m = get_departments_for_company(selected_company, config)
            dept_options_m = ["Tất cả phòng ban"] + allowed_depts_m
            selected_dept_m = st.selectbox("Lọc theo Phòng ban", dept_options_m, key="kpi_m_dept")
            
        kpi_df = display_df.copy()
        if 'NguoiChuTri' not in kpi_df.columns:
            kpi_df['NguoiChuTri'] = ''
        
        def is_in_month(d, m, y):
            import pandas as pd
            from datetime import datetime, date
            if pd.isna(d): return False
            if isinstance(d, str):
                try: d = datetime.strptime(d, "%Y-%m-%d").date()
                except: return False
            if isinstance(d, datetime): d = d.date()
            if isinstance(d, date): return d.month == m and d.year == y
            return False
            
        kpi_df = kpi_df[kpi_df['Deadline'].apply(lambda x: is_in_month(x, selected_month, selected_year))] if not kpi_df.empty else kpi_df
        
        adj_df = read_kpi_adjustments()
        adj_df = adj_df[(adj_df['Thang'] == selected_month) & (adj_df['Nam'] == selected_year)]
        
        if kpi_df.empty and adj_df.empty:
            st.info(f"Không có dữ liệu công việc hoặc điểm thưởng/phạt nào trong Tháng {selected_month}/{selected_year}.")
        else:
            import pandas as pd
            from datetime import datetime, date
            personnel_kpi = []
            
            # Find all personnel relevant to THIS company
            company_personnel = set(display_df['NguoiChuTri'].dropna().unique())
            
            all_p = set(kpi_df['NguoiChuTri'].dropna().unique())
            if 'TenNhanVien' in adj_df.columns:
                all_p.update(adj_df['TenNhanVien'].dropna().unique())
            
            # Filter to only keep those who belong to the selected company
            all_p = all_p.intersection(company_personnel)
            
            for person in all_p:
                if not str(person).strip(): continue
                group = kpi_df[kpi_df['NguoiChuTri'] == person]
                total_tasks = len(group)
                done_tasks = len(group[group['TrangThai'] == 'Hoàn thành'])
                
                group_copy = group.copy()
                group_copy['TyTrongKPI'] = pd.to_numeric(group_copy.get('TyTrongKPI', pd.Series(0, index=group_copy.index)), errors='coerce').fillna(0)
                
                for idx, row in group_copy.iterrows():
                    is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
                    is_late = False
                    dl = row['Deadline']
                    if isinstance(dl, str):
                        try: dl = datetime.strptime(dl, "%Y-%m-%d").date()
                        except: pass
                    if isinstance(dl, datetime): dl = dl.date()
                    if isinstance(dl, date): is_late = (dl < today) and not is_comp
                    
                    if is_late and row.get('PhanLoaiTreHan') == "🌍 Do khách quan":
                        group_copy.at[idx, 'PhanTramHoanThanh'] = 100
                        
                explicit_weight_sum = group_copy[group_copy['TyTrongKPI'] > 0]['TyTrongKPI'].sum()
                unweighted_count = len(group_copy[group_copy['TyTrongKPI'] <= 0])
                
                remaining_weight = max(0, 100 - explicit_weight_sum)
                if (selected_year > 2026) or (selected_year == 2026 and selected_month >= 8):
                    auto_weight = 0
                else:
                    auto_weight = remaining_weight / unweighted_count if unweighted_count > 0 else 0
                
                # Calculate score dynamically based on NguonGiaoViec (70/30 rule)
                if 'NguonGiaoViec' not in group_copy.columns:
                    group_copy['NguonGiaoViec'] = 'Công việc được giao / định kì'
                ke_hoach_tasks = group_copy[~group_copy['NguonGiaoViec'].isin(['Công việc trong "Giao ban"', 'CV giao ban / VB đến'])]
                giao_ban_tasks = group_copy[group_copy['NguonGiaoViec'].isin(['Công việc trong "Giao ban"', 'CV giao ban / VB đến'])]
                
                def calc_score_for_group(grp):
                    if grp.empty: return 0
                    t_score = 0
                    total_w = 0
                    for idx, row in grp.iterrows():
                        is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
                        w = row['TyTrongKPI'] if row['TyTrongKPI'] > 0 else auto_weight
                        
                        if is_comp:
                            p = 100
                        else:
                            if "khách quan" in str(row.get('PhanLoaiTreHan')).lower():
                                cc = row.get('MucDoGhiNhan', '0% (Không ghi nhận)')
                                if cc == "Miễn trừ (Loại bỏ KPI)":
                                    w = 0
                                    p = 0
                                elif cc == "50%": p = 50
                                elif cc == "80%": p = 80
                                elif cc == "90%": p = 90
                                else: p = 0
                            else:
                                p = 0
                        
                        if pd.isna(p): p = 0
                        t_score += (p / 100.0) * w
                        total_w += w
                        
                    if total_w > 0:
                        return (t_score / total_w) * 100
                    return 0

                if len(giao_ban_tasks) > 0:
                    kh_score = calc_score_for_group(ke_hoach_tasks)
                    gb_score = calc_score_for_group(giao_ban_tasks)
                    task_score = kh_score * 0.7 + gb_score * 0.3
                else:
                    task_score = calc_score_for_group(ke_hoach_tasks)

                
                p_adj_df = adj_df[adj_df['TenNhanVien'] == person] if 'TenNhanVien' in adj_df.columns else pd.DataFrame()
                adj_score = 0
                if not p_adj_df.empty:
                    for _, r in p_adj_df.iterrows():
                        loai = str(r.get('LoaiHanhVi', '')).lower()
                        diem = int(r.get('DiemDieuChinh', 0))
                        if 'thưởng' in loai:
                            adj_score += diem
                        elif 'phạt' in loai:
                            adj_score -= diem
                        else:
                            adj_score += diem
                
                final_score = min(115, max(0, round(task_score + adj_score, 2)))
                
                # Xếp loại mới
                if final_score > 91: grade = "A"
                elif final_score > 81: grade = "B"
                elif final_score > 71: grade = "C"
                else:
                    if selected_year == today.year and selected_month == today.month:
                        grade = "-"
                    else:
                        grade = "D"
                
                pb = group['PhongBan'].iloc[0] if not group.empty else ""
                
                personnel_kpi.append({
                    "Người thực hiện": person,
                    "Phòng ban": DEPT_ABBR.get(pb, pb),
                    "Số việc": total_tasks,
                    "Điểm công việc": round(task_score, 1),
                    "Thưởng/Phạt": adj_score,
                    "TỔNG ĐIỂM": final_score,
                    "Xếp loại": grade,
                    "_kh_score": round(kh_score, 1) if len(giao_ban_tasks) > 0 else round(task_score, 1),
                    "_gb_score": round(gb_score, 1) if len(giao_ban_tasks) > 0 else 0,
                    "_has_gb": len(giao_ban_tasks) > 0
                })
                
            if personnel_kpi:
                kpi_month_df = pd.DataFrame(personnel_kpi)
                if selected_dept_m != "Tất cả phòng ban":
                    kpi_month_df = kpi_month_df[kpi_month_df["Phòng ban"] == DEPT_ABBR.get(selected_dept_m, selected_dept_m)]
                st.dataframe(
                    kpi_month_df[["Người thực hiện", "Phòng ban", "Số việc", "Điểm công việc", "Thưởng/Phạt", "TỔNG ĐIỂM", "Xếp loại"]],
                    column_config={
                        "TỔNG ĐIỂM": st.column_config.ProgressColumn("TỔNG ĐIỂM", format="%f", min_value=0, max_value=115),
                    },
                    use_container_width=True, hide_index=True
                )

                st.markdown("---")
                with st.expander("🔍 Tra cứu chi tiết điểm KPI của từng nhân sự", expanded=False):
                    st.info("💡 Tính năng này giúp Quản lý đối chiếu các đầu việc và mức độ hoàn thành của nhân sự để xác minh tính chính xác của điểm số.")
                    valid_people = sorted(list(kpi_month_df['Người thực hiện'].unique()))
                    if valid_people:
                        det_p = st.selectbox("👤 Chọn nhân sự cần tra cứu", valid_people, key="detail_person_kpi")
                        if det_p:
                            p_info = kpi_month_df[kpi_month_df['Người thực hiện'] == det_p].iloc[0]
                            st.markdown(f"### 🧮 Diễn giải công thức tính điểm của **{det_p}**")
                            
                            kh_val = p_info['_kh_score']
                            gb_val = p_info['_gb_score']
                            has_gb = p_info['_has_gb']
                            task_val = p_info['Điểm công việc']
                            adj_val = p_info['Thưởng/Phạt']
                            final_val = p_info['TỔNG ĐIỂM']
                            
                            st.markdown(f"**📝 Danh sách công việc của {det_p}:**")
                            p_tasks = kpi_df[kpi_df['NguoiChuTri'] == det_p].copy()
                            if p_tasks.empty:
                                st.warning("Không có đầu việc nào được ghi nhận trong tháng.")
                            else:
                                p_tasks['TyTrongKPI'] = pd.to_numeric(p_tasks.get('TyTrongKPI', pd.Series(0, index=p_tasks.index)), errors='coerce').fillna(0)
                                explicit_w = p_tasks[p_tasks['TyTrongKPI'] > 0]['TyTrongKPI'].sum()
                                unweighted = len(p_tasks[p_tasks['TyTrongKPI'] <= 0])
                                auto_w = max(0, 100 - explicit_w) / unweighted if unweighted > 0 else 0
                                
                                quy_dois = []
                                w_thuctes = []
                                kh_parts = []
                                gb_parts = []
                                kh_tw = 0.0
                                gb_tw = 0.0
                                
                                for idx, row in p_tasks.iterrows():
                                    is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
                                    w = row['TyTrongKPI'] if row['TyTrongKPI'] > 0 else auto_w
                                    if is_comp: 
                                        p = 100
                                        p_tasks.at[idx, 'MucDoGhiNhan'] = "-"
                                    elif "khách quan" in str(row.get('PhanLoaiTreHan')).lower():
                                        cc = str(row.get('MucDoGhiNhan', '0%'))
                                        if "Miễn trừ" in cc: w = 0; p = 0
                                        elif "50%" in cc: p = 50
                                        elif "80%" in cc: p = 80
                                        elif "90%" in cc: p = 90
                                        else: p = 0
                                    else: 
                                        p = 0
                                        p_tasks.at[idx, 'MucDoGhiNhan'] = "-"
                                    
                                    quy_dois.append(p)
                                    w_round = round(w, 2)
                                    w_thuctes.append(w_round)
                                    
                                    if row.get('NguonGiaoViec', '') in ['Công việc trong "Giao ban"', 'CV giao ban / VB đến']:
                                        if w_round > 0:
                                            gb_parts.append(f"({p} × {w_round}%)")
                                            gb_tw += w_round
                                    else:
                                        if w_round > 0:
                                            kh_parts.append(f"({p} × {w_round}%)")
                                            kh_tw += w_round
                                    
                                p_tasks['Tỷ trọng (Thực tế) %'] = w_thuctes
                                p_tasks['Điểm quy đổi'] = quy_dois
                                
                                kh_math = f"[{' + '.join(kh_parts)}] / {round(kh_tw,2)}%" if kh_parts else "0"
                                gb_math = f"[{' + '.join(gb_parts)}] / {round(gb_tw,2)}%" if gb_parts else "0"
                                
                                if has_gb:
                                    st.info(f"**1️⃣ Điểm Kế hoạch / Định kỳ ({kh_val}):** = {kh_math}\n\n"
                                            f"**2️⃣ Điểm Giao ban ({gb_val}):** = {gb_math}\n\n"
                                            f"**3️⃣ Điểm Thưởng/Phạt:** {adj_val}\n\n"
                                            f"👉 **TỔNG ĐIỂM ({final_val})** = (Điểm KH × 70% + Điểm GB × 30%) + Thưởng/Phạt = ({kh_val} × 0.7 + {gb_val} × 0.3) + ({adj_val})")
                                else:
                                    st.info(f"**1️⃣ Điểm Kế hoạch / Định kỳ ({kh_val}):** = {kh_math}\n\n"
                                            f"**2️⃣ Điểm Thưởng/Phạt:** {adj_val}\n\n"
                                            f"👉 **TỔNG ĐIỂM ({final_val})** = Điểm KH + Thưởng/Phạt = {kh_val} + ({adj_val})")

                                p_tasks_disp = p_tasks[['NguonGiaoViec', 'TenDuAn', 'TenCongViec', 'Deadline', 'TrangThai', 'PhanLoaiTreHan', 'MucDoGhiNhan', 'TyTrongKPI', 'Tỷ trọng (Thực tế) %', 'Điểm quy đổi']].copy()
                                p_tasks_disp['Deadline'] = pd.to_datetime(p_tasks_disp['Deadline'], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')
                                st.dataframe(p_tasks_disp, use_container_width=True, hide_index=True)
                                
                            st.markdown(f"**⚖️ Lịch sử Thưởng/Phạt của {det_p}:**")
                            p_adjs = adj_df[adj_df['TenNhanVien'] == det_p][['LoaiHanhVi', 'LyDo', 'DiemDieuChinh']] if 'TenNhanVien' in adj_df.columns else pd.DataFrame()
                            if p_adjs.empty:
                                st.success("Không có ghi nhận thưởng/phạt nào.")
                            else:
                                st.dataframe(p_adjs, use_container_width=True, hide_index=True)
            else:
                st.info("Không có dữ liệu cá nhân hợp lệ.")
    if 'kpi_tab2' in locals():
        with kpi_tab2:
            st.markdown("#### Tổng kết KPI Cả Năm & Xếp loại thưởng Tháng 13")
            col_y1, col_y2 = st.columns(2)
            with col_y1:
                selected_year_full = st.selectbox("Chọn Năm Tổng Kết", [today.year - 1, today.year, today.year + 1], index=1, key="year_full")
            with col_y2:
                allowed_depts_y = get_departments_for_company(selected_company, config)
                dept_options_y = ["Tất cả phòng ban"] + allowed_depts_y
                selected_dept_y = st.selectbox("Lọc theo Phòng ban", dept_options_y, key="kpi_y_dept")
        
            if st.button("🔄 Chạy / Cập nhật Báo cáo Tổng kết Năm", type="primary"):
                with st.spinner("Đang tính toán dữ liệu 12 tháng..."):
                    import pandas as pd
                    from datetime import datetime, date
                    all_personnel = set(display_df['NguoiChuTri'].dropna().unique())
                    adj_year_df = read_kpi_adjustments()
                    adj_year_df = adj_year_df[adj_year_df['Nam'] == selected_year_full]
                    if 'TenNhanVien' in adj_year_df.columns:
                        all_personnel.update(adj_year_df['TenNhanVien'].dropna().unique())
                
                    # Filter to only keep those who belong to the selected company
                    company_personnel = set(display_df['NguoiChuTri'].dropna().unique())
                    all_personnel = all_personnel.intersection(company_personnel)
                
                    all_personnel = list(all_personnel)
                    all_personnel = [p for p in all_personnel if str(p).strip()]
                    yearly_data = []
                    for person in all_personnel:
                        person_df = display_df[display_df['NguoiChuTri'] == person].copy()
                    
                        months_grades = {}
                        count_a = 0
                        count_b = 0
                        count_c = 0
                        count_d = 0
                    
                        for m in range(1, 13):
                            if selected_year_full > today.year or (selected_year_full == today.year and m > today.month):
                                months_grades[f"Tháng {m}"] = "-"
                                continue
                        
                            def is_in_m(d):
                                if pd.isna(d): return False
                                if isinstance(d, str):
                                    try: d = datetime.strptime(d, "%Y-%m-%d").date()
                                    except: return False
                                if isinstance(d, datetime): d = d.date()
                                if isinstance(d, date): return d.month == m and d.year == selected_year_full
                                return False
                            
                            m_df = person_df[person_df['Deadline'].apply(is_in_m)] if not person_df.empty else person_df
                            m_adj_df = adj_year_df[(adj_year_df['TenNhanVien'] == person) & (adj_year_df['Thang'] == m)] if 'TenNhanVien' in adj_year_df.columns else pd.DataFrame()
                        
                            if m_df.empty and m_adj_df.empty:
                                months_grades[f"Tháng {m}"] = "-"
                                continue
                            
                            m_df_copy = m_df.copy()
                            m_df_copy['TyTrongKPI'] = pd.to_numeric(m_df_copy.get('TyTrongKPI', pd.Series(0, index=m_df_copy.index)), errors='coerce').fillna(0)
                        
                            for idx, row in m_df_copy.iterrows():
                                is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
                                is_late = False
                                dl = row['Deadline']
                                if isinstance(dl, str):
                                    try: dl = datetime.strptime(dl, "%Y-%m-%d").date()
                                    except: pass
                                if isinstance(dl, datetime): dl = dl.date()
                                if isinstance(dl, date): is_late = (dl < today) and not is_comp
                                if is_late and row.get('PhanLoaiTreHan') == "🌍 Do khách quan":
                                    m_df_copy.at[idx, 'PhanTramHoanThanh'] = 100
                                
                            explicit_weight = m_df_copy[m_df_copy['TyTrongKPI'] > 0]['TyTrongKPI'].sum()
                            uw_count = len(m_df_copy[m_df_copy['TyTrongKPI'] <= 0])
                            if (selected_year_full > 2026) or (selected_year_full == 2026 and m >= 8):
                                auto_w = 0
                            else:
                                auto_w = max(0, 100 - explicit_weight) / uw_count if uw_count > 0 else 0
                        
                            if 'NguonGiaoViec' not in m_df_copy.columns:
                                m_df_copy['NguonGiaoViec'] = 'Công việc được giao / định kì'
                            ke_hoach_tasks_y = m_df_copy[~m_df_copy['NguonGiaoViec'].isin(['Công việc trong "Giao ban"', 'CV giao ban / VB đến'])]
                            giao_ban_tasks_y = m_df_copy[m_df_copy['NguonGiaoViec'].isin(['Công việc trong "Giao ban"', 'CV giao ban / VB đến'])]
                        
                            def calc_score_for_group_y(grp, auto_w):
                                if grp.empty: return 0
                                score = 0
                                total_w = 0
                                for idx, row in grp.iterrows():
                                    is_comp = (str(row.get('TrangThai')).strip() == 'Hoàn thành')
                                    w = row['TyTrongKPI'] if row['TyTrongKPI'] > 0 else auto_w
                                
                                    if is_comp:
                                        p = 100
                                    else:
                                        if "khách quan" in str(row.get('PhanLoaiTreHan')).lower():
                                            cc = row.get('MucDoGhiNhan', '0% (Không ghi nhận)')
                                            if cc == "Miễn trừ (Loại bỏ KPI)":
                                                w = 0
                                                p = 0
                                            elif cc == "50%": p = 50
                                            elif cc == "80%": p = 80
                                            elif cc == "90%": p = 90
                                            else: p = 0
                                        else:
                                            p = 0
                                        
                                    if pd.isna(p): p = 0
                                    score += (p / 100.0) * w
                                    total_w += w
                                
                                if total_w > 0:
                                    return (score / total_w) * 100
                                return 0
                            
                            if len(giao_ban_tasks_y) > 0:
                                kh_score_y = calc_score_for_group_y(ke_hoach_tasks_y, auto_w)
                                gb_score_y = calc_score_for_group_y(giao_ban_tasks_y, auto_w)
                                t_score = kh_score_y * 0.7 + gb_score_y * 0.3
                            else:
                                t_score = calc_score_for_group_y(ke_hoach_tasks_y, auto_w)
                        
                            f_score = min(115, max(0, round(t_score + m_adj_df['DiemDieuChinh'].sum(), 2)))
                        
                            if f_score > 91: 
                                grade = "A"
                                count_a += 1
                            elif f_score > 81: 
                                grade = "B"
                                count_b += 1
                            elif f_score > 71:
                                grade = "C"
                                count_c += 1
                            else:
                                if selected_year_full == today.year and m == today.month:
                                    grade = "-"
                                else:
                                    grade = "D"
                                    count_d += 1
                            
                            months_grades[f"Tháng {m}"] = grade
                        
                        # Logic xếp loại năm mới
                        evaluated = count_a + count_b + count_c + count_d
                        if evaluated == 0:
                            final_grade = "-"
                            bonus = "-"
                        elif evaluated < 12 and selected_year_full >= today.year:
                            final_grade = "Đang tích lũy"
                            bonus = "-"
                        else:
                            if count_c > 0 or count_d > 0:
                                final_grade = "C"
                                bonus = "60%"
                            elif count_b >= 2:
                                final_grade = "B"
                                bonus = "80%"
                            elif count_a >= 11:
                                final_grade = "A"
                                bonus = "100%"
                            else:
                                final_grade = "B"
                                bonus = "80%"
                            
                        row_data = {
                            "Người thực hiện": person,
                            "Phòng ban": DEPT_ABBR.get(person_df['PhongBan'].mode()[0], person_df['PhongBan'].mode()[0]) if not person_df.empty else ""
                        }
                        row_data.update(months_grades)
                        row_data["Xếp loại Năm"] = final_grade
                        row_data["Mức hưởng T13"] = bonus
                        yearly_data.append(row_data)
                    
                    if yearly_data:
                        yearly_df = pd.DataFrame(yearly_data)
                        if selected_dept_y != "Tất cả phòng ban":
                            yearly_df = yearly_df[yearly_df["Phòng ban"] == DEPT_ABBR.get(selected_dept_y, selected_dept_y)]
                        st.dataframe(yearly_df, use_container_width=True, hide_index=True)
                    
                        excel_data = kpi_reports.generate_yearly_excel(yearly_df, selected_year_full)
                        st.download_button("📥 Xuất Báo cáo Excel", data=excel_data, file_name=f"TongKet_KPI_{selected_year_full}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else:
                        st.info("Không có dữ liệu.")

    if role_mode in ["Quản lý", "HR"] and st.session_state.is_admin_authenticated:
        with kpi_tab3:
            st.markdown("#### ⚖️ Điều chỉnh Điểm Thưởng / Phạt")
            all_p_list = []
            if selected_company == "Tất cả đơn vị":
                for dept, persons in config.get("personnel_by_department", {}).items():
                    all_p_list.extend(persons)
            else:
                valid_depts = get_departments_for_company(selected_company, config)
                for dept in valid_depts:
                    persons = get_personnel_for_company_dept(selected_company, dept, config)
                    all_p_list.extend(persons)
                    
                is_marina_co = "CTY CP DMT - MARINA" in selected_company or "Du thuyền Happy Yacht" in selected_company
                is_traffic_co = "XÂY DỰNG CÔNG TRÌNH GIAO THÔNG ĐN-MT" in selected_company
                if not is_marina_co and not is_traffic_co:
                    c_personnel = set(display_df['NguoiChuTri'].dropna().unique())
                    leads = DEPT_LEADS.get(selected_company, {}).values()
                    valid_people = c_personnel.union(set(leads))
                    all_p_list = [p for p in all_p_list if p in valid_people]
                    
            all_p_list = sorted(list(set(all_p_list)))
            if not all_p_list:
                all_p_list = ["(Chưa có nhân sự)"]
                
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                adj_person = st.selectbox("Tên nhân viên", all_p_list)
                adj_month = st.selectbox("Tháng áp dụng", list(range(1, 13)), index=today.month - 1)
                adj_year = st.selectbox("Năm áp dụng", [today.year - 1, today.year, today.year + 1], index=1)
            
            with col_f2:
                adj_template = st.selectbox("Lý do mẫu", ["Đi trễ, về sớm", "Quên chấm công", "Lý do khác"])
                
                if adj_template == "Đi trễ, về sớm":
                    so_lan = st.number_input("Tổng số lần trong tháng", min_value=1, value=1)
                    sugg_val = max(0, so_lan - 5) * 2
                    
                    st.info(f"ℹ️ Bạn đang nhập tổng cộng {so_lan} lần vi phạm trong tháng {adj_month}.")
                    if so_lan >= 6:
                        st.warning(f"⚠️ Từ lần 6 trở đi: Đề xuất trừ tổng cộng {sugg_val} điểm.")
                    else:
                        st.success("✅ Dưới 6 lần: Chưa bị trừ điểm.")
                        
                    adj_type = "🛑 Phạt điểm"
                    adj_val = st.number_input("Tổng số điểm trừ", min_value=0, max_value=100, value=sugg_val)
                    adj_reason = st.text_input("Ghi chú thêm (Tùy chọn)")
                    
                elif adj_template == "Quên chấm công":
                    so_lan = st.number_input("Tổng số lần trong tháng", min_value=1, value=1)
                    sugg_val = max(0, so_lan - 2) * 1
                    
                    st.info(f"ℹ️ Bạn đang nhập tổng cộng {so_lan} lần vi phạm trong tháng {adj_month}.")
                    if so_lan >= 3:
                        st.warning(f"⚠️ Từ lần 3 trở đi: Đề xuất trừ tổng cộng {sugg_val} điểm.")
                    else:
                        st.success("✅ Dưới 3 lần: Chưa bị trừ điểm.")
                        
                    adj_type = "🛑 Phạt điểm"
                    adj_val = st.number_input("Tổng số điểm trừ", min_value=0, max_value=100, value=sugg_val)
                    adj_reason = st.text_input("Ghi chú thêm (Tùy chọn)")
                    
                else:
                    adj_type = st.radio("Phân loại hành vi", ["⭐ Thưởng điểm", "🛑 Phạt điểm"], horizontal=True)
                    adj_val = st.number_input("Số điểm", min_value=0, max_value=15, value=5)
                    adj_reason = st.text_area("Lý do chi tiết (Bắt buộc)")
                    
            if st.button("Lưu Điểm Điều Chỉnh", type="primary"):
                if adj_template == "Lý do khác" and not adj_reason.strip():
                    st.error("⚠️ Vui lòng nhập lý do chi tiết!")
                else:
                    actual_val = adj_val if adj_type == "⭐ Thưởng điểm" else -adj_val
                    if adj_template != "Lý do khác":
                        final_reason = f"[{adj_template}] ({so_lan} lần) {adj_reason.strip()}".strip()
                    else:
                        final_reason = adj_reason.strip()
                    add_kpi_adjustment(adj_person, adj_month, adj_year, adj_type, actual_val, final_reason)
                    st.session_state["success_msg"] = "🎉 Đã lưu điều chỉnh điểm thành công!"
                    st.rerun()

        with kpi_tab4:
            st.markdown("### Tùy chọn Xuất Báo Cáo & Phân tích")
            col_x1, col_x2 = st.columns(2)
            
            with col_x1:
                st.markdown("#### 1. Báo Cáo Phòng Ban (Excel)")
                if st.button("Tải Báo cáo Phòng Ban"):
                    data_rows = []
                    for i, person in enumerate(all_p_list):
                        data_rows.append({
                            'HoTen': person, 'ChucVu': 'Nhân viên',
                            'SoLanTre': 0, 'SoLanSom': 0, 'SoLanKhongCC': 0,
                            'DiemTruTre': 0, 'DiemTruSom': 0, 'DiemTruKhongCC': 0,
                            'TongTru': 0, 'DiemConLai': 100, 'XepLoai': 'A', 'GhiChu': ''
                        })
                    excel_data = kpi_reports.generate_department_excel(selected_company, selected_month, selected_year, data_rows)
                    st.download_button("📥 Tải Báo Cáo Excel", data=excel_data, file_name=f"KPI_Thang_{selected_month}_{selected_year}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            with col_x2:
                st.markdown("#### 2. Phiếu KPI Cá Nhân (Word)")
                st.info(f"Đang xuất dữ liệu của: **Tháng {selected_month}/{selected_year}** (Để xuất tháng khác, vui lòng quay lại tab 'Đánh giá theo Tháng' để chọn).")
                emp_to_export = st.selectbox("Chọn nhân viên", all_p_list, key='emp_export')
                if st.button("Tạo Phiếu Đánh Giá"):
                    # Collect real tasks and penalties
                    emp_tasks = []
                    kpi_score = 100
                    if personnel_kpi:
                        for p in personnel_kpi:
                            if p['Người thực hiện'] == emp_to_export:
                                kpi_score = p['Điểm công việc']
                                break
                    
                    e_kpi_df = kpi_df[kpi_df['NguoiChuTri'] == emp_to_export].copy()
                    e_kpi_df['TyTrongKPI'] = pd.to_numeric(e_kpi_df['TyTrongKPI'], errors='coerce').fillna(0)
                    e_kpi_df['PhanTramHoanThanh'] = pd.to_numeric(e_kpi_df['PhanTramHoanThanh'], errors='coerce').fillna(0)
                    # Calc weights again if needed, or just display raw tasks
                    explicit_weight_sum = e_kpi_df[e_kpi_df['TyTrongKPI'] > 0]['TyTrongKPI'].sum()
                    unweighted_count = len(e_kpi_df[e_kpi_df['TyTrongKPI'] <= 0])
                    remaining_weight = max(0, 100 - explicit_weight_sum)
                    auto_weight = remaining_weight / unweighted_count if unweighted_count > 0 else 0
                    
                    for idx, row in e_kpi_df.iterrows():
                        w = row['TyTrongKPI'] if row['TyTrongKPI'] > 0 else auto_weight
                        pt = row.get('PhanTramHoanThanh', 0)
                        if pd.isna(pt): pt = 0
                        
                        diem_tru = w - (pt / 100.0 * w)
                        emp_tasks.append({
                            'TenCV': row['TenCongViec'],
                            'TgianYC': str(row['Deadline']),
                            'KetQua': f"{pt}% (Tỷ trọng: {w:.1f}%)",
                            'DiemTru': round(diem_tru, 1)
                        })
                        
                    e_adj_df = adj_df[adj_df['TenNhanVien'] == emp_to_export] if 'TenNhanVien' in adj_df.columns else pd.DataFrame()
                    penalties = e_adj_df.to_dict('records')
                    
                    def _get_pb(name):
                        # Try to find from current company's departments
                        depts = get_departments_for_company(selected_company, config)
                        if depts:
                            for d in depts:
                                p_list = get_personnel_for_company_dept(selected_company, d, config)
                                if name in p_list: return d
                        # Fallback to global config
                        for d, p_list in config.get("personnel_by_department", {}).items():
                            if name in p_list: return d
                        return "Khác"

                    emp_pb = _get_pb(emp_to_export)
                    word_data = kpi_reports.generate_individual_docx(emp_to_export, selected_month, selected_year, kpi_score, emp_tasks, penalties, "Nhân viên", emp_pb)
                    st.download_button("📥 Tải Phiếu Cá Nhân (Word)", data=word_data, file_name=f"Phieu_KPI_{emp_to_export}_Thang_{selected_month}_{selected_year}.docx", mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
                    
            st.divider()
            st.info("Để xem Biểu đồ Phân tích, vui lòng qua tab 'Tổng kết KPI Cả Năm' và bấm 'Chạy / Cập nhật' trước.")

                        
            st.markdown("##### Lịch sử Thưởng / Phạt")
            hist_df = read_kpi_adjustments()
            if not hist_df.empty:
                # Filter out personnel not belonging to the current company
                hist_df = hist_df[hist_df["TenNhanVien"].isin(all_p_list)]
                
                if hist_df.empty:
                    st.info("Chưa có lịch sử điều chỉnh cho đơn vị này.")
                else:
                    def _get_pb(name):
                        # Try to find from current company's departments
                        if selected_company != "Tất cả đơn vị":
                            depts = get_departments_for_company(selected_company, config)
                            for d in depts:
                                p_list = get_personnel_for_company_dept(selected_company, d, config)
                                if name in p_list: return d
                        # Fallback to global config
                        for d, p_list in config.get("personnel_by_department", {}).items():
                            if name in p_list: return d
                        return "Khác"
                        
                    hist_df["Phòng ban"] = hist_df["TenNhanVien"].apply(_get_pb)
                    
                    # Reorder columns to put Phòng ban next to TenNhanVien
                    cols = list(hist_df.columns)
                    if "Phòng ban" in cols:
                        cols.insert(cols.index("TenNhanVien") + 1, cols.pop(cols.index("Phòng ban")))
                        hist_df = hist_df[cols]
                    
                    # Lọc (Filter)
                    col_flt1, col_flt2 = st.columns(2)
                    with col_flt1:
                        f_pb = st.selectbox("Lọc Phòng Ban", ["Tất cả"] + sorted(list(set(hist_df["Phòng ban"]))), key="flt_adj_pb")
                    with col_flt2:
                        f_thang = st.selectbox("Lọc Tháng", ["Tất cả"] + sorted(list(set(hist_df["Thang"])), reverse=True), key="flt_adj_thang")
                    
                    hist_display_df = hist_df.copy()
                    if f_pb != "Tất cả":
                        hist_display_df = hist_display_df[hist_display_df["Phòng ban"] == f_pb]
                    if f_thang != "Tất cả":
                        hist_display_df = hist_display_df[hist_display_df["Thang"] == f_thang]
                    
                    st.dataframe(hist_display_df.sort_values(by=["Phòng ban", "Thang", "ID"], ascending=[True, False, False]), use_container_width=True, hide_index=True)
                    
                    st.markdown("---")
                    st.markdown("##### 🗑️ Xóa Điều Chỉnh KPI")
                    st.info("Nhập số ID tương ứng trong bảng trên để xóa dữ liệu.")
                    col_del1, col_del2 = st.columns([1, 3])
                    with col_del1:
                        del_id = st.number_input("Nhập ID cần xóa", min_value=1, value=1)
                    with col_del2:
                        st.write("") # Spacer
                        st.write("")
                        if st.button("❌ Xóa dòng này", type="primary"):
                            success, msg = delete_kpi_adjustment(del_id)
                            if success:
                                st.success(f"Đã xóa thành công điều chỉnh có ID: {del_id}")
                                st.rerun()
                            else:
                                st.error(msg)

            else:
                st.info("Chưa có lịch sử điều chỉnh.")

# ----------------- 6. QUẢN LÝ CẤU HÌNH -----------------# ----------------- 6. QUẢN LÝ CẤU HÌNH -----------------
elif menu in ["✅ Duyệt việc Khách quan", "⚖️ Duyệt việc Khách quan"]:
    st.header("✅ Duyệt lý do trễ hạn khách quan")
    
    if not (st.session_state.is_admin_authenticated or st.session_state.get("is_manager_authenticated", False)):
        st.warning("⚠️ Vui lòng nhập **Mật khẩu Quản lý** ở thanh bên trái (cột menu) để truy cập tính năng này.")
    else:
        df = read_db()
        if df.empty:
            st.info("Chưa có dữ liệu công việc.")
        else:
            col_thang, col_nam, col_phong = st.columns(3)
            with col_thang:
                thang_opts = list(range(1, 13))
                sel_thang = st.selectbox("Chọn Tháng", thang_opts, index=today.month - 1)
            with col_nam:
                nam_opts = [today.year - 1, today.year, today.year + 1]
                sel_nam = st.selectbox("Chọn Năm", nam_opts, index=1)
            with col_phong:
                config = load_settings()
                valid_depts = get_departments_for_company(selected_company, config)
                phong_opts = ["Tất cả"] + valid_depts
                sel_phong = st.selectbox("Lọc theo Phòng/Ban", phong_opts)
                
            st.markdown("---")
            
            # Filter logic
            def is_in_selected_month(d_str):
                if not d_str: return False
                try:
                    d = pd.to_datetime(d_str)
                    return d.month == sel_thang and d.year == sel_nam
                except:
                    return False
                    
            df['is_in_month'] = df['Deadline'].apply(is_in_selected_month)
            
            # Condition: Deadline in month, objective reason
            mask = df['is_in_month'] & df['PhanLoaiTreHan'].astype(str).str.lower().str.contains("khách quan")
            if sel_phong != "Tất cả":
                mask = mask & (df['PhongBan'] == sel_phong)
                
            filtered_df = df[mask].copy()
            
            if filtered_df.empty:
                st.success(f"🎉 Không có công việc nào báo cáo Khách quan trong tháng {sel_thang}/{sel_nam}!")
            else:
                st.info(f"Đang hiển thị **{len(filtered_df)}** công việc báo cáo lý do Khách quan.")
                
                # Setup Editor
                edit_cols = ["ID", "PhongBan", "NguoiChuTri", "TenCongViec", "Deadline", "TrangThai", "GiaiTrinhDeXuat", "MucDoGhiNhan"]
                disp_df = filtered_df[edit_cols].copy()
                
                # We need to make all columns disabled EXCEPT MucDoGhiNhan
                col_config = {
                    "ID": st.column_config.TextColumn("Mã CV", disabled=True),
                    "PhongBan": st.column_config.TextColumn("Phòng/Ban", disabled=True),
                    "NguoiChuTri": st.column_config.TextColumn("Người Phụ Trách", disabled=True),
                    "TenCongViec": st.column_config.TextColumn("Tên Công Việc", disabled=True),
                    "Deadline": st.column_config.DateColumn("Hạn Chót", disabled=True, format="DD/MM/YYYY"),
                    "TrangThai": st.column_config.TextColumn("Trạng Thái", disabled=True),
                    "GiaiTrinhDeXuat": st.column_config.TextColumn("Giải Trình Khách Quan", disabled=True),
                    "MucDoGhiNhan": st.column_config.SelectboxColumn(
                        "Mức độ Ghi nhận KPI",
                        help="Chọn mức điểm đánh giá theo lý do khách quan (Chỉ dành cho Quản lý)",
                        options=["0% (Không ghi nhận)", "Miễn trừ (Loại bỏ KPI)", "50%", "80%", "90%"],
                        required=True
                    )
                }
                
                edited_df = st.data_editor(
                    disp_df,
                    column_config=col_config,
                    hide_index=True,
                    use_container_width=True,
                    num_rows="fixed",
                    key=f"editor_approve_{sel_thang}_{sel_nam}"
                )
                
                if st.button("💾 Lưu tất cả thay đổi", type="primary"):
                    with acquire_db_lock():
                        
                        fresh_df = read_db()
                        changed = False
                        for idx, row in edited_df.iterrows():
                            task_id = row['ID']
                            new_val = row['MucDoGhiNhan']
                            # Some tasks might not exist if deleted concurrently, but for robustness:
                            if task_id in fresh_df['ID'].values:
                                old_val = fresh_df.loc[fresh_df['ID'] == task_id, 'MucDoGhiNhan'].values[0]
                                if new_val != old_val:
                                    fresh_df.loc[fresh_df['ID'] == task_id, 'MucDoGhiNhan'] = new_val
                                    changed = True
                                
                        if changed:
                            fresh_df = fresh_df.drop(columns=['is_in_month'], errors='ignore')
                            if save_db(fresh_df):
                                st.success("✅ Đã lưu toàn bộ phê duyệt thành công!")
                                st.rerun()
                        else:
                            st.info("Chưa có thay đổi nào cần lưu.")

elif menu == "🔍 Quản lý & Đối chiếu JD":
    st.markdown("### 🔍 Quản lý & Đối chiếu JD (Trí tuệ nhân tạo)")
    st.info("💡 Hệ thống sử dụng **Trí tuệ nhân tạo (Google Gemini)** để phân tích tự động việc nhân sự làm có đúng chuyên môn trong Bản Mô tả công việc (JD) hay không.")
    
    tab_hr, tab_ai = st.tabs(["📝 1. Cập nhật Mô tả công việc (Dành cho HR)", "🔍 2. AI Đối chiếu & Báo cáo (Dành cho Sếp)"])
    
    with tab_hr:
        st.markdown("#### Khai báo JD nguyên bản cho Nhân sự")
        st.write("Vui lòng mở file Word Mô tả công việc của nhân sự, copy phần **TRÁCH NHIỆM CÔNG VIỆC** và dán vào đây.")
        
        # Select personnel
        if selected_company == "Tất cả đơn vị":
            st.warning("⚠️ Vui lòng chọn cụ thể Công ty ở cột trái.")
        else:
            comp_data = config.get("companies", {}).get(selected_company, {})
            all_depts = comp_data.get("departments", [])
            sel_dept = st.selectbox("Chọn Phòng ban", all_depts, key="jd_dept")
            
            personnel = comp_data.get("personnel_by_department", {}).get(sel_dept, [])
            if personnel:
                sel_person = st.selectbox("Chọn Nhân sự", personnel, key="jd_person")
                
                if "job_descriptions" not in config:
                    config["job_descriptions"] = {}
                if selected_company not in config["job_descriptions"]:
                    config["job_descriptions"][selected_company] = {}
                    
                existing_jd_data = config["job_descriptions"][selected_company].get(sel_person, "")
                if isinstance(existing_jd_data, str):
                    existing_jd_text = existing_jd_data
                else:
                    existing_jd_text = existing_jd_data.get("jd_text", "")
                    
                st.write("---")
                st.markdown("**Cách 1: Nhập văn bản hoặc dán (Copy/Paste)**")
                jd_text = st.text_area("Nội dung Mô tả công việc:", value=existing_jd_text, height=200, key=f"jd_text_{sel_person}")
                
                st.markdown("**Cách 2: Tải lên file Word/PDF (Tự động đọc nội dung)**")
                uploaded_file = st.file_uploader("Kéo thả file JD vào đây", type=['docx', 'pdf'], key=f"jd_upload_{sel_person}")
                
                if uploaded_file is not None:
                    if st.button("Trích xuất nội dung từ File"):
                        with st.spinner("Đang đọc file..."):
                            try:
                                text = ""
                                if uploaded_file.name.endswith(".docx"):
                                    import docx
                                    doc = docx.Document(uploaded_file)
                                    text = "\n".join([p.text for p in doc.paragraphs])
                                elif uploaded_file.name.endswith(".pdf"):
                                    import pypdf
                                    pdf = pypdf.PdfReader(uploaded_file)
                                    text = "\n".join([page.extract_text() for page in pdf.pages if page.extract_text()])
                                
                                if not text.strip():
                                    st.warning("⚠️ Không thể đọc được chữ từ file này (có thể đây là file scan/ảnh). Vui lòng copy và dán văn bản thủ công vào ô phía trên.")
                                else:
                                    st.session_state[f"extracted_text_{sel_person}"] = text
                            except Exception as e:
                                st.error(f"Lỗi đọc file: {e}")
                
                if st.session_state.get(f"extracted_text_{sel_person}"):
                    st.success("Đã trích xuất thành công! Bạn có thể xem và chỉnh sửa trước khi lưu:")
                    jd_text = st.text_area("Nội dung trích xuất", value=st.session_state[f"extracted_text_{sel_person}"], height=200, key=f"jd_text_ext_{sel_person}")
                
                if st.button("💾 Lưu Mô tả công việc", type="primary"):
                    if not jd_text.strip():
                        st.error("⚠️ Nội dung Mô tả công việc đang trống! Vui lòng nhập nội dung hoặc trích xuất từ file trước khi lưu.")
                    else:
                        if isinstance(existing_jd_data, dict):
                            new_data = existing_jd_data.copy()
                            new_data["jd_text"] = jd_text
                        else:
                            new_data = {"jd_text": jd_text}
                            
                        config["job_descriptions"][selected_company][sel_person] = new_data
                        if save_config(config):
                            st.success(f"✅ Đã lưu Bản mô tả công việc (JD) thành công cho nhân sự **{sel_person}**!")
                            import time
                            time.sleep(1)
                            st.rerun()
            else:
                st.warning("Phòng ban này chưa có nhân sự.")

    with tab_ai:
        st.markdown("#### 🔍 Phân tích độ phủ công việc thực tế với JD")
        if selected_company == "Tất cả đơn vị":
            st.warning("⚠️ Vui lòng chọn cụ thể Công ty ở cột trái.")
        else:
            comp_data = config.get("companies", {}).get(selected_company, {})
            
            months = set()
            if not display_df.empty:
                for _, row in display_df.iterrows():
                    if pd.notna(row.get('Deadline')) and hasattr(row['Deadline'], 'strftime'):
                        months.add(row['Deadline'].strftime('%m/%Y'))
            month_options = sorted(list(months), key=lambda x: datetime.strptime(x, '%m/%Y'), reverse=True)
            if not month_options: month_options = [today.strftime('%m/%Y')]
            
            c1, c2, c3 = st.columns(3)
            with c1: ai_month = st.selectbox("Tháng đánh giá", month_options, key="ai_month_select")
            with c2: ai_dept = st.selectbox("Phòng ban", comp_data.get("departments", []), key="ai_dept_select")
            
            ai_personnel = comp_data.get("personnel_by_department", {}).get(ai_dept, [])
            with c3:
                if ai_personnel:
                    ai_person = st.selectbox("Nhân sự", ai_personnel, key="ai_person_select")
                else:
                    ai_person = None
                    st.warning("Trống")
            

            # --- BATCH AI SCAN ---
            st.markdown("---")
            with st.expander("⚡ Quét nhanh toàn bộ Phòng ban (Batch AI Scan)", expanded=False):
                st.info("Tính năng này sẽ tự động kiểm tra JD của tất cả nhân sự trong phòng ban. Những ai chưa có kết quả sẽ tự động gọi AI để phân tích. Khuyên dùng khi bạn muốn kiểm tra tổng thể cả phòng.")
                
                batch_api_key = ""
                try:
                    if "gemini" in st.secrets and "api_key" in st.secrets["gemini"]:
                        batch_api_key = st.secrets["gemini"]["api_key"]
                except:
                    pass
                if not batch_api_key:
                    import os
                    batch_api_key = os.environ.get("GEMINI_API_KEY", "")
                    
                if not batch_api_key:
                    batch_api_key = st.text_input("🔑 Nhập khóa API Gemini để quét hàng loạt:", type="password", key="batch_api_key_input")
                
                if st.button("🚀 Bắt đầu Quét toàn bộ", type="primary"):
                    if not batch_api_key:
                        st.error("Vui lòng nhập API Key!")
                    elif not ai_personnel:
                        st.warning("Phòng ban không có nhân sự.")
                    else:
                        import google.generativeai as genai
                        import hashlib
                        import json
                        import time
                        
                        genai.configure(api_key=batch_api_key, transport='rest')
                        valid_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
                        model_name = 'gemini-3.6-flash' if 'models/gemini-3.6-flash' in valid_models else ('gemini-1.5-flash' if 'models/gemini-1.5-flash' in valid_models else 'gemini-pro')
                        model = genai.GenerativeModel(model_name)
                        
                        results = []
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        def is_same_person_batch(db_name, target_name):
                            db_str = str(db_name).strip().lower()
                            tgt_str = str(target_name).strip().lower()
                            if db_str == tgt_str: return True
                            tgt_parts = tgt_str.split()
                            if len(tgt_parts) >= 2:
                                return tgt_parts[0] in db_str and tgt_parts[-1] in db_str
                            return False
                        
                        total_people = len(ai_personnel)
                        
                        for idx, p in enumerate(ai_personnel):
                            status_text.text(f"Đang phân tích ({idx+1}/{total_people}): {p}...")
                            
                            # Lọc công việc
                            p_tasks = display_df[
                                (display_df['NguoiChuTri'].apply(lambda x: is_same_person_batch(x, p))) & 
                                (display_df['Deadline'].apply(lambda x: x.strftime('%m/%Y') if pd.notna(x) and hasattr(x, 'strftime') else '') == ai_month)
                            ]
                            
                            if p_tasks.empty:
                                results.append({"Nhân sự": p, "Kết quả": "Trống", "Tỷ lệ khớp": None, "Chi tiết": "Không có công việc trong tháng này"})
                            else:
                                jd_data = config.get("job_descriptions", {}).get(selected_company, {}).get(p, "")
                                jd_str = jd_data if isinstance(jd_data, str) else jd_data.get("jd_text", "")
                                
                                if not jd_str.strip():
                                    results.append({"Nhân sự": p, "Kết quả": "Thiếu JD", "Tỷ lệ khớp": None, "Chi tiết": "Chưa khai báo Mô tả công việc"})
                                else:
                                    tasks_list = "\n".join([f"- {row['TenCongViec']}" for _, row in p_tasks.iterrows()])
                                    prompt = f"""
                                    Đóng vai một Giám đốc nhân sự cực kỳ tinh tế. 
                                    Dưới đây là Bản Mô tả công việc (JD) của nhân viên {p}:
                                
                                    [BẢN MÔ TẢ CÔNG VIỆC]
                                    {jd_str}
                                    [KẾT THÚC JD]
                                
                                    Và đây là danh sách công việc họ thực hiện trong tháng:
                                    {tasks_list}
                                
                                    NHIỆM VỤ CỦA BẠN:
                                    1. Đối chiếu TỪNG công việc xem nó có KHỚP với chuyên môn quy định trong JD không. 
                                    (Lưu ý: Tên công việc thực tế có thể chi tiết và từ ngữ khác biệt so với JD văn xuôi. Hãy dùng tư duy suy luận về bản chất và mục đích để phán đoán).
                                    2. Nếu khớp, giải thích vì nó phục vụ cho mục nào trong JD. Nếu ngoài JD, ghi rõ là công việc phát sinh.
                                    3. Format kết quả đầu ra thành đúng định dạng chuỗi JSON thô như sau (chỉ trả về JSON, không chứa dấu tick markdown ```json):
                                    {{
                                        "ty_le_khop": <số nguyên từ 0-100, ví dụ 80>,
                                        "chi_tiet": [
                                            {{
                                                "ten_cong_viec": "<Tên công việc y nguyên trong danh sách>",
                                                "phan_loai": "<Chỉ điền 'Khớp JD' hoặc 'Ngoài JD'>",
                                                "nhan_xet": "<Phân tích ngắn gọn 1-2 câu>"
                                            }}
                                        ]
                                    }}
                                    """
                                    
                                    prompt_hash = hashlib.md5(prompt.encode('utf-8')).hexdigest()
                                    cache_file = f".ai_cache_{prompt_hash}.txt"
                                    
                                    try:
                                        if os.path.exists(cache_file):
                                            with open(cache_file, "r", encoding="utf-8") as f:
                                                raw_text = f.read()
                                        else:
                                            max_retries = 3
                                            for attempt in range(max_retries):
                                                try:
                                                    response = model.generate_content(
                                                        prompt, 
                                                        generation_config={"temperature": 0.0},
                                                        request_options={"timeout": 60.0}
                                                    )
                                                    raw_text = response.text
                                                    break
                                                except Exception as api_err:
                                                    if attempt == max_retries - 1:
                                                        raise api_err
                                                    time.sleep(3 * (attempt + 1))
                                                    
                                            if raw_text:
                                                with open(cache_file, "w", encoding="utf-8") as f:
                                                    f.write(raw_text)
                                            time.sleep(3) # Tránh rate limit

                                            
                                        cleaned = raw_text.strip()
                                        if cleaned.startswith("```json"):
                                            cleaned = cleaned[7:]
                                        if cleaned.endswith("```"):
                                            cleaned = cleaned[:-3]
                                        
                                        data = json.loads(cleaned)
                                        ty_le = data.get("ty_le_khop", 0)
                                        ngoai_jd_count = sum(1 for c in data.get("chi_tiet", []) if "Ngoài JD" in c.get("phan_loai", ""))
                                        
                                        if ty_le == 100:
                                            res_text = "🟢 Tốt (100% khớp)"
                                        elif ty_le >= 50:
                                            res_text = f"🟡 Cảnh báo ({ty_le}% khớp)"
                                        else:
                                            res_text = f"🔴 Lệch JD ({ty_le}% khớp)"
                                            
                                        chi_tiet_text = f"{ngoai_jd_count} việc ngoài JD" if ngoai_jd_count > 0 else "Hoàn toàn khớp"
                                        
                                        results.append({"Nhân sự": p, "Kết quả": res_text, "Tỷ lệ khớp": ty_le, "Chi tiết": chi_tiet_text})
                                    except Exception as e:
                                        results.append({"Nhân sự": p, "Kết quả": "Lỗi AI", "Tỷ lệ khớp": None, "Chi tiết": str(e)})
                            
                            progress_bar.progress((idx + 1) / total_people)
                            
                        status_text.success("✅ Đã hoàn thành phân tích toàn bộ phòng ban!")
                        
                        if results:
                            df_res = pd.DataFrame(results)
                            st.dataframe(df_res, use_container_width=True, hide_index=True)
                            
            st.markdown("---")

            if ai_person:
                jd_source_data = config.get("job_descriptions", {}).get(selected_company, {}).get(ai_person, "")
                if isinstance(jd_source_data, str):
                    jd_source = jd_source_data
                else:
                    jd_source = jd_source_data.get("jd_text", "")
                    
                if not jd_source.strip():
                    st.error(f"⚠️ Nhân sự **{ai_person}** chưa được khai báo Mô tả công việc. Vui lòng sang tab bên cạnh để cập nhật JD trước khi AI có thể quét.")
                else:
                    with st.expander("Xem trước JD gốc (Làm cơ sở chấm) 👀", expanded=False):
                        st.text(jd_source)
                        
                    # Filter tasks for this person in this month flexibly to handle name changes (e.g. Lê Ngọc Tú Uyên vs Lê Thị Tú Uyên)
                    def is_same_person(db_name, target_name):
                        db_str = str(db_name).strip().lower()
                        tgt_str = str(target_name).strip().lower()
                        if db_str == tgt_str: return True
                        
                        # Match first and last name if exact match fails
                        tgt_parts = tgt_str.split()
                        if len(tgt_parts) >= 2:
                            return tgt_parts[0] in db_str and tgt_parts[-1] in db_str
                        return False
                        
                    ai_tasks = display_df[
                        (display_df['NguoiChuTri'].apply(lambda x: is_same_person(x, ai_person))) & 
                        (display_df['Deadline'].apply(lambda x: x.strftime('%m/%Y') if pd.notna(x) and hasattr(x, 'strftime') else '') == ai_month)
                    ]
                    
                    if ai_tasks.empty:
                        st.info(f"Không có công việc nào được đăng ký trong tháng {ai_month}.")
                    else:
                        st.write(f"Tìm thấy **{len(ai_tasks)}** đầu công việc do nhân sự đăng ký trong tháng.")
                        
                        import os
                        api_key = ""
                        try:
                            if "gemini" in st.secrets and "api_key" in st.secrets["gemini"]:
                                api_key = st.secrets["gemini"]["api_key"]
                        except:
                            pass
                            
                        if not api_key:
                            api_key = os.environ.get("GEMINI_API_KEY", "")
                            
                        if not api_key:
                            api_key = st.text_input("🔑 Nhập khóa API Gemini (API Key) của bạn để tiếp tục:", type="password")
                            
                        if not api_key:
                            st.warning("⚠️ Vui lòng cấu hình API Key hoặc nhập vào ô trống bên trên để sử dụng AI.")
                        else:
                            if st.button("🔍 CHẠY AI QUÉT ĐỘ PHỦ (GEMINI)", type="primary"):
                                with st.spinner("🧠 AI đang đọc JD và suy luận công việc... (Có thể mất 5-10 giây)"):
                                    try:
                                        import google.generativeai as genai
                                        genai.configure(api_key=api_key, transport='rest')
                                        
                                        # Tự động dò model khả dụng cho API Key này
                                        valid_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
                                        model_name = 'gemini-3.6-flash'
                                        if 'models/gemini-3.6-flash' in valid_models:
                                            model_name = 'gemini-3.6-flash'
                                        elif 'models/gemini-1.5-flash' in valid_models:
                                            model_name = 'gemini-1.5-flash'
                                        elif 'models/gemini-pro' in valid_models:
                                            model_name = 'gemini-pro'
                                        elif valid_models:
                                            # Tránh chọn các model cũ hoặc bị deprecate nằm ở đầu danh sách
                                            model_name = valid_models[-1].replace('models/', '')
                                            
                                        model = genai.GenerativeModel(model_name)
                                        
                                        # Rút gọn danh sách công việc
                                        tasks_list = "\n".join([f"- {row['TenCongViec']}" for _, row in ai_tasks.iterrows()])
                                    
                                        prompt = f"""
                                        Đóng vai một Giám đốc nhân sự cực kỳ tinh tế. 
                                        Dưới đây là Bản Mô tả công việc (JD) của nhân viên {ai_person}:
                                    
                                        [BẢN MÔ TẢ CÔNG VIỆC]
                                        {jd_source}
                                        [KẾT THÚC JD]
                                    
                                        Và đây là danh sách công việc họ thực hiện trong tháng:
                                        {tasks_list}
                                    
                                        NHIỆM VỤ CỦA BẠN:
                                        1. Đối chiếu TỪNG công việc xem nó có KHỚP với chuyên môn quy định trong JD không. 
                                        (Lưu ý: Tên công việc thực tế có thể chi tiết và từ ngữ khác biệt so với JD văn xuôi. Hãy dùng tư duy suy luận về bản chất và mục đích để phán đoán).
                                        2. Nếu khớp, giải thích vì nó phục vụ cho mục nào trong JD. Nếu ngoài JD, ghi rõ là công việc phát sinh.
                                        3. Format kết quả đầu ra thành đúng định dạng chuỗi JSON thô như sau (chỉ trả về JSON, không chứa dấu tick markdown ```json):
                                        {{
                                            "ty_le_khop": <số nguyên từ 0-100, ví dụ 80>,
                                            "chi_tiet": [
                                                {{
                                                    "ten_cong_viec": "<Tên công việc y nguyên trong danh sách>",
                                                    "phan_loai": "<Chỉ điền 'Khớp JD' hoặc 'Ngoài JD'>",
                                                    "nhan_xet": "<Phân tích ngắn gọn 1-2 câu>"
                                                }}, ...
                                            ]
                                        }}
                                        """
                                    
                                        import hashlib
                                        import os
                                        
                                        prompt_hash = hashlib.md5(prompt.encode('utf-8')).hexdigest()
                                        cache_file = f".ai_cache_{prompt_hash}.txt"
                                        
                                        if os.path.exists(cache_file):
                                            with open(cache_file, "r", encoding="utf-8") as f:
                                                raw_text = f.read()
                                        else:
                                            max_retries = 3
                                            for attempt in range(max_retries):
                                                try:
                                                    response = model.generate_content(
                                                        prompt, 
                                                        generation_config={"temperature": 0.0},
                                                        request_options={"timeout": 60.0}
                                                    )
                                                    raw_text = response.text
                                                    break
                                                except Exception as api_err:
                                                    if attempt == max_retries - 1:
                                                        raise api_err
                                                    import time
                                                    time.sleep(4 * (attempt + 1))
                                                    
                                            if raw_text:
                                                with open(cache_file, "w", encoding="utf-8") as f:
                                                    f.write(raw_text)
                                    
                                        import re
                                        json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
                                        if json_match:
                                            res_json = json.loads(json_match.group())
                                        
                                            st.markdown("### 📊 KẾT QUẢ ĐỐI CHIẾU JD VÀ CÔNG VIỆC ĐĂNG KÝ")
                                        
                                            # Pie chart
                                            match_rate = res_json.get("ty_le_khop", 0)
                                            m_data = pd.DataFrame({
                                                "Phân loại": ["Khớp chuyên môn (JD)", "Công việc ngoài JD"],
                                                "Tỷ lệ": [match_rate, 100 - match_rate]
                                            })
                                            fig = px.pie(m_data, values='Tỷ lệ', names='Phân loại', color='Phân loại',
                                                         color_discrete_map={"Khớp chuyên môn (JD)": "#22c55e", "Công việc ngoài JD": "#f97316"},
                                                         title=f"Độ phủ JD Tháng {ai_month}", hole=0.4)
                                            st.plotly_chart(fig, use_container_width=True)
                                        
                                            # Table
                                            res_df = pd.DataFrame(res_json.get("chi_tiet", []))
                                            if not res_df.empty:
                                                # Format columns for display
                                                res_df = res_df.rename(columns={
                                                    "ten_cong_viec": "Công việc",
                                                    "phan_loai": "Đánh giá của AI",
                                                    "nhan_xet": "Nhận xét chi tiết"
                                                })
                                                res_df.insert(0, 'STT', range(1, len(res_df) + 1))
                                            
                                                def color_ph(val):
                                                    if "Khớp" in str(val):
                                                        return 'color: #166534; background-color: #dcfce7; font-weight: bold; border-radius: 4px;'
                                                    else:
                                                        return 'color: #9a3412; background-color: #ffedd5; font-weight: bold; border-radius: 4px;'
                                                    
                                                st.dataframe(res_df.style.map(color_ph, subset=['Đánh giá của AI']), use_container_width=True, hide_index=True)
                                        else:
                                            st.error("Lỗi: AI trả về kết quả không mong muốn. Vui lòng thử lại.")
                                            with st.expander("Dữ liệu thô AI trả về"):
                                                st.write(raw_text)
                                        
                                    except ImportError:
                                        st.error("Chưa cài đặt thư viện `google-generativeai`. Vui lòng chạy `pip install google-generativeai`.")
                                    except Exception as e:
                                        st.error(f"Lỗi hệ thống khi gọi AI: {e}")

elif menu == "⚙️ Quản Lý Cấu Hình":
    st.markdown("### ⚙️ Quản Lý Cấu Hình Hệ Thống")
    
    tab_proj, tab_dept, tab_gsheets = st.tabs(["📁 Quản lý Dự án", "🏢 Quản lý Phòng ban", "📊 Đồng bộ Google Sheets"])
    
    with tab_proj:
        if selected_company == "Tất cả đơn vị":
            st.warning("⚠️ Vui lòng chọn cụ thể một **Công ty / Đơn vị** ở menu bên trái để tiến hành cấu hình (Không áp dụng cho 'Tất cả đơn vị').")
        else:
            st.info(f"Đang cấu hình dữ liệu cho: **{selected_company}**")
            # Load current company's config
            comp_config = config.get("companies", {}).get(selected_company, {})
            comp_projects_by_cat = comp_config.get("projects_by_category", {})
            
            st.markdown(f"#### Quản lý Danh mục Dự án - {selected_company}")
            
            cats = list(comp_projects_by_cat.keys())
            if not cats:
                st.info("Chưa có lĩnh vực dự án nào. Vui lòng cập nhật cấu trúc JSON hoặc thêm mới.")
            else:
                sel_cat = st.selectbox("Chọn Lĩnh vực dự án", cats)
                projs_in_cat = comp_projects_by_cat.get(sel_cat, [])
                
                st.markdown(f"**Danh sách dự án hiện tại trong [{sel_cat}]:**")
                st.write(", ".join(projs_in_cat) if projs_in_cat else "Chưa có dự án nào")
                
                st.markdown("---")
                
                col_add, col_edit, col_del = st.columns(3)
                
                with col_add:
                    st.markdown("**➕ Thêm dự án mới**")
                    new_proj_name = st.text_input("Tên dự án mới", key="admin_add_proj")
                    if st.button("Thêm dự án", type="primary"):
                        if new_proj_name.strip():
                            if new_proj_name.strip() not in projs_in_cat:
                                config["companies"][selected_company]["projects_by_category"][sel_cat].append(new_proj_name.strip())
                                if save_config(config):
                                    st.success(f"Đã thêm dự án: {new_proj_name}")
                                    
                                    st.rerun()
                            else:
                                st.error("Dự án đã tồn tại!")
                        else:
                            st.error("Tên dự án không được để trống!")
                            
                with col_edit:
                    st.markdown("**✏️ Đổi tên dự án**")
                    if projs_in_cat:
                        proj_to_edit = st.selectbox("Chọn dự án cần sửa", projs_in_cat, key="admin_edit_proj_sel")
                        edited_proj_name = st.text_input("Tên dự án mới", value=proj_to_edit, key="admin_edit_proj_val")
                        if st.button("Lưu đổi tên"):
                            if edited_proj_name.strip():
                                idx = config["companies"][selected_company]["projects_by_category"][sel_cat].index(proj_to_edit)
                                config["companies"][selected_company]["projects_by_category"][sel_cat][idx] = edited_proj_name.strip()
                                if save_config(config):
                                    st.success(f"Đã đổi tên thành: {edited_proj_name}")
                                    
                                    st.rerun()
                            else:
                                st.error("Tên mới không được để trống!")
                    else:
                        st.write("Không có dự án để sửa.")
                        
                with col_del:
                    st.markdown("**🗑️ Xóa dự án**")
                    if projs_in_cat:
                        proj_to_del = st.selectbox("Chọn dự án cần xóa", projs_in_cat, key="admin_del_proj_sel")
                        if st.button("Xác nhận xóa dự án", type="secondary"):
                            config["companies"][selected_company]["projects_by_category"][sel_cat].remove(proj_to_del)
                            if save_config(config):
                                st.success(f"Đã xóa dự án: {proj_to_del}")
                                
                                st.rerun()
                    else:
                        st.write("Không có dự án để xóa.")
    
    with tab_dept:
        if selected_company == "Tất cả đơn vị":
            st.warning("⚠️ Vui lòng chọn cụ thể một **Công ty / Đơn vị** ở menu bên trái để tiến hành cấu hình (Không áp dụng cho 'Tất cả đơn vị').")
        else:
            comp_config = config.get("companies", {}).get(selected_company, {})
            comp_depts = comp_config.get("departments", [])
            comp_personnel = comp_config.get("personnel_by_department", {})
            
            st.markdown(f"#### Quản lý Danh sách Phòng ban - {selected_company}")
            st.markdown(f"**Danh sách phòng ban hiện tại ({len(comp_depts)} phòng ban):**")
            st.write(", ".join(comp_depts) if comp_depts else "Chưa có phòng ban nào")
            
            st.markdown("---")
            
            col_d_add, col_d_edit, col_d_del = st.columns(3)
            
            with col_d_add:
                st.markdown("**➕ Thêm phòng ban mới**")
                new_dept_name = st.text_input("Tên phòng ban mới", key="admin_add_dept")
                if st.button("Thêm phòng ban", type="primary"):
                    if new_dept_name.strip():
                        if new_dept_name.strip() not in comp_depts:
                            config["companies"][selected_company]["departments"].append(new_dept_name.strip())
                            if save_config(config):
                                st.success(f"Đã thêm phòng ban: {new_dept_name}")
                                
                                st.rerun()
                        else:
                            st.error("Phòng ban đã tồn tại!")
                    else:
                        st.error("Tên không được để trống!")
                        
            with col_d_edit:
                st.markdown("**✏️ Đổi tên phòng ban**")
                if comp_depts:
                    dept_to_edit = st.selectbox("Chọn phòng ban cần sửa", comp_depts, key="admin_edit_dept_sel")
                    edited_dept_name = st.text_input("Tên phòng ban mới", value=dept_to_edit, key="admin_edit_dept_val")
                    if st.button("Lưu đổi tên phòng"):
                        if edited_dept_name.strip():
                            idx = config["companies"][selected_company]["departments"].index(dept_to_edit)
                            config["companies"][selected_company]["departments"][idx] = edited_dept_name.strip()
                            # Update personnel keys as well
                            if dept_to_edit in config["companies"][selected_company]["personnel_by_department"]:
                                config["companies"][selected_company]["personnel_by_department"][edited_dept_name.strip()] = config["companies"][selected_company]["personnel_by_department"].pop(dept_to_edit, [])
                            if save_config(config):
                                st.success(f"Đã đổi tên thành: {edited_dept_name}")
                                
                                st.rerun()
                        else:
                            st.error("Tên mới không được để trống!")
                else:
                    st.write("Không có phòng ban để sửa.")
                    
            with col_d_del:
                st.markdown("**🗑️ Xóa phòng ban**")
                if comp_depts:
                    dept_to_del = st.selectbox("Chọn phòng ban cần xóa", comp_depts, key="admin_del_dept_sel")
                    if st.button("Xác nhận xóa phòng", type="secondary"):
                        config["companies"][selected_company]["departments"].remove(dept_to_del)
                        # Remove personnel mapping too
                        config["companies"][selected_company]["personnel_by_department"].pop(dept_to_del, None)
                        if save_config(config):
                            st.success(f"Đã xóa phòng ban: {dept_to_del}")
                            
                            st.rerun()
                else:
                    st.write("Không có phòng ban để xóa.")
    
            st.markdown("---")
            st.markdown(f"#### 👥 Quản lý Nhân sự theo Phòng ban - {selected_company}")
            
            if comp_depts:
                sel_dept_p = st.selectbox("Chọn phòng ban để quản lý nhân sự", comp_depts, key="admin_sel_dept_p")
                    
                # Load personnel list
                current_p_list = comp_personnel.get(sel_dept_p, [])
                
                st.markdown(f"**Danh sách nhân sự thuộc [{sel_dept_p}] ({len(current_p_list)} người):**")
                st.write(", ".join(current_p_list) if current_p_list else "Chưa có nhân sự nào")
                
                st.markdown("---")
                
                col_p_add, col_p_edit, col_p_del = st.columns(3)
                
                with col_p_add:
                    st.markdown("**➕ Thêm nhân sự mới**")
                    new_p_name = st.text_input("Tên nhân sự mới", key="admin_add_p_name")
                    if st.button("Thêm nhân sự", type="primary", key="btn_admin_add_p"):
                        if new_p_name.strip():
                            if new_p_name.strip() not in current_p_list:
                                if sel_dept_p not in config["companies"][selected_company]["personnel_by_department"]:
                                    config["companies"][selected_company]["personnel_by_department"][sel_dept_p] = []
                                config["companies"][selected_company]["personnel_by_department"][sel_dept_p].append(new_p_name.strip())
                                if save_config(config):
                                    st.success(f"Đã thêm nhân sự: {new_p_name.strip()}")
                                    
                                    st.rerun()
                            else:
                                st.error("Nhân sự đã tồn tại trong phòng ban này!")
                        else:
                            st.error("Tên nhân sự không được để trống!")
                            
                with col_p_edit:
                    st.markdown("**✏️ Sửa tên nhân sự**")
                    if current_p_list:
                        p_to_edit = st.selectbox("Chọn nhân sự cần sửa", current_p_list, key="admin_edit_p_sel")
                        edited_p_name = st.text_input("Tên nhân sự mới", value=p_to_edit, key="admin_edit_p_val")
                        if st.button("Lưu thay đổi", key="btn_admin_edit_p"):
                            if edited_p_name.strip():
                                if edited_p_name.strip() not in current_p_list or edited_p_name.strip() == p_to_edit:
                                    idx = current_p_list.index(p_to_edit)
                                    config["companies"][selected_company]["personnel_by_department"][sel_dept_p][idx] = edited_p_name.strip()
                                    if save_config(config):
                                        st.success(f"Đã cập nhật tên nhân sự thành: {edited_p_name.strip()}")
                                        
                                        st.rerun()
                                else:
                                    st.error("Tên mới đã tồn tại trong phòng ban này!")
                            else:
                                st.error("Tên mới không được để trống!")
                    else:
                        st.write("Không có nhân sự để sửa.")
                        
                with col_p_del:
                    st.markdown("**🗑️ Xóa nhân sự**")
                    if current_p_list:
                        p_to_del = st.selectbox("Chọn nhân sự cần xóa", current_p_list, key="admin_del_p_sel")
                        if st.button("Xác nhận xóa", type="secondary", key="btn_admin_del_p"):
                            config["companies"][selected_company]["personnel_by_department"][sel_dept_p].remove(p_to_del)
                            if save_config(config):
                                st.success(f"Đã xóa nhân sự: {p_to_del}")
                                
                                st.rerun()
                    else:
                        st.write("Không có nhân sự để xóa.")
            else:
                st.warning("Vui lòng tạo ít nhất một phòng ban trước khi cấu hình nhân sự.")

    with tab_gsheets:
        st.markdown("#### 📊 Cấu hình kết nối Google Sheets")
        if is_gsheets_configured():
            st.success("🎉 Hệ thống đã kết nối thành công với Google Sheets! Mọi thay đổi dữ liệu sẽ được tự động đồng bộ thời gian thực.")
            try:
                st.info(f"**Spreadsheet URL:** `{st.secrets['connections']['gsheets']['spreadsheet']}`")
            except Exception:
                pass
        else:
            st.warning("⚠️ Hiện tại hệ thống đang hoạt động ở chế độ ngoại tuyến (Offline) bằng file Excel cục bộ.")
            
        st.markdown("""
        ### 📝 Hướng dẫn kết nối Google Sheets trên Streamlit Cloud
        Để đồng bộ dữ liệu trực tuyến, vui lòng thực hiện các bước sau:
        
        1. **Tạo Google Sheet**: Tạo một bảng tính Google Sheets mới. Đảm bảo sheet đầu tiên có tên là `Sheet1`.
        2. **Định cấu hình cột mẫu**: Bạn có thể tải file Excel hiện tại từ sidebar xuống và copy cấu trúc cột sang Google Sheets. Các cột gồm:
           `ID`, `DonVi`, `PhongBan`, `NguoiChuTri`, `TenDuAn`, `MocTienDo`, `SanPhamBanGiao`, `TenCongViec`, `PhanLoaiChiSo`, `NgayBatDau`, `Deadline`, `DoUuTien`, `PhanTramHoanThanh`, `TrangThai`, `LinkKetQua`, `GiaiTrinhDeXuat`, `NgayCapNhat`
        3. **Chia sẻ quyền chỉnh sửa**: Chia sẻ Google Sheet đó với tài khoản **Google Service Account** của bạn (cấp quyền **Editor**).
        4. **Cấu hình Secrets trên Streamlit Cloud**:
           - Truy cập Dashboard của Streamlit Cloud -> Vào ứng dụng -> Chọn **Settings** -> **Secrets**.
           - Nhập thông tin cấu hình Service Account theo định dạng sau:
           ```toml
           [connections.gsheets]
           spreadsheet = "https://docs.google.com/spreadsheets/d/your-spreadsheet-id"
           type = "service_account"
           project_id = "your-project-id"
           private_key_id = "your-private-key-id"
           private_key = "-----BEGIN PRIVATE KEY-----\\nyour-private-key-details\\n-----END PRIVATE KEY-----\\n"
           client_email = "your-service-account-email@your-project.iam.gserviceaccount.com"
           client_id = "your-client-id"
           auth_uri = "https://accounts.google.com/o/oauth2/auth"
           token_uri = "https://oauth2.googleapis.com/token"
           auth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"
           client_x509_cert_url = "https://www.googleapis.com/workspace/3pid/cert"
           ```
        5. **Khởi động lại (Reboot) app**: Lưu lại Secrets, Streamlit sẽ tự động đồng bộ và nạp dữ liệu từ Google Sheets.
        """)

# ----------------- 6. SỔ TAY HƯỚNG DẪN -----------------
elif menu == "📖 Sổ tay Hướng dẫn":
    st.markdown("## 📖 Sổ tay Hướng dẫn sử dụng phần mềm KPI")
    st.markdown("Chọn vai trò của bạn để xem hướng dẫn chi tiết:")
    
    tab_nv, tab_ql, tab_tc = st.tabs(["👨‍💼 Hướng dẫn Nhân viên", "👔 Hướng dẫn Quản lý", "🌟 Tiêu chí Đánh giá & Xếp loại"])
    
    with tab_nv:
        st.info("""
        **1️⃣ Đăng ký công việc (Đầu tháng)**
        - 🕒 **Thời gian:** Từ ngày 30 tháng trước đến ngày 3 tháng này.
        - 🖱️ **Thao tác:** Vào mục **Thêm / Cập nhật công việc**.
        - 📝 **Nội dung:** Tự khai báo các đầu việc chính trong tháng. Hệ thống sẽ tự động tính toán và chia đều tỷ trọng KPI cho tất cả các công việc của bạn.
        """)
        
        st.success("""
        **2️⃣ Báo cáo tiến độ và Hoàn thành**
        - 🖱️ Khi thực hiện xong công việc, vào mục **Thêm / Cập nhật công việc**, đánh dấu tick vào ô **☑️ Công việc đã hoàn thành**.
        - 📌 **Lưu ý quan trọng:** Bạn cần dán kèm Link minh chứng kết quả (từ Google Drive, OneDrive...) hoặc ghi tên/số hiệu văn bản vào ô khai báo kết quả.
        """)
        
        st.warning("""
        **3️⃣ Xử lý Trễ hạn / Có vướng mắc**
        - Nếu rủi ro trễ hạn, đổi trạng thái thành **Có vướng mắc** và ghi rõ lý do tại ô *Giải trình / Đề xuất*.
        - 🌍 **Do khách quan**: Hệ thống gửi yêu cầu để Quản lý xem xét lý do và đánh giá lại mức điểm (50%, 80%, 90%...).
        - 🌧️ **Do chủ quan**: Công việc bị tính là chưa hoàn thành và nhận 0 điểm KPI.
        """)
        
        st.error("""
        **4️⃣ Theo dõi công việc hàng ngày**
        - 🖱️ Vào mục **Bảng theo dõi tiến độ công việc**.
        - Xem thẻ **CÔNG VIỆC TỚI HẠN** để biết việc nào sắp đến hạn (màu vàng) hoặc đã trễ hạn (màu đỏ) để ưu tiên xử lý.
        """)
        
    with tab_ql:
        st.success("""
        **1️⃣ Xem xét và Đánh giá lý do Khách quan**
        - 🖱️ Vào mục **✅ Duyệt việc Khách quan**.
        - Hệ thống liệt kê các công việc nhân viên báo cáo trễ hạn với lý do **Khách quan**.
        - Bạn xem xét giải trình, click trực tiếp vào ô *Mức độ KPI ghi nhận* để chọn điểm phù hợp (Miễn trừ, 50%, 80%, 90%...).
        - Nhấn **💾 Lưu toàn bộ phê duyệt** ở cuối danh sách.
        """)
        
        st.info("""
        **2️⃣ Xem Báo cáo Xếp loại KPI (Ngày 1-3 đầu tháng)**
        - 🕒 **Thời gian:** Từ ngày 1 đến ngày 3 hàng tháng.
        - 🎯 **Mục đích:** Xác nhận điểm số KPI của tháng trước để Phòng HCNS lưu kết quả.
        - Xem biểu đồ tổng quan và Bảng dữ liệu tự động Xếp loại cho từng nhân sự.
        """)

    with tab_tc:
        st.info("""
        **🌟 TIÊU CHÍ ĐÁNH GIÁ VÀ XẾP LOẠI KPI**
        
        🧮 **1. Công thức tính điểm KPI Tổng:**
        > :blue[**Điểm KPI**] = (:green[**Điểm trung bình công việc Định kỳ**] × **70%**) + (:orange[**Điểm trung bình công việc Giao ban**] × **30%**) + :red[**Điểm thưởng/phạt**]
        
        *(Lưu ý: Nếu không có công việc Giao ban, hệ thống sẽ tự động điều chỉnh 100% trọng số cho công việc Định kỳ).*
        
        📊 **2. Phân loại và Quy đổi Điểm Xếp loại:**
        - Tổng điểm **> 91**: Xếp loại **A** (Xuất sắc)
        - Tổng điểm **> 81**: Xếp loại **B** (Tốt)
        - Tổng điểm **> 71**: Xếp loại **C** (Khá)
        - Tổng điểm **<= 71**: Xếp loại **D** (Kém)
        
        ⚖️ **3. Về Điểm Thưởng / Phạt:**
        - **Cộng điểm (+):** Áp dụng cho các công việc hoàn thành xuất sắc vượt tiến độ, hoặc có sáng kiến mang lại hiệu quả cao.
        - **Trừ điểm (-):** Áp dụng khi vi phạm nội quy, chậm trễ báo cáo, hoặc có sai sót nghiệp vụ gây ảnh hưởng.
        - *Quản lý trực tiếp hoặc HCNS sẽ rà soát và cập nhật quỹ điểm Thưởng/Phạt này trước thời điểm chốt sổ cuối tháng.*
        """)

